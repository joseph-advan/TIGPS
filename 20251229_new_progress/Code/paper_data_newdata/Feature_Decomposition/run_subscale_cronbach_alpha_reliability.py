from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_binary_drop_then_split_baseline as fd

SCRIPT_PATH = Path(__file__).resolve()
OUT_DIR = SCRIPT_PATH.parent
CONFIG_PATH = OUT_DIR / "subscale_definitions_w2_w3.json"
RELIABILITY_OUT_DIR = OUT_DIR / "outputs" / "reliability"
OUT_XLSX = RELIABILITY_OUT_DIR / "subscale_cronbach_alpha_reliability.xlsx"
OUT_MD = RELIABILITY_OUT_DIR / "subscale_cronbach_alpha_reliability_summary.md"
OUT_JSON = RELIABILITY_OUT_DIR / "subscale_cronbach_alpha_reliability_details.json"


def min_valid_count(n_items: int) -> int:
    return max(1, math.ceil(n_items * 0.5))


def numeric_items(df: pd.DataFrame, items: list[str]) -> tuple[pd.DataFrame, list[str], list[str]]:
    found, missing = fd.resolve_existing_items(df, items)
    if not found:
        return pd.DataFrame(index=df.index), found, missing
    data = df[found].apply(pd.to_numeric, errors="coerce")
    return data, found, missing


def cronbach_alpha(data: pd.DataFrame) -> tuple[float, int]:
    complete = data.dropna(axis=0, how="any")
    n_complete = int(len(complete))
    k = int(complete.shape[1])
    if k < 2 or n_complete < 2:
        return np.nan, n_complete
    item_var = complete.var(axis=0, ddof=1)
    total_var = complete.sum(axis=1).var(ddof=1)
    if pd.isna(total_var) or total_var <= 0:
        return np.nan, n_complete
    alpha = k / (k - 1) * (1 - item_var.sum() / total_var)
    return float(alpha), n_complete


def standardized_alpha(data: pd.DataFrame) -> float:
    complete = data.dropna(axis=0, how="any")
    k = int(complete.shape[1])
    if k < 2 or len(complete) < 2:
        return np.nan
    corr = complete.corr()
    vals = corr.where(~np.eye(k, dtype=bool)).stack().dropna()
    if vals.empty:
        return np.nan
    rbar = float(vals.mean())
    denom = 1 + (k - 1) * rbar
    if denom == 0:
        return np.nan
    return float(k * rbar / denom)


def corrected_item_total_corr(complete: pd.DataFrame, item: str) -> float:
    if item not in complete.columns or complete.shape[1] < 2 or len(complete) < 2:
        return np.nan
    other_total = complete.drop(columns=[item]).sum(axis=1)
    if complete[item].nunique(dropna=True) < 2 or other_total.nunique(dropna=True) < 2:
        return np.nan
    return float(complete[item].corr(other_total))


def alpha_rating(alpha: float, n_items: int) -> str:
    if n_items < 2:
        return "Not applicable: single-item subscale"
    if pd.isna(alpha):
        return "Not estimable"
    if alpha < 0:
        return "Problematic: negative alpha; check item direction or dimensionality"
    if alpha >= 0.80:
        return "Good internal consistency"
    if alpha >= 0.70:
        return "Acceptable internal consistency"
    if alpha >= 0.60:
        return "Questionable but usable for exploratory work"
    return "Low internal consistency; review subscale"


def alpha_flag(alpha: float, n_items: int, min_item_total: float) -> str:
    flags: list[str] = []
    if n_items < 2:
        flags.append("single_item_no_alpha")
    elif pd.isna(alpha):
        flags.append("alpha_not_estimable")
    elif alpha < 0:
        flags.append("negative_alpha_check_reverse_or_multidimensional")
    elif alpha < 0.60:
        flags.append("alpha_below_0.60_review")
    elif alpha < 0.70:
        flags.append("alpha_0.60_to_0.69_questionable")
    if n_items >= 2 and not pd.isna(min_item_total):
        if min_item_total < 0:
            flags.append("negative_item_total_correlation")
        elif min_item_total < 0.20:
            flags.append("item_total_below_0.20")
    return "; ".join(flags) if flags else "ok"


def build_item_diagnostics(
    *,
    wave: str,
    parent_group: str,
    formal_en: str,
    subscale_code: str,
    subscale_en: str,
    data: pd.DataFrame,
) -> list[dict[str, Any]]:
    complete = data.dropna(axis=0, how="any")
    rows: list[dict[str, Any]] = []
    for item in data.columns:
        alpha_drop = np.nan
        if data.shape[1] > 2:
            alpha_drop, _ = cronbach_alpha(data.drop(columns=[item]))
        rows.append(
            {
                "Wave": wave,
                "Parent Group": parent_group,
                "Parent Scale English Name": formal_en,
                "Subscale Code": subscale_code,
                "Subscale English Name": subscale_en,
                "Item": item,
                "N non-missing item": int(data[item].notna().sum()),
                "Corrected Item-Total Correlation": corrected_item_total_corr(complete, item),
                "Alpha if Item Deleted": alpha_drop,
            }
        )
    return rows


def evaluate_scale(
    *,
    wave: str,
    parent_group: str,
    formal_zh: str,
    formal_en: str,
    scale_code: str,
    scale_zh: str,
    scale_en: str,
    items: list[str],
    df: pd.DataFrame,
    scale_type: str,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    data, found, missing = numeric_items(df, items)
    n_defined = len(items)
    n_found = len(found)
    valid_required = min_valid_count(n_found) if n_found else np.nan
    n_valid_50 = int((data.notna().sum(axis=1) >= valid_required).sum()) if n_found else 0
    alpha, n_complete = cronbach_alpha(data)
    std_alpha = standardized_alpha(data)
    item_rows = build_item_diagnostics(
        wave=wave,
        parent_group=parent_group,
        formal_en=formal_en,
        subscale_code=scale_code,
        subscale_en=scale_en,
        data=data,
    ) if n_found else []
    item_corrs = [r["Corrected Item-Total Correlation"] for r in item_rows if not pd.isna(r["Corrected Item-Total Correlation"])]
    min_item_total = min(item_corrs) if item_corrs else np.nan
    mean_item_total = float(np.mean(item_corrs)) if item_corrs else np.nan
    row = {
        "Wave": wave,
        "Scale Type": scale_type,
        "Parent Group": parent_group,
        "Parent Scale Chinese Name": formal_zh,
        "Parent Scale English Name": formal_en,
        "Scale Code": scale_code,
        "Scale Chinese Name": scale_zh,
        "Scale English Name": scale_en,
        "Defined Item Count": n_defined,
        "Found Item Count": n_found,
        "Missing Items": "; ".join(missing),
        "Used Items": "; ".join(found),
        "Minimum Valid Items for 50% Rule": valid_required,
        "N rows": int(len(df)),
        "N complete cases for alpha": n_complete,
        "N meeting >=50% valid items": n_valid_50,
        "Cronbach alpha": alpha,
        "Standardized alpha": std_alpha,
        "Mean corrected item-total correlation": mean_item_total,
        "Minimum corrected item-total correlation": min_item_total,
        "Interpretation": alpha_rating(alpha, n_found),
        "Review Flag": alpha_flag(alpha, n_found, min_item_total),
    }
    return row, item_rows


def load_data() -> dict[str, pd.DataFrame]:
    return {
        "W2": fd.normalize_student_id(pd.read_csv(fd.W2_DATA_PATH, encoding="utf-8-sig", low_memory=False)),
    }


def collect_reliability() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    config = fd.load_subscale_config(CONFIG_PATH)
    data_by_wave = load_data()
    scale_rows: list[dict[str, Any]] = []
    parent_rows: list[dict[str, Any]] = []
    item_rows: list[dict[str, Any]] = []

    for wave in ["W2"]:
        df = data_by_wave[wave]
        for parent_group, info in config["waves"].get(wave, {}).items():
            formal_zh = info.get("formal_group_name_zh", "")
            formal_en = info.get("formal_group_name_en", "")
            parent_items: list[str] = []
            for subscale_code, subscale in info.get("subscales", {}).items():
                sub_items = list(subscale.get("items", []))
                parent_items.extend(sub_items)
                row, items = evaluate_scale(
                    wave=wave,
                    parent_group=parent_group,
                    formal_zh=formal_zh,
                    formal_en=formal_en,
                    scale_code=subscale_code,
                    scale_zh=subscale.get("subscale_name_zh", ""),
                    scale_en=subscale.get("subscale_name_en", ""),
                    items=sub_items,
                    df=df,
                    scale_type="Configured subscale",
                )
                scale_rows.append(row)
                item_rows.extend(items)
            # Parent alpha is a diagnostic only. It is not the modeling feature when the group is configured for splitting.
            parent_items = list(dict.fromkeys(parent_items))
            parent_row, _ = evaluate_scale(
                wave=wave,
                parent_group=parent_group,
                formal_zh=formal_zh,
                formal_en=formal_en,
                scale_code=parent_group,
                scale_zh=formal_zh,
                scale_en=formal_en,
                items=parent_items,
                df=df,
                scale_type="Full parent scale diagnostic",
            )
            parent_rows.append(parent_row)

    subscale_df = pd.DataFrame(scale_rows)
    parent_df = pd.DataFrame(parent_rows)
    item_df = pd.DataFrame(item_rows)
    review_df = subscale_df.loc[subscale_df["Review Flag"].ne("ok")].copy()
    review_df = review_df.sort_values(["Wave", "Parent Group", "Scale Code"])
    return subscale_df, parent_df, item_df, review_df


def build_w2_compact_summary(subscale_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Parent Group",
        "Parent Scale English Name",
        "Scale Code",
        "Scale English Name",
        "Defined Item Count",
        "Cronbach alpha",
        "Minimum corrected item-total correlation",
        "Interpretation",
        "Review Flag",
    ]
    return subscale_df[cols].sort_values(["Parent Group", "Scale Code"]).reset_index(drop=True)


def write_xlsx(sheets: dict[str, pd.DataFrame]) -> None:
    RELIABILITY_OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    wb = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = max(len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
    wb.save(OUT_XLSX)


def write_markdown(subscale_df: pd.DataFrame, parent_df: pd.DataFrame, review_df: pd.DataFrame) -> None:
    total = len(subscale_df)
    single = int(subscale_df["Defined Item Count"].eq(1).sum())
    estimable = subscale_df[subscale_df["Defined Item Count"].ge(2)].copy()
    good = int(estimable["Cronbach alpha"].ge(0.80).sum())
    acceptable = int(estimable["Cronbach alpha"].between(0.70, 0.799999, inclusive="both").sum())
    questionable = int(estimable["Cronbach alpha"].between(0.60, 0.699999, inclusive="both").sum())
    low = int(estimable["Cronbach alpha"].lt(0.60).sum())

    lines = [
        "# Subscale Cronbach's Alpha Reliability Check",
        "",
        "## Purpose",
        "",
        "This reliability check extends the Feature_Decomposition workflow. The current main-paper plan uses W2 predictors only for W2->W2 and W2->W3 tasks, so Cronbach's alpha is calculated only for W2 configured subscales.",
        "",
        "## Data Used",
        "",
        f"- W2 cleaned data: `{fd.W2_DATA_PATH}`",
        f"- Subscale config: `{CONFIG_PATH}`",
        "",
        "## Scoring and Reliability Rules",
        "",
        "- Cronbach's alpha is calculated on complete cases for the items within each subscale.",
        "- `N meeting >=50% valid items` is also reported because the modeling pipeline uses multi-item scale scores; this count shows how many students have enough valid item responses for a 50% valid-item rule.",
        "- Single-item subscales cannot have Cronbach's alpha and are marked as not applicable.",
        "- Low alpha does not automatically invalidate a subscale, but it means the item grouping should be reviewed against theory and item wording.",
        "- W3 subscale reliability is intentionally not calculated in this output because W3 questionnaire features are not used as predictors in the current two-task design.",
        "",
        "## Overall Summary",
        "",
        f"- Total configured subscales checked: `{total}`",
        f"- Single-item subscales where alpha is not applicable: `{single}`",
        f"- Good alpha >= 0.80: `{good}`",
        f"- Acceptable alpha 0.70-0.79: `{acceptable}`",
        f"- Questionable alpha 0.60-0.69: `{questionable}`",
        f"- Low alpha < 0.60: `{low}`",
        "",
        "## Review Flags",
        "",
    ]
    if review_df.empty:
        lines.append("No subscale was flagged for review.")
    else:
        cols = ["Wave", "Parent Group", "Scale Code", "Scale English Name", "Defined Item Count", "Cronbach alpha", "Minimum corrected item-total correlation", "Review Flag"]
        lines.append(review_df[cols].to_markdown(index=False, floatfmt=".3f"))
    lines += [
        "",
        "## How to Use This Result",
        "",
        "Use this as supporting evidence for the decomposition strategy. The prediction comparison shows whether splitting improves model performance; Cronbach's alpha shows whether each proposed subscale is internally coherent. If a theoretically important subscale has low alpha, it can still be retained, but the limitation should be noted and the item wording should be reviewed.",
        "",
        "## Output Files",
        "",
        f"- `{OUT_XLSX.name}`",
        f"- `{OUT_JSON.name}`",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    subscale_df, parent_df, item_df, review_df = collect_reliability()
    compact_df = build_w2_compact_summary(subscale_df)
    readme = pd.DataFrame(
        [
            {"Sheet": "W2_SubscaleAlpha", "Description": "Cronbach's alpha for each configured W2 subscale."},
            {"Sheet": "W2_ParentScaleAlpha", "Description": "Diagnostic alpha for each full original W2 parent item group before splitting."},
            {"Sheet": "W2_ItemDiagnostics", "Description": "Corrected item-total correlation and alpha-if-item-deleted diagnostics for W2 items."},
            {"Sheet": "W2_ReviewFlags", "Description": "W2 subscales flagged for low alpha, single-item alpha not applicable, or weak item-total correlations."},
            {"Sheet": "W2_CompactSummary", "Description": "Compact W2-only summary for quick review."},
        ]
    )
    write_xlsx(
        {
            "ReadMe": readme,
            "W2_SubscaleAlpha": subscale_df,
            "W2_ParentScaleAlpha": parent_df,
            "W2_ItemDiagnostics": item_df,
            "W2_ReviewFlags": review_df,
            "W2_CompactSummary": compact_df,
        }
    )
    details = {
        "subscale_alpha": subscale_df.replace({np.nan: None}).to_dict(orient="records"),
        "parent_scale_alpha": parent_df.replace({np.nan: None}).to_dict(orient="records"),
        "review_flags": review_df.replace({np.nan: None}).to_dict(orient="records"),
    }
    OUT_JSON.write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(subscale_df, parent_df, review_df)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
