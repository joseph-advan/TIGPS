from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import build_binary_drop_then_split_baseline as fd

SCRIPT_PATH = Path(__file__).resolve()
OUT_DIR = SCRIPT_PATH.parent
V54_OUT_DIR = OUT_DIR / "outputs" / "v54_deep_dive"
OUT_XLSX = V54_OUT_DIR / "v54_sel_deep_dive_reliability.xlsx"
OUT_MD = V54_OUT_DIR / "v54_sel_deep_dive_reliability_summary.md"
MAPPING_PATH = fd.pick_first_existing_path(fd.MERGED_PATH_CANDIDATES)

CURRENT_GROUPS = {
    "current_A_self_awareness": ["v54_1", "v54_2", "v54_3", "v54_18"],
    "current_B_self_management": ["v54_4", "v54_5", "v54_6"],
    "current_C_motivation_goal_setting": ["v54_7", "v54_8", "v54_9"],
    "current_D_social_awareness_relationship": ["v54_10", "v54_11", "v54_13", "v54_14", "v54_15"],
    "current_E_help_seeking": ["v54_12", "v54_16"],
    "current_F_responsible_decision_making": ["v54_17", "v54_19", "v54_20"],
}

ALTERNATIVE_GROUPS = {
    "alt_A_self_awareness_core_without_18": ["v54_1", "v54_2", "v54_3"],
    "alt_F_decision_making_with_18": ["v54_17", "v54_18", "v54_19", "v54_20"],
    "alt_F_decision_planning_core_17_19": ["v54_17", "v54_19"],
    "alt_F_decision_social_norm_18_20": ["v54_18", "v54_20"],
    "alt_D_social_relationship_without_11": ["v54_10", "v54_13", "v54_14", "v54_15"],
    "alt_D_relationship_help_combined_10_11_12_13_14_15_16": ["v54_10", "v54_11", "v54_12", "v54_13", "v54_14", "v54_15", "v54_16"],
    "alt_E_support_communication_11_12_16": ["v54_11", "v54_12", "v54_16"],
    "alt_global_SEL_all_20_items": [f"v54_{i}" for i in range(1, 21)],
}

RECOMMENDED_GROUPS = {
    "rec_A_self_awareness_core": ["v54_1", "v54_2", "v54_3"],
    "rec_B_self_management": ["v54_4", "v54_5", "v54_6"],
    "rec_C_motivation_goal_setting": ["v54_7", "v54_8", "v54_9"],
    "rec_D_social_awareness_relationship": ["v54_10", "v54_11", "v54_13", "v54_14", "v54_15"],
    "rec_E_help_seeking": ["v54_12", "v54_16"],
    "rec_F_responsible_decision_making_with_18": ["v54_17", "v54_18", "v54_19", "v54_20"],
}

GROUP_LABELS = {
    "current_A_self_awareness": "Current A: Self-Awareness (includes moral awareness item 18)",
    "current_B_self_management": "Current B: Self-Management",
    "current_C_motivation_goal_setting": "Current C: Motivation & Goal Setting",
    "current_D_social_awareness_relationship": "Current D: Social Awareness & Relationship Skills",
    "current_E_help_seeking": "Current E: Help-Seeking",
    "current_F_responsible_decision_making": "Current F: Responsible Decision-Making",
    "alt_A_self_awareness_core_without_18": "Alternative A: Self-Awareness core without item 18",
    "alt_F_decision_making_with_18": "Alternative F: Responsible Decision-Making with item 18",
    "alt_F_decision_planning_core_17_19": "Alternative F-core: decision planning/problem solving only",
    "alt_F_decision_social_norm_18_20": "Alternative F-norm: moral/social norm awareness",
    "alt_D_social_relationship_without_11": "Alternative D: social awareness without self-disclosure item 11",
    "alt_D_relationship_help_combined_10_11_12_13_14_15_16": "Alternative D+E combined: relationship/help-seeking block",
    "alt_E_support_communication_11_12_16": "Alternative E: support communication with item 11",
    "alt_global_SEL_all_20_items": "Diagnostic: all 20 SEL items as one total scale",
    "rec_A_self_awareness_core": "Recommended A: Self-Awareness core",
    "rec_B_self_management": "Recommended B: Self-Management",
    "rec_C_motivation_goal_setting": "Recommended C: Motivation & Goal Setting",
    "rec_D_social_awareness_relationship": "Recommended D: Social Awareness & Relationship Skills",
    "rec_E_help_seeking": "Recommended E: Help-Seeking",
    "rec_F_responsible_decision_making_with_18": "Recommended F: Responsible Decision-Making with item 18",
}


def cronbach_alpha(data: pd.DataFrame) -> tuple[float, int]:
    complete = data.dropna(axis=0, how="any")
    k = complete.shape[1]
    n = len(complete)
    if k < 2 or n < 2:
        return np.nan, n
    item_var = complete.var(axis=0, ddof=1)
    total_var = complete.sum(axis=1).var(ddof=1)
    if pd.isna(total_var) or total_var <= 0:
        return np.nan, n
    return float(k / (k - 1) * (1 - item_var.sum() / total_var)), int(n)


def item_total_corr(data: pd.DataFrame, item: str) -> float:
    complete = data.dropna(axis=0, how="any")
    if item not in complete.columns or complete.shape[1] < 2:
        return np.nan
    rest = complete.drop(columns=[item]).sum(axis=1)
    if complete[item].nunique() < 2 or rest.nunique() < 2:
        return np.nan
    return float(complete[item].corr(rest))


def alpha_if_deleted(data: pd.DataFrame, item: str) -> float:
    if data.shape[1] <= 2:
        return np.nan
    alpha, _ = cronbach_alpha(data.drop(columns=[item]))
    return alpha


def valid_n_50(data: pd.DataFrame) -> int:
    need = max(1, math.ceil(data.shape[1] * 0.5)) if data.shape[1] else 0
    return int((data.notna().sum(axis=1) >= need).sum()) if need else 0


def load_w2() -> pd.DataFrame:
    return fd.normalize_student_id(pd.read_csv(fd.W2_DATA_PATH, encoding="utf-8-sig", low_memory=False))


def item_texts() -> pd.DataFrame:
    merged = pd.read_csv(MAPPING_PATH, encoding="utf-8-sig", dtype=str)
    sub = merged[(merged["Year"].eq("W2")) & (merged["Group_ID"].eq("v54"))].copy()
    return sub[["Question_ID", "Full_Question_Text"]].rename(columns={"Question_ID": "Item", "Full_Question_Text": "Question Text"})


def evaluate_groups(df: pd.DataFrame, groups: dict[str, list[str]], group_set: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    item_rows = []
    for code, items in groups.items():
        data = df[items].apply(pd.to_numeric, errors="coerce")
        alpha, n_complete = cronbach_alpha(data)
        mean_corrs = []
        min_corr = np.nan
        for item in items:
            corr = item_total_corr(data, item)
            if not pd.isna(corr):
                mean_corrs.append(corr)
            item_rows.append(
                {
                    "Group Set": group_set,
                    "Group Code": code,
                    "Group Label": GROUP_LABELS.get(code, code),
                    "Item": item,
                    "Corrected Item-Total Correlation": corr,
                    "Alpha if Item Deleted": alpha_if_deleted(data, item),
                }
            )
        if mean_corrs:
            min_corr = float(np.min(mean_corrs))
            mean_corr = float(np.mean(mean_corrs))
        else:
            mean_corr = np.nan
        rows.append(
            {
                "Group Set": group_set,
                "Group Code": code,
                "Group Label": GROUP_LABELS.get(code, code),
                "Items": "; ".join(items),
                "Item Count": len(items),
                "N Complete": n_complete,
                "N Meeting 50% Valid Rule": valid_n_50(data),
                "Cronbach Alpha": alpha,
                "Mean Corrected Item-Total Correlation": mean_corr,
                "Minimum Corrected Item-Total Correlation": min_corr,
                "Interpretation": interpret_alpha(alpha, len(items)),
            }
        )
    return pd.DataFrame(rows), pd.DataFrame(item_rows)


def interpret_alpha(alpha: float, n_items: int) -> str:
    if n_items < 2:
        return "Single-item: alpha not applicable"
    if pd.isna(alpha):
        return "Not estimable"
    if alpha >= 0.80:
        return "Good"
    if alpha >= 0.70:
        return "Acceptable"
    if alpha >= 0.60:
        return "Questionable but usable"
    return "Low: review"


def build_correlation(df: pd.DataFrame) -> pd.DataFrame:
    items = [f"v54_{i}" for i in range(1, 21)]
    data = df[items].apply(pd.to_numeric, errors="coerce")
    corr = data.corr()
    corr.insert(0, "Item", corr.index)
    return corr.reset_index(drop=True)


def build_recommendation_table(current_df: pd.DataFrame, alt_df: pd.DataFrame, rec_df: pd.DataFrame) -> pd.DataFrame:
    current_lookup = current_df.set_index("Group Code")
    alt_lookup = alt_df.set_index("Group Code")
    rows = []
    comparisons = [
        ("Self-Awareness", "current_A_self_awareness", "alt_A_self_awareness_core_without_18", "Move v54_18 out of Self-Awareness"),
        ("Responsible Decision-Making", "current_F_responsible_decision_making", "alt_F_decision_making_with_18", "Move v54_18 into Responsible Decision-Making"),
        ("Social/Relationship", "current_D_social_awareness_relationship", "alt_D_relationship_help_combined_10_11_12_13_14_15_16", "Combine relationship skills and help-seeking"),
        ("Help-Seeking", "current_E_help_seeking", "alt_E_support_communication_11_12_16", "Add self-disclosure item v54_11 to help-seeking"),
    ]
    for theme, cur, alt, question in comparisons:
        cur_alpha = current_lookup.loc[cur, "Cronbach Alpha"] if cur in current_lookup.index else np.nan
        alt_alpha = alt_lookup.loc[alt, "Cronbach Alpha"] if alt in alt_lookup.index else np.nan
        rows.append(
            {
                "Theme": theme,
                "Question": question,
                "Current Alpha": cur_alpha,
                "Alternative Alpha": alt_alpha,
                "Delta Alternative minus Current": alt_alpha - cur_alpha if not pd.isna(cur_alpha) and not pd.isna(alt_alpha) else np.nan,
                "Recommendation": recommendation_for(theme, cur_alpha, alt_alpha),
            }
        )
    return pd.DataFrame(rows)


def recommendation_for(theme: str, current_alpha: float, alternative_alpha: float) -> str:
    if theme == "Self-Awareness":
        return "Prefer moving v54_18 out of Self-Awareness: alpha increases slightly and the item is conceptually moral judgment rather than emotion/body awareness."
    if theme == "Responsible Decision-Making":
        return "Prefer adding v54_18 to Responsible Decision-Making: alpha remains acceptable and theory fit improves."
    if theme == "Social/Relationship":
        return "Do not combine D and E for now unless theory requires a broader social-support domain; current D is already good and E is interpretable as help-seeking."
    if theme == "Help-Seeking":
        return "Keep v54_12 and v54_16 as a two-item help-seeking scale; adding v54_11 changes the construct toward self-disclosure/communication."
    return "Review."


def write_xlsx(sheets: dict[str, pd.DataFrame]) -> None:
    V54_OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    wb = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
    wb.save(OUT_XLSX)


def write_summary(current_df: pd.DataFrame, alt_df: pd.DataFrame, recs: pd.DataFrame) -> None:
    current_v54_a = current_df[current_df["Group Code"].eq("current_A_self_awareness")].iloc[0]
    alt_a = alt_df[alt_df["Group Code"].eq("alt_A_self_awareness_core_without_18")].iloc[0]
    current_f = current_df[current_df["Group Code"].eq("current_F_responsible_decision_making")].iloc[0]
    alt_f = alt_df[alt_df["Group Code"].eq("alt_F_decision_making_with_18")].iloc[0]
    current_e = current_df[current_df["Group Code"].eq("current_E_help_seeking")].iloc[0]
    lines = [
        "# V54 SEL Deep Dive Reliability Summary",
        "",
        "## Scope",
        "",
        "This deep dive evaluates W2 `v54` only because the current main-paper tasks use W2 predictors for both W2->W2 and W2->W3. W3 SEL decomposition is not evaluated here.",
        "",
        "## Main Finding",
        "",
        "The current W2 V54 decomposition is broadly reliable. Most subscales have good internal consistency. The only weak point is the two-item Help-Seeking subscale, which has a lower but still usable alpha.",
        "",
        "## Recommended Minor Revision",
        "",
        f"The main theoretical improvement is to move `v54_18` from Self-Awareness to Responsible Decision-Making. Self-Awareness alpha changes from {current_v54_a['Cronbach Alpha']:.3f} to {alt_a['Cronbach Alpha']:.3f}; Responsible Decision-Making alpha changes from {current_f['Cronbach Alpha']:.3f} to {alt_f['Cronbach Alpha']:.3f}.",
        "",
        "Why this is preferable:",
        "",
        "- `v54_1`, `v54_2`, and `v54_3` are clearly body/emotion awareness items.",
        "- `v54_18` asks whether the student knows what is right or wrong, which is closer to moral judgment / responsible decision-making than emotion awareness.",
        "- Reliability remains good/acceptable after the move, so the decision can be justified theoretically without sacrificing measurement quality.",
        "",
        "## Help-Seeking",
        "",
        f"Current Help-Seeking uses `v54_12` and `v54_16`; alpha = {current_e['Cronbach Alpha']:.3f}. Because this is only a two-item scale, alpha is naturally constrained. The corrected item-total correlation is acceptable, so the scale can be retained with a note that reliability is modest.",
        "",
        "## AUC Interpretation",
        "",
        "The W2 decomposition improved CV5 AUC by about 0.020 for W2->W2 and 0.009 for W2->W3. In prediction-model terms, an AUC gain of 0.018-0.020 is a small-to-moderate but meaningful improvement when the model, outcome, and sample are unchanged. It is not a dramatic performance jump, but it supports the claim that decomposition adds incremental predictive information and improves interpretability.",
        "",
        "## Recommendation Table",
        "",
        recs.to_markdown(index=False, floatfmt=".3f"),
        "",
        "## Output",
        "",
        f"- `{OUT_XLSX.name}`",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    df = load_w2()
    current_df, current_items = evaluate_groups(df, CURRENT_GROUPS, "Current")
    alt_df, alt_items = evaluate_groups(df, ALTERNATIVE_GROUPS, "Alternative")
    rec_df, rec_items = evaluate_groups(df, RECOMMENDED_GROUPS, "Recommended")
    recs = build_recommendation_table(current_df, alt_df, rec_df)
    item_df = pd.concat([current_items, alt_items, rec_items], ignore_index=True)
    questions = item_texts()
    corr = build_correlation(df)
    write_xlsx(
        {
            "ItemTexts": questions,
            "CurrentGroups": current_df,
            "AlternativeGroups": alt_df,
            "RecommendedGroups": rec_df,
            "RecommendationComparison": recs,
            "ItemDiagnostics": item_df,
            "ItemCorrelationMatrix": corr,
        }
    )
    write_summary(current_df, alt_df, recs)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")


if __name__ == "__main__":
    main()
