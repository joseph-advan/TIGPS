from __future__ import annotations

import importlib.util
import json
import shutil
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
PAPER_DATA_DIR = PAPER_RESULTS_DIR.parent
ROOT = PAPER_DATA_DIR.parents[1]

SOURCE_TABLE1_SCRIPT = PAPER_DATA_DIR / "tables" / "table1" / "scripts" / "build_table1_drop_decomposition.py"

OUT_DIR = SECTION_DIR / "outputs"
TASK_W2W2_DIR = OUT_DIR / "01_w2_features_to_w2_distress"
TASK_W2W3_DIR = OUT_DIR / "02_w2_features_to_w3_distress"
DIAG_DIR = OUT_DIR / "diagnostics"

PAPER_READY_XLSX = OUT_DIR / "table1_prediction_aligned_group_differences.xlsx"
SUMMARY_MD = OUT_DIR / "TABLE1_PREDICTION_ALIGNED_GROUP_DIFFERENCES_SUMMARY.md"
DIAGNOSTICS_JSON = DIAG_DIR / "table1_prediction_aligned_group_differences_diagnostics.json"

TASKS = [
    {
        "task": "W2 -> W2",
        "slug": "w2_to_w2",
        "outcome_wave": "W2",
        "folder": TASK_W2W2_DIR,
        "description": "W2 baseline features compared by W2 psychological distress group.",
    },
    {
        "task": "W2 -> W3",
        "slug": "w2_to_w3",
        "outcome_wave": "W3",
        "folder": TASK_W2W3_DIR,
        "description": "W2 baseline features compared by W3 psychological distress group.",
    },
]

NETWORK_SPECS = [
    {
        "network_version": "Observed network",
        "slug": "observed_network",
        "spec_attr": "INTERPERSONAL_TABLE1_FEATURES",
        "label": "observed",
        "description": "Raw interpersonal nomination counts, ratios, and valence indicators.",
    },
]

INTERPERSONAL_KEYWORDS = (
    "Nomination",
    "Nominations",
    "Tie",
    "Ties",
    "Network Valence",
    "Positive Tie Ratio",
)

NETWORK_CONCEPT_ORDER = [
    "Online Total Nominations",
    "Offline Total Nominations",
    "Outgoing Friendship Nominations",
    "Incoming Friendship Nominations",
    "Outgoing Negative Nominations",
    "Incoming Negative Nominations",
    "Reciprocal Friendship Ties",
    "Reciprocal Negative Ties",
    "Sent Positive Tie Ratio",
    "Received Positive Tie Ratio",
    "Sent Network Valence",
    "Received Network Valence",
]

NETWORK_VERSION_ORDER = {
    "Observed network": 1,
}


def import_source_table1_module():
    if not SOURCE_TABLE1_SCRIPT.exists():
        raise FileNotFoundError(f"Missing source Table 1 script: {SOURCE_TABLE1_SCRIPT}")
    spec = importlib.util.spec_from_file_location("source_table1_drop_decomposition", SOURCE_TABLE1_SCRIPT)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to import source script: {SOURCE_TABLE1_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def reset_outputs() -> None:
    section_resolved = SECTION_DIR.resolve()
    out_resolved = OUT_DIR.resolve()
    if out_resolved.parent != section_resolved:
        raise RuntimeError(f"Refusing to delete unexpected output path: {out_resolved}")
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    for path in [TASK_W2W2_DIR, TASK_W2W3_DIR, DIAG_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def load_inputs(src) -> dict[str, Any]:
    config = json.loads(src.CONFIG_PATH.read_text(encoding="utf-8"))
    merged_path = src.t23.core.pick_first_existing_path(src.t23.core.MERGED_PATH_CANDIDATES)
    merged = pd.read_csv(merged_path, dtype=str, encoding="utf-8-sig")
    for col in ["Year", "Group_ID", "Question_ID"]:
        if col in merged.columns:
            merged[col] = merged[col].astype(str).str.strip()

    w2 = src.t23.core.normalize_student_id(pd.read_csv(src.W2_DATA, encoding="utf-8-sig", low_memory=False))
    w3 = src.t23.core.normalize_student_id(pd.read_csv(src.W3_DATA, encoding="utf-8-sig", low_memory=False))
    return {"config": config, "merged": merged, "W2": w2, "W3": w3, "merged_path": merged_path}


def make_distress_group(src, df: pd.DataFrame, wave: str) -> tuple[pd.Series, dict[str, Any]]:
    items = src.t23.TARGET_W2_ITEMS if wave == "W2" else src.t23.TARGET_W3_ITEMS
    group, diag = src.classify_psychological_distress(df, items)
    return group, diag


def map_outcome_group_to_w2_rows(w2: pd.DataFrame, outcome_df: pd.DataFrame, outcome_group: pd.Series) -> pd.Series:
    lookup = (
        pd.DataFrame({"student_id": outcome_df["student_id"].astype(str), "outcome_group": outcome_group})
        .dropna(subset=["student_id"])
        .drop_duplicates(subset=["student_id"], keep="first")
    )
    mapped = w2[["student_id"]].astype(str).merge(lookup, on="student_id", how="left")["outcome_group"]
    mapped.index = w2.index
    return mapped.astype("object")


def write_xlsx(df: pd.DataFrame, path: Path, sheet_name: str = "Table1") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    autosize_and_style_xlsx(path)


def parse_p(value: Any) -> float | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("<"):
        return 0.0005
    parsed = pd.to_numeric(text, errors="coerce")
    return float(parsed) if pd.notna(parsed) else None


def top_effect_rows(df: pd.DataFrame, n: int = 8) -> pd.DataFrame:
    if "Between-group difference" not in df.columns:
        return pd.DataFrame()
    work = df.copy()
    work["effect_numeric"] = pd.to_numeric(work["Between-group difference"], errors="coerce")
    work = work[work["effect_numeric"].notna()].copy()
    work["abs_effect"] = work["effect_numeric"].abs()
    cols = [
        "Question ID",
        "Variable",
        "High Psychological Distress",
        "Low Psychological Distress",
        "p-value",
        "Between-group difference",
        "Between-group difference type",
    ]
    return work.sort_values("abs_effect", ascending=False).head(n)[cols]


def interpersonal_rows(df: pd.DataFrame) -> pd.DataFrame:
    if "Variable" not in df.columns:
        return pd.DataFrame()
    mask = df["Variable"].astype(str).apply(lambda x: any(keyword in x for keyword in INTERPERSONAL_KEYWORDS))
    cols = [
        "Question ID",
        "Variable",
        "High Psychological Distress",
        "Low Psychological Distress",
        "p-value",
        "Between-group difference",
        "Between-group difference type",
    ]
    return df.loc[mask, [c for c in cols if c in df.columns]].copy()


def summarize_interpersonal(df: pd.DataFrame) -> dict[str, Any]:
    rows = interpersonal_rows(df)
    if rows.empty:
        return {"n_interpersonal_rows": 0, "n_p_lt_01": 0, "median_abs_effect": None, "max_abs_effect": None}
    p_values = rows["p-value"].apply(parse_p) if "p-value" in rows.columns else pd.Series(dtype=float)
    effects = pd.to_numeric(rows.get("Between-group difference"), errors="coerce").abs()
    return {
        "n_interpersonal_rows": int(len(rows)),
        "n_p_lt_01": int(p_values.dropna().lt(0.01).sum()),
        "median_abs_effect": float(effects.median()) if effects.notna().any() else None,
        "max_abs_effect": float(effects.max()) if effects.notna().any() else None,
    }


def n_row(table: pd.DataFrame) -> dict[str, Any]:
    row = table[table["Variable"].astype(str).eq("N")]
    if row.empty:
        return {"high_n": None, "low_n": None, "total_n": None}
    r = row.iloc[0]
    return {
        "high_n": int(pd.to_numeric(r.get("High Psychological Distress"), errors="coerce")),
        "low_n": int(pd.to_numeric(r.get("Low Psychological Distress"), errors="coerce")),
        "total_n": int(pd.to_numeric(r.get("Total"), errors="coerce")),
    }


def build_tables(src, inputs: dict[str, Any]) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, list[dict[str, Any]]]:
    w2 = inputs["W2"]
    w3 = inputs["W3"]
    merged = inputs["merged"]
    config = inputs["config"]

    w2_group, w2_group_diag = make_distress_group(src, w2, "W2")
    w3_group, w3_group_diag = make_distress_group(src, w3, "W3")
    groups = {
        "W2": w2_group,
        "W3": map_outcome_group_to_w2_rows(w2, w3, w3_group),
    }

    tables: dict[str, pd.DataFrame] = {}
    diag_rows: list[dict[str, Any]] = []
    output_rows: list[dict[str, Any]] = []

    for task in TASKS:
        task["folder"].mkdir(parents=True, exist_ok=True)
        outcome_wave = task["outcome_wave"]
        group = groups[outcome_wave]
        for network in NETWORK_SPECS:
            specs = getattr(src, network["spec_attr"])
            X, feature_defs, feature_diag = src.build_feature_context(
                "W2",
                w2,
                merged,
                interpersonal_specs=specs,
                interpersonal_label=network["label"],
            )
            table = src.build_table(
                "W2",
                X,
                feature_defs,
                group,
                (src.GROUP_HIGH_DISTRESS, src.GROUP_LOW_DISTRESS),
                config,
            )

            table_key = f"{task['slug']}_{network['slug']}"
            out_path = task["folder"] / f"table1_{table_key}.xlsx"
            write_xlsx(table, out_path)
            tables[table_key] = table

            n_info = n_row(table)
            diag_rows.append(
                {
                    "Task": task["task"],
                    "Predictor Wave": "W2",
                    "Outcome Wave": outcome_wave,
                    "Network Version": network["network_version"],
                    "Rows in table": len(table),
                    **n_info,
                    "Feature context rows": len(X),
                    "Feature count": len(feature_defs),
                    "Outcome missing after mapping": int(group.isna().sum()),
                    "W2 target positive": w2_group_diag.get("target_positive"),
                    "W2 target negative": w2_group_diag.get("target_negative"),
                    "W3 target positive": w3_group_diag.get("target_positive"),
                    "W3 target negative": w3_group_diag.get("target_negative"),
                    **{f"feature_diag_{k}": v for k, v in feature_diag.items()},
                }
            )
            output_rows.append(
                {
                    "Task": task["task"],
                    "Network Version": network["network_version"],
                    "Description": task["description"],
                    "Output Path": str(out_path),
                }
            )

    output_index = pd.DataFrame(output_rows)
    return tables, output_index, diag_rows


def network_concept(variable: Any) -> str:
    text = str(variable)
    for concept in NETWORK_CONCEPT_ORDER:
        if concept in text:
            return concept
    return text.replace(", Observed Count", "").replace(", Observed", "").replace(", Respondent-Class-Normalized", "").replace(", mean (SD)", "")


def build_network_comparison(tables: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for task in TASKS:
        for network in NETWORK_SPECS:
            key = f"{task['slug']}_{network['slug']}"
            net = interpersonal_rows(tables[key])
            for _, row in net.iterrows():
                rows.append(
                    {
                        "Task": task["task"],
                        "Feature Concept": network_concept(row.get("Variable")),
                        "Network Version": network["network_version"],
                        "Variable": row.get("Variable"),
                        "High Psychological Distress": row.get("High Psychological Distress"),
                        "Low Psychological Distress": row.get("Low Psychological Distress"),
                        "p-value": row.get("p-value"),
                        "Between-group difference": row.get("Between-group difference"),
                        "Between-group difference type": row.get("Between-group difference type"),
                    }
                )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    concept_order = {concept: idx for idx, concept in enumerate(NETWORK_CONCEPT_ORDER, start=1)}
    out["_task_order"] = out["Task"].map({task["task"]: idx for idx, task in enumerate(TASKS, start=1)})
    out["_concept_order"] = out["Feature Concept"].map(concept_order).fillna(999)
    out["_version_order"] = out["Network Version"].map(NETWORK_VERSION_ORDER).fillna(999)
    out = out.sort_values(["_task_order", "_concept_order", "_version_order", "Variable"]).drop(
        columns=["_task_order", "_concept_order", "_version_order"]
    )
    return out.reset_index(drop=True)


def autosize_and_style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 70))
            ws.column_dimensions[get_column_letter(idx)].width = max(12, max_len + 2)
    wb.save(path)


def write_paper_ready_workbook(
    tables: dict[str, pd.DataFrame],
    output_index: pd.DataFrame,
    diagnostics: pd.DataFrame,
) -> dict[str, Any]:
    readme = pd.DataFrame(
        [
            {
                "Section": "Main design",
                "Description": "Prediction-aligned Table 1 uses W2 baseline features only, grouped by W2 or W3 psychological distress outcomes.",
            },
            {
                "Section": "Task 1",
                "Description": "W2 -> W2: W2 baseline features compared by W2 high vs low psychological distress.",
            },
            {
                "Section": "Task 2",
                "Description": "W2 -> W3: W2 baseline features compared by W3 high vs low psychological distress.",
            },
            {
                "Section": "Network versions",
                "Description": "Each task uses observed, non-class-adjusted interpersonal nomination features.",
            },
            {
                "Section": "Excluded from main output",
                "Description": "W3 features -> W3 distress tables are intentionally excluded from this main paper-results section.",
            },
            {
                "Section": "Effect-size column",
                "Description": "Between-group difference is Cramer's V for categorical rows and Cohen's d for continuous/scale rows.",
            },
        ]
    )
    network_comparison = build_network_comparison(tables)

    with pd.ExcelWriter(PAPER_READY_XLSX, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="ReadMe", index=False)
        tables["w2_to_w2_observed_network"].to_excel(writer, sheet_name="W2toW2_Observed", index=False)
        tables["w2_to_w3_observed_network"].to_excel(writer, sheet_name="W2toW3_Observed", index=False)
        network_comparison.to_excel(writer, sheet_name="NetworkComparison", index=False)
        diagnostics.to_excel(writer, sheet_name="Diagnostics", index=False)
        output_index.to_excel(writer, sheet_name="OutputIndex", index=False)

    autosize_and_style_xlsx(PAPER_READY_XLSX)

    summary: dict[str, Any] = {}
    for task in TASKS:
        for network in NETWORK_SPECS:
            key = f"{task['slug']}_{network['slug']}"
            summary[key] = {
                "shape": list(tables[key].shape),
                "n": n_row(tables[key]),
                "top_effects": top_effect_rows(tables[key]).to_dict(orient="records"),
                "interpersonal": summarize_interpersonal(tables[key]),
            }
    return summary


def md_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_No rows available._"
    return df.to_markdown(index=False)


def format_interpersonal_summary(summary: dict[str, Any], key: str) -> str:
    info = summary[key]["interpersonal"]
    if info["median_abs_effect"] is None:
        return f"| {key} | 0 | 0 |  |  |"
    return (
        f"| {key} | {info['n_interpersonal_rows']} | {info['n_p_lt_01']} | "
        f"{info['median_abs_effect']:.3f} | {info['max_abs_effect']:.3f} |"
    )


def write_summary(workbook_summary: dict[str, Any], output_index: pd.DataFrame) -> None:
    lines = [
        "# Prediction-Aligned Table 1 Group Differences Summary",
        "",
        "## Main decision",
        "",
        "This section now aligns Table 1 with the prediction tasks used in model performance. The predictor side is always W2 baseline features.",
        "",
        "Included tasks:",
        "",
        "- W2 -> W2: W2 baseline features grouped by W2 high vs low psychological distress.",
        "- W2 -> W3: W2 baseline features grouped by W3 high vs low psychological distress.",
        "",
        "Excluded from the main output:",
        "",
        "- W3 features -> W3 distress. This is intentionally excluded because it does not match the baseline-prediction logic.",
        "",
        "## Network specifications",
        "",
        "Each prediction task is produced once with observed interpersonal features:",
        "",
        "- Observed network: raw interpersonal nomination counts, ratios, and valence features.",
        "",
        "## Main workbook",
        "",
        f"- `{PAPER_READY_XLSX}`",
        "",
        "## Output files",
        "",
    ]
    for _, row in output_index.iterrows():
        lines.append(f"- {row['Task']} / {row['Network Version']}: `{row['Output Path']}`")

    lines.extend(
        [
            "",
            "## Largest descriptive differences by table",
            "",
        ]
    )
    for task in TASKS:
        for network in NETWORK_SPECS:
            key = f"{task['slug']}_{network['slug']}"
            title = f"{task['task']} - {network['network_version']}"
            lines.extend(["", f"### {title}", ""])
            lines.append(md_table(pd.DataFrame(workbook_summary[key]["top_effects"])))

    lines.extend(
        [
            "",
            "## Interpersonal feature signal",
            "",
            "This table focuses on the stricter `p < .01` threshold. Because the sample size is large, statistical significance alone is not enough; the effect-size columns should be read at the same time.",
            "",
            "| Table key | Interpersonal rows | Rows with p < .01 | Median absolute effect size | Max absolute effect size |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for task in TASKS:
        for network in NETWORK_SPECS:
            key = f"{task['slug']}_{network['slug']}"
            lines.append(format_interpersonal_summary(workbook_summary, key))

    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "Use these tables as descriptive screening evidence before the model-based feature-importance and interaction sections. The W2 -> W3 table is especially important for the longitudinal story because it asks whether baseline W2 characteristics already distinguish students who later fall into the W3 high-distress group.",
            "",
            "For the interpersonal features, the main question is not only whether p-values pass the threshold. The more important pattern is whether the observed differences are large enough to matter. If a feature is significant but has a small absolute effect size, it should be described as a detectable descriptive difference rather than a strong substantive difference.",
            "",
            "The `NetworkComparison` sheet is arranged by interpersonal feature concept using the observed, non-class-adjusted version only.",
            "",
        ]
    )
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def write_diagnostics(inputs: dict[str, Any], diagnostics: pd.DataFrame, workbook_summary: dict[str, Any]) -> None:
    payload = {
        "section_dir": str(SECTION_DIR),
        "source_table1_script": str(SOURCE_TABLE1_SCRIPT),
        "merged_question_map": str(inputs["merged_path"]),
        "w2_data": str(import_source_table1_module().W2_DATA),
        "w3_data": str(import_source_table1_module().W3_DATA),
        "diagnostics": diagnostics.to_dict(orient="records"),
        "workbook_summary": workbook_summary,
    }
    DIAGNOSTICS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    reset_outputs()
    src = import_source_table1_module()
    inputs = load_inputs(src)
    tables, output_index, diag_rows = build_tables(src, inputs)
    diagnostics = pd.DataFrame(diag_rows)
    workbook_summary = write_paper_ready_workbook(tables, output_index, diagnostics)
    write_summary(workbook_summary, output_index)
    write_diagnostics(inputs, diagnostics, workbook_summary)

    print(f"Wrote {PAPER_READY_XLSX}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Wrote {DIAGNOSTICS_JSON}")


if __name__ == "__main__":
    main()
