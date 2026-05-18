from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
RESULT_DIR = SCRIPT_PATH.parent
OUTPUT_DIR = RESULT_DIR / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CODE_DIR = SCRIPT_PATH.parents[2]
TABLES_SCRIPT_DIR = CODE_DIR / "tables" / "scripts"
TABLE1_SCRIPT_DIR = CODE_DIR / "tables" / "table1" / "scripts"
for p in [TABLES_SCRIPT_DIR, TABLE1_SCRIPT_DIR]:
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

import build_table2_table3_drop_decomposition as t23  # noqa: E402
import build_table1_drop_decomposition as t1  # noqa: E402

TASKS = [
    {
        "task": "W2 -> W2",
        "scenario": "w2_predict_w2",
        "feature_wave": "W2",
        "target_wave": "W2",
        "feature_target_group": "v55",
        "target_items": t23.TARGET_W2_ITEMS,
    },
    {
        "task": "W2 -> W3",
        "scenario": "w2_predict_w3",
        "feature_wave": "W2",
        "target_wave": "W3",
        "feature_target_group": "54",
        "target_items": t23.TARGET_W3_ITEMS,
    },
]

FEATURE_SETS = [
    {
        "feature_set": "decomposed_features_only",
        "feature_set_label": "Decomposed features only",
        "interpersonal": False,
    },
    {
        "feature_set": "decomposed_plus_12_interpersonal",
        "feature_set_label": "Decomposed + 12 observed interpersonal features",
        "interpersonal": True,
    },
]

INTERPERSONAL_SPECS = t1.INTERPERSONAL_TABLE1_FEATURES
INTERPERSONAL_COLUMNS = [spec["column"] for spec in INTERPERSONAL_SPECS]
INTERPERSONAL_NAMES = {spec["column"]: spec["name"] for spec in INTERPERSONAL_SPECS}
INTERPERSONAL_ITEMS = {spec["column"]: spec["items"] for spec in INTERPERSONAL_SPECS}
INTERPERSONAL_VERSION = "observed_count"


def read_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    merged_path = t23.core.pick_first_existing_path(t23.core.MERGED_PATH_CANDIDATES)
    merged = pd.read_csv(merged_path, dtype=str, encoding="utf-8-sig")
    for col in ["Year", "Group_ID", "Question_ID"]:
        if col in merged.columns:
            merged[col] = merged[col].astype(str).str.strip()
    w2 = t23.core.normalize_student_id(pd.read_csv(t23.W2_DATA, encoding="utf-8-sig", low_memory=False))
    w3 = t23.core.normalize_student_id(pd.read_csv(t23.W3_DATA, encoding="utf-8-sig", low_memory=False))
    return merged, w2, w3


def make_target_for_task(task: dict[str, Any], datasets: dict[str, pd.DataFrame], feature_df: pd.DataFrame) -> tuple[pd.Series, dict[str, Any]]:
    target_df = datasets[task["target_wave"]]
    _, y_raw, target_diag = t23.make_target(target_df, task["target_items"])
    target_map = pd.DataFrame({"student_id": target_df["student_id"], "target": y_raw}).drop_duplicates("student_id", keep="first")
    y = feature_df[["student_id"]].merge(target_map, on="student_id", how="left")["target"]
    y.index = feature_df.index
    target_diag = {
        **target_diag,
        "task": task["task"],
        "target_wave": task["target_wave"],
        "aligned_target_non_missing": int(y.notna().sum()),
        "aligned_target_positive": int(y.eq(1).sum()),
        "aligned_target_negative": int(y.eq(0).sum()),
    }
    return y, target_diag


def build_feature_set(task: dict[str, Any], feature_set: dict[str, Any], merged: pd.DataFrame, datasets: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    feature_df = datasets[task["feature_wave"]]
    X, defs, diag = t23.build_drop_decomposition_features(
        feature_df,
        merged,
        task["feature_wave"],
        task["feature_target_group"],
    )
    diag = {
        **diag,
        "task": task["task"],
        "scenario": task["scenario"],
        "feature_set": feature_set["feature_set"],
        "feature_set_label": feature_set["feature_set_label"],
    }
    if feature_set["interpersonal"]:
        X, defs, diag = t1.add_interpersonal_features(
            task["feature_wave"],
            feature_df,
            X,
            defs,
            diag,
            INTERPERSONAL_SPECS,
            INTERPERSONAL_VERSION,
        )
    return X, defs, diag


def to_numeric_value(value: Any) -> float:
    if value is None:
        return np.nan
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return np.nan
    return float(pd.to_numeric(value, errors="coerce"))


def numeric_performance(perf: pd.DataFrame) -> pd.DataFrame:
    out = perf.copy()
    for col in out.columns:
        if col == "Model":
            continue
        converted = pd.to_numeric(out[col], errors="coerce")
        if converted.notna().any():
            out[col] = converted
    return out


def add_metadata_to_performance(perf: pd.DataFrame, task: dict[str, Any], feature_set: dict[str, Any]) -> pd.DataFrame:
    out = perf.copy()
    out.insert(0, "Task", task["task"])
    out.insert(1, "Scenario", task["scenario"])
    out.insert(2, "Feature Set", feature_set["feature_set"])
    out.insert(3, "Feature Set Label", feature_set["feature_set_label"])
    out.insert(4, "Feature Wave", task["feature_wave"])
    out.insert(5, "Target Wave", task["target_wave"])
    return out


def long_coefficients(coef_table: pd.DataFrame, task: dict[str, Any], feature_set: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    model_cols = {
        "Multivariable Logistic": "Multivariable Logistic Std. B",
        "LASSO Logistic": "LASSO Logistic Std. B",
        "Ridge Logistic": "Ridge Logistic Std. B",
    }
    for _, row in coef_table.iterrows():
        model_column = str(row["Model Column"])
        feature_code = str(row["Feature Code"])
        is_ip = row.get("Source Type") == "interpersonal_feature" or feature_code in INTERPERSONAL_COLUMNS
        for model, col in model_cols.items():
            if col not in coef_table.columns:
                continue
            b = to_numeric_value(row[col])
            rows.append(
                {
                    "Task": task["task"],
                    "Scenario": task["scenario"],
                    "Feature Set": feature_set["feature_set"],
                    "Feature Set Label": feature_set["feature_set_label"],
                    "Model": model,
                    "Variable": row["Variable"],
                    "Feature Code": feature_code,
                    "Model Column": model_column,
                    "Source Type": row.get("Source Type", ""),
                    "Items": row.get("Items", ""),
                    "Std. B": b,
                    "Abs Std. B": abs(b) if pd.notna(b) else np.nan,
                    "Selected by LASSO": bool(row.get("Selected by LASSO", False)),
                    "Is Interpersonal Feature": bool(is_ip),
                }
            )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["Relative Importance %"] = np.nan
    out["Rank by Abs Std. B"] = np.nan
    for keys, idx in out.groupby(["Task", "Feature Set", "Model"]).groups.items():
        sub_idx = list(idx)
        denom = out.loc[sub_idx, "Abs Std. B"].sum(skipna=True)
        if denom and denom > 0:
            out.loc[sub_idx, "Relative Importance %"] = out.loc[sub_idx, "Abs Std. B"] / denom * 100.0
        ranks = out.loc[sub_idx, "Abs Std. B"].rank(method="min", ascending=False)
        out.loc[sub_idx, "Rank by Abs Std. B"] = ranks
    out["Top 20 by Abs Std. B"] = out["Rank by Abs Std. B"].le(20)
    return out


def performance_delta(perf_all: pd.DataFrame) -> pd.DataFrame:
    metrics = [
        "CV auc mean",
        "CV accuracy mean",
        "CV balanced_accuracy mean",
        "CV precision mean",
        "CV recall mean",
        "CV f1 mean",
        "Test AUC",
        "Test Accuracy",
        "Test Balanced Accuracy",
        "Test Precision",
        "Test Recall",
        "Test F1",
    ]
    rows: list[dict[str, Any]] = []
    for (task, model), sub in perf_all.groupby(["Task", "Model"]):
        base = sub[sub["Feature Set"].eq("decomposed_features_only")]
        plus = sub[sub["Feature Set"].eq("decomposed_plus_12_interpersonal")]
        if base.empty or plus.empty:
            continue
        b = base.iloc[0]
        p = plus.iloc[0]
        row = {"Task": task, "Model": model}
        row["N features baseline"] = b.get("N features")
        row["N features plus interpersonal"] = p.get("N features")
        for metric in metrics:
            if metric in perf_all.columns:
                bval = to_numeric_value(b.get(metric))
                pval = to_numeric_value(p.get(metric))
                row[f"Baseline {metric}"] = bval
                row[f"Plus interpersonal {metric}"] = pval
                row[f"Delta {metric}"] = pval - bval if pd.notna(pval) and pd.notna(bval) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def interpersonal_selection_summary(coef_long: pd.DataFrame) -> pd.DataFrame:
    sub = coef_long[coef_long["Feature Set"].eq("decomposed_plus_12_interpersonal") & coef_long["Is Interpersonal Feature"]].copy()
    if sub.empty:
        return sub
    sub = sub.sort_values(["Task", "Model", "Rank by Abs Std. B", "Feature Code"])
    cols = [
        "Task",
        "Model",
        "Variable",
        "Feature Code",
        "Std. B",
        "Relative Importance %",
        "Rank by Abs Std. B",
        "Top 20 by Abs Std. B",
        "Selected by LASSO",
        "Items",
    ]
    return sub[[c for c in cols if c in sub.columns]]


def selection_counts(coef_long: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    sub = coef_long[coef_long["Feature Set"].eq("decomposed_plus_12_interpersonal")].copy()
    for (task, model), g in sub.groupby(["Task", "Model"]):
        ip = g[g["Is Interpersonal Feature"]]
        rows.append(
            {
                "Task": task,
                "Model": model,
                "N interpersonal features": int(len(ip)),
                "Interpersonal features in Top 20": int(ip["Top 20 by Abs Std. B"].sum()),
                "Interpersonal relative importance sum %": float(ip["Relative Importance %"].sum(skipna=True)),
                "Max interpersonal relative importance %": float(ip["Relative Importance %"].max(skipna=True)) if len(ip) else np.nan,
                "Best-ranked interpersonal feature": ip.sort_values("Rank by Abs Std. B").iloc[0]["Variable"] if len(ip) else "",
                "Best interpersonal rank": float(ip["Rank by Abs Std. B"].min(skipna=True)) if len(ip) else np.nan,
                "LASSO-selected interpersonal features": int(ip["Selected by LASSO"].sum()) if model == "LASSO Logistic" else "",
            }
        )
    return pd.DataFrame(rows)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            clean_sheet = sheet[:31]
            df.to_excel(writer, index=False, sheet_name=clean_sheet)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def fmt(value: Any, digits: int = 3) -> str:
    value = pd.to_numeric(value, errors="coerce")
    if pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def build_summary_md(perf_delta_df: pd.DataFrame, counts_df: pd.DataFrame, diag_df: pd.DataFrame) -> str:
    lines: list[str] = []
    lines.append("# Interpersonal Incremental Modeling Summary")
    lines.append("")
    lines.append("## Purpose")
    lines.append("")
    lines.append("This analysis tests whether adding the 12 respondent-class-normalized interpersonal network indicators improves prediction beyond the current drop + decomposition individual-level feature set.")
    lines.append("")
    lines.append("## Tasks and Models")
    lines.append("")
    lines.append("- Tasks: W2 -> W2 and W2 -> W3.")
    lines.append("- Feature sets: decomposed features only vs decomposed + 12 interpersonal features.")
    lines.append("- Models: plain multivariable Logistic, LASSO Logistic, and Ridge Logistic.")
    lines.append("- Main comparison metrics: CV5 AUC, CV5 F1, CV5 accuracy, and holdout test AUC/F1.")
    lines.append("")
    lines.append("## Interpersonal Features Added")
    lines.append("")
    for idx, spec in enumerate(INTERPERSONAL_SPECS, start=1):
        lines.append(f"{idx}. {spec['name']} (`{spec['column']}`): {spec['items']}.")
    lines.append("")
    lines.append("## Performance Delta: Plus Interpersonal Minus Baseline")
    lines.append("")
    show_cols = [
        "Task",
        "Model",
        "Delta CV auc mean",
        "Delta CV f1 mean",
        "Delta CV accuracy mean",
        "Delta Test AUC",
        "Delta Test F1",
        "N features baseline",
        "N features plus interpersonal",
    ]
    if not perf_delta_df.empty:
        table = perf_delta_df[show_cols].copy()
        for col in table.columns:
            if col.startswith("Delta"):
                table[col] = table[col].map(lambda v: fmt(v, 4))
        lines.append(table.to_markdown(index=False))
    lines.append("")
    lines.append("## Interpersonal Feature Selection Summary")
    lines.append("")
    if not counts_df.empty:
        show = counts_df.copy()
        for col in ["Interpersonal relative importance sum %", "Max interpersonal relative importance %", "Best interpersonal rank"]:
            if col in show.columns:
                show[col] = show[col].map(lambda v: fmt(v, 2))
        lines.append(show.to_markdown(index=False))
    lines.append("")
    lines.append("## Interpretation Guide")
    lines.append("")
    lines.append("Use this analysis as the bridge between Table 1 and the later LASSO Top 20 feature-importance section. Table 1 only shows group differences. This incremental model tests whether interpersonal indicators add predictive value after individual-level features are included.")
    lines.append("")
    lines.append("A strong argument for limited interpersonal incremental value would require: small performance deltas after adding the 12 features, few interpersonal features in the LASSO Top 20, and low Ridge relative importance for interpersonal indicators.")
    lines.append("")
    lines.append("## Diagnostics")
    lines.append("")
    if not diag_df.empty:
        diag_for_show = diag_df.rename(
            columns={
                "task": "Task",
                "feature_set": "Feature Set",
            }
        )
        wanted = [
            "Task",
            "Feature Set",
            "n_final_features",
            "table1_interpersonal_status",
            "table1_interpersonal_features_added",
            "aligned_target_non_missing",
        ]
        diag_show = diag_for_show[[c for c in wanted if c in diag_for_show.columns]].copy()
        lines.append(diag_show.to_markdown(index=False))
    lines.append("")
    return "\n".join(lines) + "\n"


def main() -> None:
    merged, w2, w3 = read_inputs()
    datasets = {"W2": w2, "W3": w3}

    perf_frames: list[pd.DataFrame] = []
    coef_frames: list[pd.DataFrame] = []
    diagnostics: list[dict[str, Any]] = []

    for task in TASKS:
        feature_df = datasets[task["feature_wave"]]
        y, target_diag = make_target_for_task(task, datasets, feature_df)
        for feature_set in FEATURE_SETS:
            X, defs, feature_diag = build_feature_set(task, feature_set, merged, datasets)
            coef_table, perf = t23.fit_model_comparison(y, X, defs)
            perf = numeric_performance(perf)
            perf_frames.append(add_metadata_to_performance(perf, task, feature_set))
            coef_frames.append(long_coefficients(coef_table, task, feature_set))
            diagnostics.append({**feature_diag, **target_diag, "n_final_features": int(X.shape[1])})

    perf_all = pd.concat(perf_frames, ignore_index=True)
    coef_long = pd.concat(coef_frames, ignore_index=True)
    diag_df = pd.DataFrame(diagnostics)
    delta_df = performance_delta(perf_all)
    selection_df = interpersonal_selection_summary(coef_long)
    counts_df = selection_counts(coef_long)

    performance_xlsx = OUTPUT_DIR / "interpersonal_incremental_model_performance.xlsx"
    selection_xlsx = OUTPUT_DIR / "interpersonal_feature_selection_summary.xlsx"
    diagnostics_json = OUTPUT_DIR / "interpersonal_incremental_modeling_diagnostics.json"
    summary_md = OUTPUT_DIR / "INTERPERSONAL_INCREMENTAL_MODELING_SUMMARY.md"

    write_workbook(
        performance_xlsx,
        {
            "Performance": perf_all,
            "PerformanceDelta": delta_df,
            "SelectionCounts": counts_df,
            "Diagnostics": diag_df,
        },
    )
    write_workbook(
        selection_xlsx,
        {
            "InterpersonalFeatures": selection_df,
            "AllCoefficientsLong": coef_long,
            "SelectionCounts": counts_df,
        },
    )
    diagnostics_json.write_text(
        json.dumps(
            {
                "feature_set_definition": {
                    "baseline": "drop + decomposition features only",
                    "incremental": "drop + decomposition plus 12 respondent-class-normalized interpersonal features",
                    "interpersonal_features": INTERPERSONAL_SPECS,
                },
                "diagnostics": diagnostics,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8-sig",
    )
    summary_md.write_text(build_summary_md(delta_df, counts_df, diag_df), encoding="utf-8")

    print("Wrote", performance_xlsx)
    print("Wrote", selection_xlsx)
    print("Wrote", diagnostics_json)
    print("Wrote", summary_md)
    print("\nPerformance delta:")
    print(delta_df.to_string(index=False))
    print("\nInterpersonal selection counts:")
    print(counts_df.to_string(index=False))


if __name__ == "__main__":
    main()
