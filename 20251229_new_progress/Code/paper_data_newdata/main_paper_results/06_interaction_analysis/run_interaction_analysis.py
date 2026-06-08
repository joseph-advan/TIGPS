from __future__ import annotations

import json
import math
import re
import shutil
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import accuracy_score, roc_auc_score


SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
ROOT = SCRIPT_PATH.parents[4]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"
TABLES_SCRIPT_DIR = CODE_DIR / "tables" / "scripts"
TABLE1_SCRIPT_DIR = CODE_DIR / "tables" / "table1" / "scripts"
for script_dir in [TABLES_SCRIPT_DIR, TABLE1_SCRIPT_DIR]:
    if str(script_dir) not in sys.path:
        sys.path.insert(0, str(script_dir))

import build_table2_table3_drop_decomposition as t23  # noqa: E402
import build_table1_drop_decomposition as t1  # noqa: E402


OUT_DIR = SECTION_DIR / "outputs"
DIAG_DIR = OUT_DIR / "diagnostics"
TEACHER_COMBINED_XLSX = OUT_DIR / "teacher_formula_interaction_models_combined.xlsx"
TEACHER_COMBINED_SUMMARY_MD = OUT_DIR / "TEACHER_FORMULA_INTERACTION_SUMMARY_ZH.md"
TEACHER_COMBINED_DIAGNOSTICS_JSON = DIAG_DIR / "teacher_formula_interaction_diagnostics.json"
ONLINE_ACTIVITY_SINGLE_XLSX = OUT_DIR / "teacher_formula_online_activity_single_feature_interaction_models.xlsx"
ONLINE_ACTIVITY_SINGLE_SUMMARY_MD = OUT_DIR / "TEACHER_FORMULA_ONLINE_ACTIVITY_SINGLE_FEATURE_INTERACTION_SUMMARY_ZH.md"
ONLINE_ACTIVITY_SINGLE_DIAGNOSTICS_JSON = DIAG_DIR / "teacher_formula_online_activity_single_feature_interaction_diagnostics.json"
ONLINE_ACTIVITY_ADJUSTED_XLSX = OUT_DIR / "teacher_formula_online_activity_top20_adjusted_interaction_models.xlsx"
ONLINE_ACTIVITY_ADJUSTED_SUMMARY_MD = OUT_DIR / "TEACHER_FORMULA_ONLINE_ACTIVITY_TOP20_ADJUSTED_INTERACTION_SUMMARY_ZH.md"
ONLINE_ACTIVITY_ADJUSTED_DIAGNOSTICS_JSON = DIAG_DIR / "teacher_formula_online_activity_top20_adjusted_interaction_diagnostics.json"

TOP20_XLSX = (
    PAPER_RESULTS_DIR
    / "04_feature_importance_top20"
    / "outputs"
    / "lasso_top20_feature_importance_with_categories.xlsx"
)

ONLINE_ACTIVITY_ITEMS_W2 = ["v21_3", "v21_4", "v21_5", "v21_6"]
INTERPERSONAL_SPECS = t1.INTERPERSONAL_TABLE1_FEATURES
INTERPERSONAL_VERSION = "observed_count"

MODERATOR_SPECS = [
    {
        "id": "online_activity",
        "name": "Online Activity",
        "high_label": "High Online Activity",
        "low_label": "Low Online Activity",
        "high_group_text": "high-online group",
        "low_group_text": "low-online group",
        "definition_sheet_name": "OnlineActivityDefinition",
        "definition_label": "W2 Online Activity",
        "main_question": "Among the LASSO top 20 features from section 04, which features are risk-amplifying or protective among high-online-activity students?",
        "skip_feature_codes": set(),
    },
]

ANALYSIS_MODES = [
    {
        "id": "single_feature",
        "label": "Single-feature + gender interaction",
        "xlsx": ONLINE_ACTIVITY_SINGLE_XLSX,
        "summary_md": ONLINE_ACTIVITY_SINGLE_SUMMARY_MD,
        "diagnostics_json": ONLINE_ACTIVITY_SINGLE_DIAGNOSTICS_JSON,
        "model_description": "Distress ~ one LASSO Top20 feature + W2 High Online Activity + Feature x W2 High Online Activity + Gender.",
    },
    {
        "id": "top20_adjusted",
        "label": "Top20-adjusted interaction",
        "xlsx": ONLINE_ACTIVITY_ADJUSTED_XLSX,
        "summary_md": ONLINE_ACTIVITY_ADJUSTED_SUMMARY_MD,
        "diagnostics_json": ONLINE_ACTIVITY_ADJUSTED_DIAGNOSTICS_JSON,
        "model_description": "Distress ~ task-specific LASSO Top20 main effects + W2 High Online Activity + one Feature x W2 High Online Activity interaction.",
    },
]

TASKS = [
    {
        "Task": "W2 -> W2",
        "Feature Wave": "W2",
        "Target Wave": "W2",
        "Feature Target Group": "v55",
        "Target Items": t23.TARGET_W2_ITEMS,
    },
    {
        "Task": "W2 -> W3",
        "Feature Wave": "W2",
        "Target Wave": "W3",
        "Feature Target Group": "54",
        "Target Items": t23.TARGET_W3_ITEMS,
    },
]

def reset_outputs() -> None:
    out_resolved = OUT_DIR.resolve()
    section_resolved = SECTION_DIR.resolve()
    if out_resolved.parent != section_resolved:
        raise RuntimeError(f"Refusing to remove unexpected output path: {out_resolved}")
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    DIAG_DIR.mkdir(parents=True, exist_ok=True)


def load_project_inputs() -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    merged_path = t23.core.pick_first_existing_path(t23.core.MERGED_PATH_CANDIDATES)
    merged = pd.read_csv(merged_path, dtype=str, encoding="utf-8-sig")
    for col in ["Year", "Group_ID", "Question_ID"]:
        if col in merged.columns:
            merged[col] = merged[col].astype(str).str.strip()
    datasets = {
        "W2": t23.core.normalize_student_id(pd.read_csv(t23.W2_DATA, encoding="utf-8-sig", low_memory=False)),
        "W3": t23.core.normalize_student_id(pd.read_csv(t23.W3_DATA, encoding="utf-8-sig", low_memory=False)),
    }
    return merged, datasets


def make_high_online_activity_w2(w2: pd.DataFrame) -> tuple[pd.Series, pd.Series, dict[str, Any]]:
    values = w2[ONLINE_ACTIVITY_ITEMS_W2].apply(pd.to_numeric, errors="coerce")
    complete = values.notna().sum(axis=1).eq(len(ONLINE_ACTIVITY_ITEMS_W2))
    score = values.sum(axis=1, min_count=len(ONLINE_ACTIVITY_ITEMS_W2))
    median = float(score.loc[complete].median(skipna=True))
    binary = pd.Series(np.nan, index=w2.index, dtype=float)
    binary.loc[complete] = score.loc[complete].gt(median).astype(float)
    diag = {
        "online_activity_items": ONLINE_ACTIVITY_ITEMS_W2,
        "online_activity_definition": "sum(v21_3 to v21_6) > W2 median",
        "online_activity_complete_rows": int(complete.sum()),
        "online_activity_median": median,
        "high_online_n": int(binary.eq(1).sum()),
        "low_online_n": int(binary.eq(0).sum()),
        "missing_online_n": int(binary.isna().sum()),
    }
    return score, binary, diag


def infer_category(code: str) -> str:
    if code.startswith("v54") or code == "v52":
        return "SEL / Resilience"
    if code in {"v5", "v6", "v19"}:
        return "Family / Parenting"
    if code.startswith(("v22", "v23", "v25", "v26", "v27")) or code in {"v28", "v49"}:
        return "Online / Digital Life"
    if code in {"v34", "v36", "v38", "v40"}:
        return "Bullying / Victimization"
    if code == "v42":
        return "Delinquency / Risk Behavior"
    return "Other"


def load_top20_features(feature_defs: pd.DataFrame) -> pd.DataFrame:
    if not TOP20_XLSX.exists():
        raise FileNotFoundError(f"Missing LASSO top20 workbook: {TOP20_XLSX}")
    defs = feature_defs.copy()
    defs["feature_code"] = defs["feature_code"].astype(str)
    top20 = pd.read_excel(TOP20_XLSX, sheet_name="LASSO_Top20_Combined")
    top20["Feature Code"] = top20["Feature Code"].astype(str)
    top20 = top20.sort_values(["Task", "Rank by Abs Std. B", "Feature Code"])
    rows: list[dict[str, Any]] = []
    for _, top in top20.iterrows():
        code = str(top["Feature Code"])
        match_def = defs[defs["feature_code"].eq(code)]
        if match_def.empty:
            continue
        feature_meta = match_def.iloc[0].to_dict()
        rows.append(
            {
                "Task": top["Task"],
                "Feature Code": code,
                "Variable": top.get("Variable", feature_meta.get("feature_name", code)),
                "Model Column": feature_meta.get("model_column", f"feature_{code}"),
                "Category": top.get("Category", infer_category(code)),
                "LASSO Top20 Rank": top.get("Rank by Abs Std. B", np.nan),
                "LASSO Std. B": top.get("Std. B", np.nan),
                "LASSO Relative Importance %": top.get("Relative Importance %", np.nan),
                "LASSO Direction": top.get("Direction", ""),
                "Selected by LASSO": top.get("Selected by LASSO", np.nan),
                "Is Interpersonal Feature": top.get("Is Interpersonal Feature", np.nan),
                "Items": top.get("Items", feature_meta.get("items", "")),
                "Source Type": feature_meta.get("source_type", ""),
            }
        )
    return pd.DataFrame(rows)


def make_target_for_task(
    task: dict[str, Any],
    datasets: dict[str, pd.DataFrame],
    feature_df: pd.DataFrame,
) -> tuple[pd.Series, dict[str, Any]]:
    target_df = datasets[task["Target Wave"]]
    _, y_raw, diag = t23.make_target(target_df, task["Target Items"])
    target_map = pd.DataFrame({"student_id": target_df["student_id"], "target": y_raw}).drop_duplicates("student_id", keep="first")
    y = feature_df[["student_id"]].merge(target_map, on="student_id", how="left")["target"]
    y.index = feature_df.index
    return y, diag


def is_binary_series(values: pd.Series) -> bool:
    unique = pd.to_numeric(values, errors="coerce").dropna().unique()
    if len(unique) == 0:
        return False
    return set(float(v) for v in unique).issubset({0.0, 1.0})


def prepare_model_frame(
    task: dict[str, Any],
    feature_df: pd.DataFrame,
    X: pd.DataFrame,
    y: pd.Series,
    high_online: pd.Series,
    feature: pd.Series,
    feature_code: str,
    feature_name: str,
    moderator_spec: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    frame = pd.DataFrame(
        {
            "student_id": feature_df["student_id"],
            "y": pd.to_numeric(y, errors="coerce"),
            "high_online": pd.to_numeric(high_online, errors="coerce"),
            "feature_raw": pd.to_numeric(feature, errors="coerce"),
        },
        index=feature_df.index,
    )
    if "feature_v1_male" in X.columns and feature_code != "v1_male":
        frame["gender_male"] = pd.to_numeric(X["feature_v1_male"], errors="coerce")

    required = ["y", "high_online", "feature_raw"]
    if "gender_male" in frame.columns:
        required.append("gender_male")
    frame = frame.dropna(subset=required).copy()

    binary = is_binary_series(frame["feature_raw"])
    if binary:
        frame["feature_model"] = frame["feature_raw"]
        scale_note = "binary_0_1"
        raw_mean = float(frame["feature_raw"].mean())
        raw_sd = float(frame["feature_raw"].std(ddof=0)) if len(frame) > 1 else np.nan
    else:
        raw_mean = float(frame["feature_raw"].mean())
        raw_sd = float(frame["feature_raw"].std(ddof=0))
        if raw_sd == 0 or pd.isna(raw_sd):
            frame["feature_model"] = np.nan
        else:
            frame["feature_model"] = (frame["feature_raw"] - raw_mean) / raw_sd
        scale_note = "z_score"

    frame = frame.dropna(subset=["feature_model"]).copy()
    frame["feature_x_high_online"] = frame["feature_model"] * frame["high_online"]

    diag = {
        "Task": task["Task"],
        "Moderator": moderator_spec["name"],
        "Feature": feature_name,
        "Feature Code": feature_code,
        "N model rows": int(len(frame)),
        "Target positive n": int(frame["y"].eq(1).sum()),
        "Target negative n": int(frame["y"].eq(0).sum()),
        f"{moderator_spec['high_label']} n": int(frame["high_online"].eq(1).sum()),
        f"{moderator_spec['low_label']} n": int(frame["high_online"].eq(0).sum()),
        "Feature scale": scale_note,
        "Feature raw mean": raw_mean,
        "Feature raw SD": raw_sd,
        "Feature raw min": float(frame["feature_raw"].min()) if len(frame) else np.nan,
        "Feature raw max": float(frame["feature_raw"].max()) if len(frame) else np.nan,
    }
    return frame, diag


def fit_interaction_logit(frame: pd.DataFrame) -> tuple[Any | None, str, pd.DataFrame, list[str]]:
    terms = ["feature_model", "high_online", "feature_x_high_online"]
    if "gender_male" in frame.columns and frame["gender_male"].nunique(dropna=True) > 1:
        terms.append("gender_male")
    model_df = frame.dropna(subset=["y"] + terms).copy()
    if len(model_df) < 100 or model_df["y"].nunique() < 2:
        return None, "insufficient_n_or_target_variation", model_df, terms
    for term in terms:
        if model_df[term].nunique(dropna=True) < 2:
            return None, f"insufficient_variation: {term}", model_df, terms
    try:
        fit = sm.Logit(model_df["y"], sm.add_constant(model_df[terms], has_constant="add")).fit(disp=False, maxiter=300)
        return fit, "ok", model_df, terms
    except Exception as exc:
        try:
            fit = sm.GLM(
                model_df["y"],
                sm.add_constant(model_df[terms], has_constant="add"),
                family=sm.families.Binomial(),
            ).fit()
            return fit, f"glm_fallback_after_logit_error: {type(exc).__name__}", model_df, terms
        except Exception as exc2:
            return None, f"fit_failed: {type(exc).__name__}; glm_failed: {type(exc2).__name__}", model_df, terms


def normal_p_from_z(z: float) -> float:
    if pd.isna(z):
        return np.nan
    return float(math.erfc(abs(z) / math.sqrt(2.0)))


def term_stats(fit: Any | None, term: str) -> dict[str, float]:
    if fit is None or term not in fit.params.index:
        return {"B": np.nan, "SE": np.nan, "p-value": np.nan}
    return {
        "B": float(fit.params[term]),
        "SE": float(fit.bse[term]),
        "p-value": float(fit.pvalues[term]),
    }


def high_online_slope_stats(fit: Any | None) -> dict[str, float]:
    if fit is None or "feature_model" not in fit.params.index or "feature_x_high_online" not in fit.params.index:
        return {"B": np.nan, "SE": np.nan, "p-value": np.nan}
    b = float(fit.params["feature_model"] + fit.params["feature_x_high_online"])
    cov = fit.cov_params()
    var = (
        float(cov.loc["feature_model", "feature_model"])
        + float(cov.loc["feature_x_high_online", "feature_x_high_online"])
        + 2.0 * float(cov.loc["feature_model", "feature_x_high_online"])
    )
    se = math.sqrt(var) if var >= 0 else np.nan
    z = b / se if se and not pd.isna(se) else np.nan
    return {"B": b, "SE": se, "p-value": normal_p_from_z(z)}


def safe_auc(y: pd.Series, prob: pd.Series) -> float:
    try:
        if y.nunique(dropna=True) < 2:
            return np.nan
        return float(roc_auc_score(y, prob))
    except Exception:
        return np.nan


def model_metrics(fit: Any | None, model_df: pd.DataFrame, terms: list[str]) -> dict[str, float]:
    if fit is None or model_df.empty:
        return {"AUC apparent": np.nan, "Accuracy apparent": np.nan}
    prob = pd.Series(fit.predict(sm.add_constant(model_df[terms], has_constant="add")), index=model_df.index)
    return {
        "AUC apparent": safe_auc(model_df["y"], prob),
        "Accuracy apparent": float(accuracy_score(model_df["y"], prob.ge(0.5).astype(int))),
    }


def moderator_effect_at_feature_levels(fit: Any | None, moderator_spec: dict[str, Any]) -> list[dict[str, Any]]:
    if fit is None:
        return []
    rows: list[dict[str, Any]] = []
    interaction = term_stats(fit, "feature_x_high_online")
    moderator = term_stats(fit, "high_online")
    for label, feature_value in [("Low feature (-1 SD)", -1.0), ("Mean feature", 0.0), ("High feature (+1 SD)", 1.0)]:
        b = moderator["B"] + interaction["B"] * feature_value
        rows.append(
            {
                "Feature Level": label,
                "Feature Model Value": feature_value,
                f"{moderator_spec['high_label']} B at Feature Level": b,
            }
        )
    return rows


def direction_label(b: float, p: float, group: str) -> str:
    if pd.isna(b) or pd.isna(p):
        return "Not available"
    if p >= 0.05:
        return f"Not statistically clear in {group}"
    if b < 0:
        return f"Protective association in {group}"
    return f"Risk association in {group}"


def interaction_label(low_b: float, high_b: float, diff_b: float, p: float, high_label: str) -> str:
    if pd.isna(low_b) or pd.isna(high_b) or pd.isna(diff_b) or pd.isna(p):
        return "Not available"
    if p >= 0.10:
        return "No clear high-vs-low slope difference"

    strength = "significantly" if p < 0.05 else "marginally"
    prefix = f"{high_label} slope is"
    if low_b > 0 and high_b > 0:
        direction = "weaker as a risk association" if abs(high_b) < abs(low_b) else "stronger as a risk association"
    elif low_b < 0 and high_b < 0:
        direction = "stronger as a protective association" if abs(high_b) > abs(low_b) else "weaker as a protective association"
    elif low_b >= 0 and high_b < 0:
        direction = "shifted from risk/non-protective to protective"
    elif low_b < 0 and high_b >= 0:
        direction = "shifted from protective to risk/non-protective"
    else:
        direction = "different from low-online slope"
    return f"{prefix} {strength} {direction}"


def linear_combination_stats(fit: Any | None, weights: dict[str, float]) -> dict[str, float]:
    if fit is None:
        return {"B": np.nan, "SE": np.nan, "p-value": np.nan}
    for term in weights:
        if term not in fit.params.index:
            return {"B": np.nan, "SE": np.nan, "p-value": np.nan}
    b = float(sum(float(weight) * float(fit.params[term]) for term, weight in weights.items()))
    cov = fit.cov_params()
    var = 0.0
    for term_i, weight_i in weights.items():
        for term_j, weight_j in weights.items():
            var += float(weight_i) * float(weight_j) * float(cov.loc[term_i, term_j])
    se = math.sqrt(var) if var >= 0 else np.nan
    z = b / se if se and not pd.isna(se) else np.nan
    return {"B": b, "SE": se, "p-value": normal_p_from_z(z)}


def odds_ratio(b: float) -> float:
    if pd.isna(b):
        return np.nan
    try:
        return float(math.exp(b))
    except OverflowError:
        return np.nan


def safe_term_suffix(value: Any) -> str:
    suffix = re.sub(r"[^0-9A-Za-z]+", "_", str(value)).strip("_")
    return suffix or "feature"


def fit_adjusted_interaction_logit(
    frame: pd.DataFrame,
    terms: list[str],
) -> tuple[Any | None, str, pd.DataFrame, list[str]]:
    model_df = frame.dropna(subset=["y"] + terms).copy()
    if len(model_df) < 100 or model_df["y"].nunique(dropna=True) < 2:
        return None, "insufficient_n_or_target_variation", model_df, terms
    usable_terms: list[str] = []
    dropped_terms: list[str] = []
    for term in terms:
        if model_df[term].nunique(dropna=True) < 2:
            dropped_terms.append(term)
        else:
            usable_terms.append(term)
    required_terms = {"feature_model", "high_online", "feature_x_high_online"}
    if not required_terms.issubset(set(usable_terms)):
        return None, f"insufficient_variation_in_required_terms: {','.join(sorted(required_terms - set(usable_terms)))}", model_df, usable_terms
    status_suffix = ""
    if dropped_terms:
        status_suffix = f"; dropped_zero_variance_terms: {','.join(dropped_terms)}"
    try:
        fit = sm.Logit(model_df["y"], sm.add_constant(model_df[usable_terms], has_constant="add")).fit(disp=False, maxiter=300)
        return fit, f"ok{status_suffix}", model_df, usable_terms
    except Exception as exc:
        try:
            fit = sm.GLM(
                model_df["y"],
                sm.add_constant(model_df[usable_terms], has_constant="add"),
                family=sm.families.Binomial(),
            ).fit()
            return fit, f"glm_fallback_after_logit_error: {type(exc).__name__}{status_suffix}", model_df, usable_terms
        except Exception as exc2:
            return None, f"fit_failed: {type(exc).__name__}; glm_failed: {type(exc2).__name__}{status_suffix}", model_df, usable_terms


def prepare_adjusted_model_frame(
    task: dict[str, Any],
    feature_df: pd.DataFrame,
    X: pd.DataFrame,
    y: pd.Series,
    moderator_binary: pd.Series,
    task_candidates: pd.DataFrame,
    focal_candidate: pd.Series,
    moderator_spec: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any], list[str], list[dict[str, Any]]]:
    focal_code = str(focal_candidate["Feature Code"])
    raw_frame = pd.DataFrame(
        {
            "student_id": feature_df["student_id"],
            "y": pd.to_numeric(y, errors="coerce"),
            "high_online": pd.to_numeric(moderator_binary, errors="coerce"),
        },
        index=feature_df.index,
    )
    term_meta: list[dict[str, Any]] = []
    missing_cols: list[str] = []

    for _, cand in task_candidates.iterrows():
        code = str(cand["Feature Code"])
        model_col = str(cand["Model Column"])
        if model_col not in X.columns:
            missing_cols.append(f"{code}:{model_col}")
            continue
        term = "feature_model" if code == focal_code else f"adj_{safe_term_suffix(code)}"
        raw_col = f"raw_{term}"
        raw_frame[raw_col] = pd.to_numeric(X[model_col], errors="coerce")
        term_meta.append(
            {
                "Feature Code": code,
                "Feature": cand["Variable"],
                "Category": cand["Category"],
                "Model Column": model_col,
                "Raw Column": raw_col,
                "Term": term,
                "Is Focal": code == focal_code,
            }
        )

    if missing_cols:
        raise KeyError(f"Missing Top20 model columns for {task['Task']}: {', '.join(missing_cols)}")
    if not any(meta["Is Focal"] for meta in term_meta):
        raise KeyError(f"Focal feature not available in Top20 model columns: {focal_code}")

    required = ["y", "high_online"] + [str(meta["Raw Column"]) for meta in term_meta]
    frame = raw_frame.dropna(subset=required).copy()
    terms: list[str] = []
    scale_rows: list[dict[str, Any]] = []
    for meta in term_meta:
        raw_col = str(meta["Raw Column"])
        term = str(meta["Term"])
        raw = frame[raw_col]
        raw_mean = float(raw.mean())
        raw_sd = float(raw.std(ddof=0)) if len(raw) > 1 else np.nan
        if is_binary_series(raw):
            frame[term] = raw
            scale_note = "binary_0_1"
        elif raw_sd == 0 or pd.isna(raw_sd):
            frame[term] = np.nan
            scale_note = "zero_variance"
        else:
            frame[term] = (raw - raw_mean) / raw_sd
            scale_note = "z_score"
        terms.append(term)
        scale_rows.append(
            {
                **meta,
                "Feature scale": scale_note,
                "Raw mean": raw_mean,
                "Raw SD": raw_sd,
                "Raw min": float(raw.min()) if len(raw) else np.nan,
                "Raw max": float(raw.max()) if len(raw) else np.nan,
            }
        )

    frame = frame.dropna(subset=terms).copy()
    frame["feature_raw"] = frame["raw_feature_model"]
    frame["feature_x_high_online"] = frame["feature_model"] * frame["high_online"]
    ordered_terms = ["feature_model"] + [t for t in terms if t != "feature_model"] + ["high_online", "feature_x_high_online"]
    focal_scale = next(row for row in scale_rows if row["Is Focal"])
    diag = {
        "Task": task["Task"],
        "Moderator": moderator_spec["name"],
        "Feature": focal_candidate["Variable"],
        "Feature Code": focal_code,
        "N model rows": int(len(frame)),
        "Target positive n": int(frame["y"].eq(1).sum()),
        "Target negative n": int(frame["y"].eq(0).sum()),
        f"{moderator_spec['high_label']} n": int(frame["high_online"].eq(1).sum()),
        f"{moderator_spec['low_label']} n": int(frame["high_online"].eq(0).sum()),
        "Feature scale": focal_scale["Feature scale"],
        "Feature raw mean": focal_scale["Raw mean"],
        "Feature raw SD": focal_scale["Raw SD"],
        "Feature raw min": focal_scale["Raw min"],
        "Feature raw max": focal_scale["Raw max"],
        "Adjusted main effects count": int(len(terms)),
        "Adjusted predictor count including moderator and interaction": int(len(ordered_terms)),
        "Adjusted feature codes": ";".join(str(meta["Feature Code"]) for meta in term_meta),
        "Adjusted feature names": ";".join(str(meta["Feature"]) for meta in term_meta),
    }
    return frame, diag, ordered_terms, scale_rows


def teacher_interpretation(b1: float, b3: float, slope0: float, slope1: float, p3: float, high_label: str) -> str:
    if pd.isna(b1) or pd.isna(b3) or pd.isna(slope0) or pd.isna(slope1) or pd.isna(p3):
        return "Not available"
    if p3 < 0.05:
        strength = "significantly"
    elif p3 < 0.10:
        strength = "marginally"
    else:
        return "No clear interaction evidence; b3 is not statistically clear."

    if slope0 > 0 and slope1 > 0:
        direction = "weaker risk slope" if abs(slope1) < abs(slope0) else "stronger risk slope"
    elif slope0 < 0 and slope1 < 0:
        direction = "stronger protective slope" if abs(slope1) > abs(slope0) else "weaker protective slope"
    elif slope0 >= 0 and slope1 < 0:
        direction = "shift from risk/non-protective to protective slope"
    elif slope0 < 0 and slope1 >= 0:
        direction = "shift from protective to risk/non-protective slope"
    else:
        direction = "different slope"
    return f"{high_label}=1 has a {strength} {direction} compared with {high_label}=0."


def build_teacher_row(
    task: dict[str, Any],
    cand: pd.Series,
    frame: pd.DataFrame,
    diag: dict[str, Any],
    moderator_spec: dict[str, Any],
    terms: list[str] | None = None,
    analysis_mode: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    analysis_mode = analysis_mode or {"id": "single_feature", "label": "Single-feature + gender interaction"}
    if terms is None:
        fit, status, model_df, terms = fit_interaction_logit(frame)
        adjusted_note = "Single-feature interaction model."
        model_formula = "logit(P(High Psychological Distress=1)) = b0 + b1*Feature + b2*ModeratorHigh + b3*Feature*ModeratorHigh + gender_male"
        main_effect_count = 1
        predictor_count = len(terms)
    else:
        fit, status, model_df, terms = fit_adjusted_interaction_logit(frame, terms)
        adjusted_note = "Top20-adjusted interaction model: all LASSO Top20 main effects are included, then one Feature x Moderator interaction is added."
        model_formula = "logit(P(High Psychological Distress=1)) = Top20 main effects + ModeratorHigh + Feature*ModeratorHigh"
        main_effect_count = diag.get("Adjusted main effects count", np.nan)
        predictor_count = diag.get("Adjusted predictor count including moderator and interaction", len(terms))
    metrics = model_metrics(fit, model_df, terms)
    high_label = moderator_spec["high_label"]
    low_label = moderator_spec["low_label"]

    b0 = term_stats(fit, "const")
    b1 = term_stats(fit, "feature_model")
    b2 = term_stats(fit, "high_online")
    b3 = term_stats(fit, "feature_x_high_online")
    intercept0 = b0
    intercept1 = linear_combination_stats(fit, {"const": 1.0, "high_online": 1.0})
    slope0 = b1
    slope1 = linear_combination_stats(fit, {"feature_model": 1.0, "feature_x_high_online": 1.0})
    gender = term_stats(fit, "gender_male") if "gender_male" in terms else {"B": np.nan, "SE": np.nan, "p-value": np.nan}

    common = {
        "Analysis Mode": analysis_mode["label"],
        "Analysis Mode ID": analysis_mode["id"],
        "Task": task["Task"],
        "Moderator": moderator_spec["name"],
        "Moderator High Group Label": high_label,
        "Moderator Low Group Label": low_label,
        "Feature": cand["Variable"],
        "Feature Code": cand["Feature Code"],
        "Category": cand["Category"],
        "Feature scale": diag["Feature scale"],
        "N": int(len(model_df)),
        "Target positive n": int(model_df["y"].eq(1).sum()) if not model_df.empty else 0,
        "Target negative n": int(model_df["y"].eq(0).sum()) if not model_df.empty else 0,
        f"{high_label} n": int(model_df["high_online"].eq(1).sum()) if not model_df.empty else 0,
        f"{low_label} n": int(model_df["high_online"].eq(0).sum()) if not model_df.empty else 0,
        "LASSO Top20 Rank": cand.get("LASSO Top20 Rank", np.nan),
        "LASSO Std. B": cand.get("LASSO Std. B", np.nan),
        "LASSO Relative Importance %": cand.get("LASSO Relative Importance %", np.nan),
        "LASSO Direction": cand.get("LASSO Direction", ""),
        "Is Interpersonal Feature": cand.get("Is Interpersonal Feature", np.nan),
        "Items": cand.get("Items", ""),
        "Fit status": status,
        "Model formula": model_formula,
        "Covariate note": adjusted_note,
        "Adjusted main effects count": main_effect_count,
        "Adjusted predictor count including moderator and interaction": predictor_count,
        "Adjusted feature codes": diag.get("Adjusted feature codes", ""),
        "Adjusted feature names": diag.get("Adjusted feature names", ""),
    }

    row = {
        **common,
        "b0 Intercept B": b0["B"],
        "b0 Intercept SE": b0["SE"],
        "b0 Intercept p-value": b0["p-value"],
        "b1 Feature Main Effect B": b1["B"],
        "b1 Feature Main Effect SE": b1["SE"],
        "b1 Feature Main Effect p-value": b1["p-value"],
        "b2 Moderator Main Effect B": b2["B"],
        "b2 Moderator Main Effect SE": b2["SE"],
        "b2 Moderator Main Effect p-value": b2["p-value"],
        "b3 Feature x Moderator B": b3["B"],
        "b3 Feature x Moderator SE": b3["SE"],
        "b3 Feature x Moderator p-value": b3["p-value"],
        "Gender Male B": gender["B"],
        "Gender Male SE": gender["SE"],
        "Gender Male p-value": gender["p-value"],
        "Intercept when Moderator=0": intercept0["B"],
        "Intercept when Moderator=0 SE": intercept0["SE"],
        "Intercept when Moderator=0 p-value": intercept0["p-value"],
        "Slope when Moderator=0": slope0["B"],
        "Slope when Moderator=0 SE": slope0["SE"],
        "Slope when Moderator=0 p-value": slope0["p-value"],
        "Intercept when Moderator=1": intercept1["B"],
        "Intercept when Moderator=1 SE": intercept1["SE"],
        "Intercept when Moderator=1 p-value": intercept1["p-value"],
        "Slope when Moderator=1": slope1["B"],
        "Slope when Moderator=1 SE": slope1["SE"],
        "Slope when Moderator=1 p-value": slope1["p-value"],
        "Slope Difference Moderator1 minus Moderator0": b3["B"],
        "Slope Difference p-value": b3["p-value"],
        "OR Slope Moderator=0": odds_ratio(slope0["B"]),
        "OR Slope Moderator=1": odds_ratio(slope1["B"]),
        "OR Interaction b3": odds_ratio(b3["B"]),
        "AUC apparent": metrics["AUC apparent"],
        "Accuracy apparent": metrics["Accuracy apparent"],
        "Teacher Formula Interpretation": teacher_interpretation(
            b1["B"],
            b3["B"],
            slope0["B"],
            slope1["B"],
            b3["p-value"],
            high_label,
        ),
    }

    term_rows = build_coefficient_terms(common, fit, terms, intercept1, slope1, metrics, moderator_spec)
    prediction_rows = build_predicted_probability_rows(common, fit, terms, diag["Feature scale"], moderator_spec)
    return row, term_rows, prediction_rows


def build_coefficient_terms(
    common: dict[str, Any],
    fit: Any | None,
    terms: list[str],
    intercept1: dict[str, float],
    slope1: dict[str, float],
    metrics: dict[str, float],
    moderator_spec: dict[str, Any],
) -> list[dict[str, Any]]:
    if fit is None:
        return [{**common, "Coefficient Label": "", "Raw Term": "", "B": np.nan, "SE": np.nan, "p-value": np.nan, **metrics}]
    labels = {
        "const": "b0 Intercept",
        "feature_model": "b1 Feature Main Effect",
        "high_online": "b2 Moderator Main Effect",
        "feature_x_high_online": "b3 Feature x Moderator",
        "gender_male": "Covariate: Gender Male",
    }
    rows: list[dict[str, Any]] = []
    ordered_terms = ["const", "feature_model", "high_online", "feature_x_high_online"]
    ordered_terms.extend([term for term in terms if term.startswith("adj_")])
    ordered_terms.append("gender_male")
    for term in ordered_terms:
        if term not in fit.params.index:
            continue
        st = term_stats(fit, term)
        label = labels.get(term, f"Adjusted Top20 Main Effect: {term.replace('adj_', '')}")
        rows.append({**common, "Coefficient Label": label, "Raw Term": term, "B": st["B"], "SE": st["SE"], "p-value": st["p-value"], **metrics})
    rows.extend(
        [
            {
                **common,
                "Coefficient Label": "Derived: Intercept when Moderator=1 = b0 + b2",
                "Raw Term": "const + high_online",
                "B": intercept1["B"],
                "SE": intercept1["SE"],
                "p-value": intercept1["p-value"],
                **metrics,
            },
            {
                **common,
                "Coefficient Label": "Derived: Slope when Moderator=1 = b1 + b3",
                "Raw Term": "feature_model + feature_x_high_online",
                "B": slope1["B"],
                "SE": slope1["SE"],
                "p-value": slope1["p-value"],
                **metrics,
            },
        ]
    )
    return rows


def inverse_logit(value: float) -> float:
    if pd.isna(value):
        return np.nan
    if value >= 0:
        z = math.exp(-value)
        return float(1 / (1 + z))
    z = math.exp(value)
    return float(z / (1 + z))


def build_predicted_probability_rows(
    common: dict[str, Any],
    fit: Any | None,
    terms: list[str],
    feature_scale: str,
    moderator_spec: dict[str, Any],
) -> list[dict[str, Any]]:
    if fit is None:
        return []
    feature_values = [0.0, 1.0] if feature_scale == "binary_0_1" else [-2.0, -1.0, 0.0, 1.0, 2.0]
    rows: list[dict[str, Any]] = []
    for moderator_value in [0.0, 1.0]:
        for feature_value in feature_values:
            values = {
                "const": 1.0,
                "feature_model": feature_value,
                "high_online": moderator_value,
                "feature_x_high_online": feature_value * moderator_value,
                "gender_male": 0.0,
            }
            linear_predictor = 0.0
            for term, coefficient in fit.params.items():
                linear_predictor += float(coefficient) * float(values.get(term, 0.0))
            rows.append(
                {
                    **common,
                    "Moderator Value": int(moderator_value),
                    "Moderator Group": moderator_spec["high_label"] if moderator_value == 1 else moderator_spec["low_label"],
                    "Feature Model Value": feature_value,
                    "Feature Value Label": "1" if feature_scale == "binary_0_1" and feature_value == 1 else "0" if feature_scale == "binary_0_1" else f"{feature_value:+.0f} SD",
                    "Gender Male Setting": 0 if "gender_male" in terms else np.nan,
                    "Gender Setting Note": "Predicted probability shown at gender_male=0 when gender is an adjustment covariate." if "gender_male" in terms else "No separate gender adjustment in this model.",
                    "Linear Predictor Logit": linear_predictor,
                    "Predicted Probability": inverse_logit(linear_predictor),
                }
            )
    return rows


def build_base_inputs() -> tuple[pd.DataFrame, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, dict[str, Any], pd.DataFrame]:
    merged, datasets = load_project_inputs()
    feature_df = datasets["W2"]
    X, feature_defs, feature_diag = t23.build_drop_decomposition_features(feature_df, merged, "W2", "v55")
    X, feature_defs, feature_diag = t1.add_interpersonal_features(
        "W2",
        feature_df,
        X,
        feature_defs,
        feature_diag,
        INTERPERSONAL_SPECS,
        INTERPERSONAL_VERSION,
    )
    candidates = load_top20_features(feature_defs)
    return feature_df, datasets, X, feature_defs, feature_diag, candidates


def build_outputs_for_moderator(
    *,
    moderator_spec: dict[str, Any],
    analysis_mode: dict[str, Any],
    feature_df: pd.DataFrame,
    datasets: dict[str, pd.DataFrame],
    X: pd.DataFrame,
    feature_defs: pd.DataFrame,
    feature_diag: dict[str, Any],
    candidates: pd.DataFrame,
    moderator_binary: pd.Series,
    moderator_diag: dict[str, Any],
) -> dict[str, pd.DataFrame]:
    coefficient_rows: list[dict[str, Any]] = []
    term_rows: list[dict[str, Any]] = []
    prediction_rows: list[dict[str, Any]] = []
    diag_rows: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    scale_rows_all: list[dict[str, Any]] = []
    skip_feature_codes = set(moderator_spec.get("skip_feature_codes", set()))
    model_mode = analysis_mode["id"]

    for task in TASKS:
        y, target_diag = make_target_for_task(task, datasets, feature_df)
        task_candidates = candidates[candidates["Task"].eq(task["Task"])].copy()
        for _, cand in task_candidates.iterrows():
            feature_code = str(cand["Feature Code"])
            if feature_code in skip_feature_codes:
                skipped = {
                    "Task": task["Task"],
                    "Moderator": moderator_spec["name"],
                    "Feature": cand["Variable"],
                    "Feature Code": feature_code,
                    "Reason": (
                        f"{feature_code} is used to define {moderator_spec['high_label']} / "
                        f"{moderator_spec['low_label']}; self-interaction is excluded."
                    ),
                }
                skipped_rows.append(skipped)
                diag_rows.append({**skipped, "Status": "skipped_self_interaction"})
                continue
            model_col = str(cand["Model Column"])
            if model_col not in X.columns:
                diag_rows.append(
                    {
                        "Task": task["Task"],
                        "Moderator": moderator_spec["name"],
                        "Feature": cand["Variable"],
                        "Feature Code": feature_code,
                        "Status": f"missing model column: {model_col}",
                    }
                )
                continue
            if model_mode == "single_feature":
                frame, diag = prepare_model_frame(
                    task=task,
                    feature_df=feature_df,
                    X=X,
                    y=y,
                    high_online=moderator_binary,
                    feature=X[model_col],
                    feature_code=feature_code,
                    feature_name=str(cand["Variable"]),
                    moderator_spec=moderator_spec,
                )
                terms = None
                scale_rows = [
                    {
                        "Feature Code": feature_code,
                        "Feature": cand["Variable"],
                        "Category": cand["Category"],
                        "Model Column": model_col,
                        "Raw Column": "feature_raw",
                        "Term": "feature_model",
                        "Is Focal": True,
                        "Feature scale": diag["Feature scale"],
                        "Raw mean": diag["Feature raw mean"],
                        "Raw SD": diag["Feature raw SD"],
                        "Raw min": diag["Feature raw min"],
                        "Raw max": diag["Feature raw max"],
                    }
                ]
            elif model_mode == "top20_adjusted":
                frame, diag, terms, scale_rows = prepare_adjusted_model_frame(
                    task=task,
                    feature_df=feature_df,
                    X=X,
                    y=y,
                    moderator_binary=moderator_binary,
                    task_candidates=task_candidates,
                    focal_candidate=cand,
                    moderator_spec=moderator_spec,
                )
            else:
                raise ValueError(f"Unsupported analysis mode: {model_mode}")
            diag = {**diag, **{f"target_{k}": v for k, v in target_diag.items()}}
            diag["Analysis Mode"] = analysis_mode["label"]
            diag["Analysis Mode ID"] = analysis_mode["id"]
            diag_rows.append(diag)
            scale_rows_all.extend(
                {
                    "Analysis Mode": analysis_mode["label"],
                    "Analysis Mode ID": analysis_mode["id"],
                    "Task": task["Task"],
                    "Moderator": moderator_spec["name"],
                    "Focal Feature Code": feature_code,
                    "Focal Feature": cand["Variable"],
                    **row,
                }
                for row in scale_rows
            )
            coefficient_row, coefficients_long, predicted = build_teacher_row(
                task,
                cand,
                frame,
                diag,
                moderator_spec,
                terms=terms,
                analysis_mode=analysis_mode,
            )
            coefficient_rows.append(coefficient_row)
            term_rows.extend(coefficients_long)
            prediction_rows.extend(predicted)

    coefficients = pd.DataFrame(coefficient_rows)
    terms_long = pd.DataFrame(term_rows)
    predicted_prob = pd.DataFrame(prediction_rows)
    diagnostics = pd.DataFrame(diag_rows)
    skipped_df = pd.DataFrame(skipped_rows)
    scale_df = pd.DataFrame(scale_rows_all)

    if not coefficients.empty:
        coefficients["Interaction significant p<.05"] = pd.to_numeric(coefficients["b3 Feature x Moderator p-value"], errors="coerce").lt(0.05)
        coefficients["Interaction marginal p<.10"] = pd.to_numeric(coefficients["b3 Feature x Moderator p-value"], errors="coerce").lt(0.10)
        coefficients = coefficients.sort_values(
            ["Task", "b3 Feature x Moderator p-value", "Feature Code"],
            ascending=[True, True, True],
        ).reset_index(drop=True)

    readme = pd.DataFrame(
        [
            {
                "Item": "Teacher formula",
                "Description": analysis_mode["model_description"],
            },
            {
                "Item": "Moderator=0 intercept and slope",
                "Description": "Intercept = b0; slope = b1.",
            },
            {
                "Item": "Moderator=1 intercept and slope",
                "Description": "Intercept = b0 + b2; slope = b1 + b3.",
            },
            {
                "Item": "Adjusted feature set",
                "Description": (
                    "Each row uses one focal Top20 feature plus gender adjustment."
                    if model_mode == "single_feature"
                    else "Each row uses the task-specific LASSO Top20 main effects. One interaction term is added at a time, so W2->W2 has 20 models and W2->W3 has 20 models."
                ),
            },
            {
                "Item": "Outcome scale",
                "Description": "Because the outcome is binary high psychological distress, B values are logistic-regression log-odds coefficients.",
            },
            {
                "Item": "Feature scaling",
                "Description": "Continuous features are z-scored; binary features remain 0/1.",
            },
        ]
    )

    return {
        "ReadMe": readme,
        "TeacherFormulaCoefficients": coefficients,
        "CoefficientTermsLong": terms_long,
        "PredictedProbabilities": predicted_prob,
        "LASSOTop20Features": candidates,
        "SkippedFeatures": skipped_df,
        "Diagnostics": diagnostics,
        "FeatureScaling": scale_df,
        moderator_spec["definition_sheet_name"]: pd.DataFrame([moderator_diag]),
        "FeatureSetDiagnostics": pd.DataFrame([feature_diag]),
    }


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 80))
            ws.column_dimensions[letter].width = max(12, max_len + 2)
    wb.save(path)


def format_p(value: Any) -> str:
    if pd.isna(value):
        return ""
    value = float(value)
    if value < 0.001:
        return "<0.001"
    if value < 0.10:
        return f"{value:.4f}"
    return f"{value:.3f}"


def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df.empty:
        return "_No rows available._"
    show = df if max_rows is None else df.head(max_rows)
    out = show.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            if "p-value" in col:
                out[col] = out[col].map(format_p)
            else:
                out[col] = out[col].map(lambda v: "" if pd.isna(v) else f"{float(v):.4f}")
    return out.to_markdown(index=False)


def formula_explanation(row: pd.Series, moderator_spec: dict[str, Any]) -> list[str]:
    feature = str(row["Feature"])
    task = str(row["Task"])
    high_label = str(moderator_spec["high_label"])
    low_label = str(moderator_spec["low_label"])
    b0 = float(row["b0 Intercept B"])
    b1 = float(row["b1 Feature Main Effect B"])
    b2 = float(row["b2 Moderator Main Effect B"])
    b3 = float(row["b3 Feature x Moderator B"])
    p3 = float(row["b3 Feature x Moderator p-value"])
    int0 = float(row["Intercept when Moderator=0"])
    slope0 = float(row["Slope when Moderator=0"])
    int1 = float(row["Intercept when Moderator=1"])
    slope1 = float(row["Slope when Moderator=1"])
    or0 = float(row["OR Slope Moderator=0"])
    or1 = float(row["OR Slope Moderator=1"])
    or3 = float(row["OR Interaction b3"])
    analysis_id = str(row.get("Analysis Mode ID", ""))
    if analysis_id == "single_feature":
        formula_main = "= b0 + b1 * Feature + b2 * ModeratorHigh + b3 * Feature * ModeratorHigh + gender_male"
        substituted_main = f"= {b0:.4f} + ({b1:.4f}) * Feature + ({b2:.4f}) * ModeratorHigh + ({b3:.4f}) * Feature * ModeratorHigh + gender_male"
        low_slope_sentence = f"- `{low_label}` 中，在控制性別後，`{feature}` 每增加 1 SD，高心理困擾的 log-odds 改變 `{slope0:.4f}`，對應 OR = `{or0:.3f}`。"
        high_slope_sentence = f"- `{high_label}` 中，在控制性別後，`{feature}` 每增加 1 SD，高心理困擾的 log-odds 改變 `{slope1:.4f}`，對應 OR = `{or1:.3f}`。"
    else:
        formula_main = "= all Top20 main effects + b2 * ModeratorHigh + b3 * Feature * ModeratorHigh"
        substituted_main = f"= Top20 main effects, including ({b1:.4f}) * Feature + ({b2:.4f}) * ModeratorHigh + ({b3:.4f}) * Feature * ModeratorHigh"
        low_slope_sentence = f"- `{low_label}` 中，在控制同一任務 Top20 其他主效應後，`{feature}` 每增加 1 SD，高心理困擾的 log-odds 改變 `{slope0:.4f}`，對應 OR = `{or0:.3f}`。"
        high_slope_sentence = f"- `{high_label}` 中，在控制同一任務 Top20 其他主效應後，`{feature}` 每增加 1 SD，高心理困擾的 log-odds 改變 `{slope1:.4f}`，對應 OR = `{or1:.3f}`。"

    if slope0 > 0 and slope1 > 0:
        direction = "兩組都是風險斜率"
        change = "較弱" if abs(slope1) < abs(slope0) else "較強"
    elif slope0 < 0 and slope1 < 0:
        direction = "兩組都是保護斜率"
        change = "較強" if abs(slope1) > abs(slope0) else "較弱"
    elif slope0 >= 0 and slope1 < 0:
        direction = "低組為風險或接近無效，高組轉為保護斜率"
        change = "方向改變"
    elif slope0 < 0 and slope1 >= 0:
        direction = "低組為保護斜率，高組轉為風險或接近無效"
        change = "方向改變"
    else:
        direction = "兩組斜率方向不同"
        change = "不同"

    return [
        f"### {task}: {feature}",
        "",
        "老師公式：",
        "",
        "```text",
        "logit(P(High Psychological Distress = 1))",
        formula_main,
        "```",
        "",
        "代入本結果：",
        "",
        "```text",
        f"b0 = {b0:.4f}",
        f"b1 = {b1:.4f}",
        f"b2 = {b2:.4f}",
        f"b3 = {b3:.4f}, p = {format_p(p3)}",
        "",
        "logit(P(High Psychological Distress = 1))",
        substituted_main,
        "```",
        "",
        f"當 `{moderator_spec['name']} = 0`，也就是 `{low_label}`：",
        "",
        "```text",
        f"intercept = b0 = {int0:.4f}",
        f"slope = b1 = {slope0:.4f}",
        "```",
        "",
        f"當 `{moderator_spec['name']} = 1`，也就是 `{high_label}`：",
        "",
        "```text",
        f"intercept = b0 + b2 = {b0:.4f} + {b2:.4f} = {int1:.4f}",
        f"slope = b1 + b3 = {b1:.4f} + {b3:.4f} = {slope1:.4f}",
        "```",
        "",
        "解釋：",
        "",
        f"- `b3 = {b3:.4f}` 且 `p = {format_p(p3)}`，表示 `{high_label}` 會顯著改變 `{feature}` 與高心理困擾之間的斜率。",
        low_slope_sentence,
        high_slope_sentence,
        f"- interaction OR = `exp(b3) = {or3:.3f}`。",
        f"- 整體來看，{direction}；在 `{high_label}` 中，這個 feature 的斜率比 `{low_label}` {change}。",
        "",
    ]


def formula_explanations(significant: pd.DataFrame, moderator_spec: dict[str, Any]) -> str:
    if significant.empty:
        return "_沒有 b3 p < .05 的顯著 interaction 結果。_"
    lines: list[str] = []
    for _, row in significant.iterrows():
        lines.extend(formula_explanation(row, moderator_spec))
    return "\n".join(lines)


def write_summary(sheets: dict[str, pd.DataFrame], moderator_spec: dict[str, Any]) -> None:
    coef = sheets["TeacherFormulaCoefficients"].copy()
    display_cols = [
        "Analysis Mode",
        "Task",
        "Feature",
        "Category",
        "b1 Feature Main Effect B",
        "b2 Moderator Main Effect B",
        "b3 Feature x Moderator B",
        "b3 Feature x Moderator p-value",
        "Intercept when Moderator=0",
        "Slope when Moderator=0",
        "Intercept when Moderator=1",
        "Slope when Moderator=1",
        "Teacher Formula Interpretation",
    ]
    significant = coef[pd.to_numeric(coef["b3 Feature x Moderator p-value"], errors="coerce").lt(0.05)].copy()
    analysis_id = moderator_spec.get("analysis_id", "")
    if analysis_id == "single_feature":
        model_text = "`logit(P(High Psychological Distress=1)) = b0 + b1*Feature + b2*ModeratorHigh + b3*Feature*ModeratorHigh + gender_male`"
        setup_bullets = [
            "- 每一列都是一個 single-feature interaction model：只放入一個 focal Top20 feature、High Online Activity、交互作用項與性別。",
            "- 因此 W2 -> W2 跑 20 個模型，W2 -> W3 跑 20 個模型，共 40 個模型。",
        ]
        slope_note = "- 連續特徵已標準化為 z-score，因此 slope 表示在控制性別後，該特徵每增加 1 SD 的 log-odds 變化。"
    else:
        model_text = "`logit(P(High Psychological Distress=1)) = task-specific LASSO Top20 main effects + b2*ModeratorHigh + b3*Feature*ModeratorHigh`"
        setup_bullets = [
            "- 每一列都是一個 adjusted interaction model：先放入該任務的 LASSO Top20 主效應，再一次加入一個 `Feature x ModeratorHigh` 交互作用項。",
            "- 因此 W2 -> W2 跑 20 個模型，W2 -> W3 跑 20 個模型，共 40 個模型。",
        ]
        slope_note = "- 連續特徵已標準化為 z-score，因此 slope 表示在控制其他 Top20 主效應後，該特徵每增加 1 SD 的 log-odds 變化。"
    lines = [
        f"# Teacher Formula Interaction Summary: {moderator_spec['name']} - {moderator_spec.get('analysis_label', '')}",
        "",
        "## 模型",
        "",
        model_text,
        "",
        "## 老師公式對應",
        "",
        *setup_bullets,
        "- Moderator = 0: `intercept = b0`, `slope = b1`。",
        "- Moderator = 1: `intercept = b0 + b2`, `slope = b1 + b3`。",
        "- 因為 outcome 是 binary high psychological distress，所以 B 是 log-odds coefficient。",
        slope_note,
        "- 這裡的 p-value 是未做多重比較校正的 exploratory interaction screening；因為總共檢查 40 個 interaction，寫論文時建議保守解讀。",
        "",
        "## b3 interaction 顯著結果 p < .05",
        "",
        md_table(significant[display_cols] if not significant.empty else significant),
        "",
        "## 顯著結果公式代入與解釋",
        "",
        formula_explanations(significant, moderator_spec),
        "",
        "## Outputs",
        "",
        f"- Workbook: `{moderator_spec['xlsx']}`",
        f"- Diagnostics: `{moderator_spec['diagnostics_json']}`",
    ]
    moderator_spec["summary_md"].write_text("\n".join(lines), encoding="utf-8")


def write_outputs(sheets: dict[str, pd.DataFrame], moderator_spec: dict[str, Any]) -> None:
    output_xlsx = moderator_spec["xlsx"]
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet[:31])
    format_workbook(output_xlsx)
    write_summary(sheets, moderator_spec)
    payload = {
        "moderator": moderator_spec["name"],
        "output_xlsx": str(output_xlsx),
        "summary_md": str(moderator_spec["summary_md"]),
        "top20_xlsx": str(TOP20_XLSX),
        "sheets": {name: {"rows": int(len(df)), "columns": list(df.columns)} for name, df in sheets.items()},
    }
    moderator_spec["diagnostics_json"].write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_combined_outputs(all_sheets: dict[str, dict[str, pd.DataFrame]]) -> None:
    comparison = build_single_vs_adjusted_comparison(all_sheets)
    with pd.ExcelWriter(TEACHER_COMBINED_XLSX, engine="openpyxl") as writer:
        for analysis_id, sheets in all_sheets.items():
            prefix = "SingleFeature" if analysis_id == "single_feature" else "Top20Adjusted"
            sheets["TeacherFormulaCoefficients"].to_excel(writer, index=False, sheet_name=f"{prefix}_Coefficients")
            sheets["PredictedProbabilities"].to_excel(writer, index=False, sheet_name=f"{prefix}_PredictedProb")
            sheets["SkippedFeatures"].to_excel(writer, index=False, sheet_name=f"{prefix}_Skipped")
        if not comparison.empty:
            comparison.to_excel(writer, index=False, sheet_name="Single_vs_Adjusted")
        pd.concat(
            [sheets["TeacherFormulaCoefficients"] for sheets in all_sheets.values()],
            ignore_index=True,
        ).to_excel(writer, index=False, sheet_name="All_Coefficients")
    format_workbook(TEACHER_COMBINED_XLSX)
    payload = {
        "combined_xlsx": str(TEACHER_COMBINED_XLSX),
        "analysis_modes": list(all_sheets.keys()),
        "rows": {k: int(len(v["TeacherFormulaCoefficients"])) for k, v in all_sheets.items()},
        "comparison_rows": int(len(comparison)),
    }
    TEACHER_COMBINED_DIAGNOSTICS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_single_vs_adjusted_comparison(all_sheets: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
    if "single_feature" not in all_sheets or "top20_adjusted" not in all_sheets:
        return pd.DataFrame()
    cols = [
        "Task",
        "Feature Code",
        "Feature",
        "Category",
        "b1 Feature Main Effect B",
        "b2 Moderator Main Effect B",
        "b3 Feature x Moderator B",
        "b3 Feature x Moderator p-value",
        "Slope when Moderator=0",
        "Slope when Moderator=1",
        "AUC apparent",
    ]
    single = all_sheets["single_feature"]["TeacherFormulaCoefficients"][cols].copy()
    adjusted = all_sheets["top20_adjusted"]["TeacherFormulaCoefficients"][cols].copy()
    rename_single = {
        "b1 Feature Main Effect B": "Single b1 Feature B",
        "b2 Moderator Main Effect B": "Single b2 Online Activity B",
        "b3 Feature x Moderator B": "Single b3 Interaction B",
        "b3 Feature x Moderator p-value": "Single b3 Interaction p-value",
        "Slope when Moderator=0": "Single Low Online Activity Slope",
        "Slope when Moderator=1": "Single High Online Activity Slope",
        "AUC apparent": "Single Model AUC apparent",
    }
    rename_adjusted = {
        "Feature": "Adjusted Feature",
        "Category": "Adjusted Category",
        "b1 Feature Main Effect B": "Adjusted b1 Feature B",
        "b2 Moderator Main Effect B": "Adjusted b2 Online Activity B",
        "b3 Feature x Moderator B": "Adjusted b3 Interaction B",
        "b3 Feature x Moderator p-value": "Adjusted b3 Interaction p-value",
        "Slope when Moderator=0": "Adjusted Low Online Activity Slope",
        "Slope when Moderator=1": "Adjusted High Online Activity Slope",
        "AUC apparent": "Adjusted Model AUC apparent",
    }
    single = single.rename(columns=rename_single)
    adjusted = adjusted.rename(columns=rename_adjusted)
    merged = single.merge(adjusted, on=["Task", "Feature Code"], how="outer")
    merged["Feature"] = merged["Feature"].fillna(merged.get("Adjusted Feature"))
    merged["Category"] = merged["Category"].fillna(merged.get("Adjusted Category"))
    merged = merged.drop(columns=[c for c in ["Adjusted Feature", "Adjusted Category"] if c in merged.columns])
    merged["Single significant p<.05"] = pd.to_numeric(merged["Single b3 Interaction p-value"], errors="coerce").lt(0.05)
    merged["Adjusted significant p<.05"] = pd.to_numeric(merged["Adjusted b3 Interaction p-value"], errors="coerce").lt(0.05)

    def label(row: pd.Series) -> str:
        if row["Single significant p<.05"] and row["Adjusted significant p<.05"]:
            return "Significant in both models"
        if row["Single significant p<.05"]:
            return "Significant only in single-feature model"
        if row["Adjusted significant p<.05"]:
            return "Significant only in Top20-adjusted model"
        return "Not significant in either model"

    merged["Significance pattern"] = merged.apply(label, axis=1)
    merged["Adjusted minus Single b3"] = pd.to_numeric(merged["Adjusted b3 Interaction B"], errors="coerce") - pd.to_numeric(merged["Single b3 Interaction B"], errors="coerce")
    return merged.sort_values(["Task", "Adjusted b3 Interaction p-value", "Single b3 Interaction p-value", "Feature Code"]).reset_index(drop=True)


def write_combined_summary(all_sheets: dict[str, dict[str, pd.DataFrame]]) -> None:
    comparison = build_single_vs_adjusted_comparison(all_sheets)
    lines = [
        "# Teacher Formula Interaction Analysis Summary",
        "",
        "## 這次 06 的重點",
        "",
        "這版 06 同時保留兩種老師公式 interaction models：",
        "",
        "1. Single-feature + gender: `logit(P(High Psychological Distress=1)) = b0 + b1*Feature + b2*ModeratorHigh + b3*Feature*ModeratorHigh + gender_male`。",
        "2. Top20-adjusted: `logit(P(High Psychological Distress=1)) = task-specific LASSO Top20 main effects + b2*ModeratorHigh + b3*Feature*ModeratorHigh`。",
        "",
        "- 每個任務使用該任務自己的 LASSO Top20 作為候選 focal features。",
        "- 每種模型都一次只加入一個 `Feature x ModeratorHigh` interaction term。",
        "- 每種模型各跑 40 個 interaction tests：W2 -> W2 20 個，W2 -> W3 20 個。",
        "- Moderator = 0: `intercept = b0`, `slope = b1`。",
        "- Moderator = 1: `intercept = b0 + b2`, `slope = b1 + b3`。",
        "- `b3` 是真正的 interaction effect，檢查 moderator 是否改變 feature 對 high psychological distress 的斜率。",
        "- p-value 未做多重比較校正；本段應作為 exploratory interaction screening，而不是確認性因果證據。",
        "",
    ]
    display_cols = [
        "Analysis Mode",
        "Task",
        "Moderator",
        "Feature",
        "b1 Feature Main Effect B",
        "b2 Moderator Main Effect B",
        "b3 Feature x Moderator B",
        "b3 Feature x Moderator p-value",
        "Slope when Moderator=0",
        "Slope when Moderator=1",
        "Teacher Formula Interpretation",
    ]
    for analysis_id, sheets in all_sheets.items():
        coef = sheets["TeacherFormulaCoefficients"].copy()
        significant = coef[pd.to_numeric(coef["b3 Feature x Moderator p-value"], errors="coerce").lt(0.05)]
        analysis_label = coef["Analysis Mode"].dropna().iloc[0] if not coef.empty and "Analysis Mode" in coef.columns else analysis_id
        lines.extend(
            [
                f"## {analysis_label}",
                "",
                "### b3 interaction 顯著結果 p < .05",
                "",
                md_table(significant[display_cols] if not significant.empty else significant),
                "",
            ]
        )
    if not comparison.empty:
        comparison_display_cols = [
            "Task",
            "Feature",
            "Category",
            "Single b3 Interaction B",
            "Single b3 Interaction p-value",
            "Adjusted b3 Interaction B",
            "Adjusted b3 Interaction p-value",
            "Significance pattern",
        ]
        noteworthy = comparison[
            comparison["Single significant p<.05"] | comparison["Adjusted significant p<.05"]
        ].copy()
        lines.extend(
            [
                "## Single-feature vs Top20-adjusted 對照",
                "",
                "這張表用來判斷 interaction 是否在只控制性別時顯著，或是在控制 Top20 主效應後仍顯著。",
                "",
                md_table(noteworthy[comparison_display_cols] if not noteworthy.empty else noteworthy),
                "",
            ]
        )
    lines.extend(
        [
            "## Outputs",
            "",
            f"- Combined workbook: `{TEACHER_COMBINED_XLSX}`",
            f"- Single-feature workbook: `{ONLINE_ACTIVITY_SINGLE_XLSX}`",
            f"- Top20-adjusted workbook: `{ONLINE_ACTIVITY_ADJUSTED_XLSX}`",
        ]
    )
    TEACHER_COMBINED_SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    reset_outputs()
    feature_df, datasets, X, feature_defs, feature_diag, candidates = build_base_inputs()
    _, high_online, online_diag = make_high_online_activity_w2(feature_df)
    moderator_inputs = {
        "online_activity": (high_online, online_diag),
    }

    all_sheets: dict[str, dict[str, pd.DataFrame]] = {}
    for moderator_spec in MODERATOR_SPECS:
        moderator_binary, moderator_diag = moderator_inputs[moderator_spec["id"]]
        for analysis_mode in ANALYSIS_MODES:
            output_spec = {
                **moderator_spec,
                **analysis_mode,
                "analysis_id": analysis_mode["id"],
                "analysis_label": analysis_mode["label"],
            }
            sheets = build_outputs_for_moderator(
                moderator_spec=output_spec,
                analysis_mode=analysis_mode,
                feature_df=feature_df,
                datasets=datasets,
                X=X,
                feature_defs=feature_defs,
                feature_diag=feature_diag,
                candidates=candidates,
                moderator_binary=moderator_binary,
                moderator_diag=moderator_diag,
            )
            write_outputs(sheets, output_spec)
            all_sheets[analysis_mode["id"]] = sheets
            print(f"Wrote {output_spec['xlsx']}")
            print(f"Wrote {output_spec['summary_md']}")
            print(f"Wrote {output_spec['diagnostics_json']}")
            main_cols = [
                "Analysis Mode",
                "Task",
                "Feature",
                "b1 Feature Main Effect B",
                "b2 Moderator Main Effect B",
                "b3 Feature x Moderator B",
                "b3 Feature x Moderator p-value",
                "Slope when Moderator=0",
                "Slope when Moderator=1",
            ]
            print(f"\nTop teacher-formula rows for {output_spec['name']} / {analysis_mode['label']}:")
            print(sheets["TeacherFormulaCoefficients"][main_cols].head(12).to_string(index=False))

    write_combined_outputs(all_sheets)
    write_combined_summary(all_sheets)
    print(f"Wrote {TEACHER_COMBINED_XLSX}")
    print(f"Wrote {TEACHER_COMBINED_SUMMARY_MD}")
    print(f"Wrote {TEACHER_COMBINED_DIAGNOSTICS_JSON}")

if __name__ == "__main__":
    main()


