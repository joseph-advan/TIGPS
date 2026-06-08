from __future__ import annotations

import json
import shutil
import sys
import textwrap
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression, LogisticRegressionCV
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.preprocessing import StandardScaler


SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
PAPER_DATA_DIR = PAPER_RESULTS_DIR.parent
INTERPERSONAL_DIR = PAPER_RESULTS_DIR / "03_interpersonal_incremental_modeling"
if str(INTERPERSONAL_DIR) not in sys.path:
    sys.path.insert(0, str(INTERPERSONAL_DIR))

import run_interpersonal_incremental_modeling as im3  # noqa: E402

SOURCE_XLSX = (
    PAPER_RESULTS_DIR
    / "03_interpersonal_incremental_modeling"
    / "outputs"
    / "interpersonal_feature_selection_summary.xlsx"
)

OUT_DIR = SECTION_DIR / "outputs"
DIAG_DIR = OUT_DIR / "diagnostics"
FIG_DIR = OUT_DIR / "figures"
MAIN_XLSX = OUT_DIR / "lasso_top20_feature_importance_with_categories.xlsx"
SUMMARY_MD = OUT_DIR / "LASSO_TOP20_FEATURE_IMPORTANCE_SUMMARY.md"
DIAGNOSTICS_JSON = DIAG_DIR / "lasso_top20_feature_importance_diagnostics.json"

MAIN_FEATURE_SET = "decomposed_plus_12_interpersonal"
TOP_N = 20
LASSO_NONZERO_TOL = 1e-8

DISPLAY_NAME_OVERRIDES = {
    "v22": "Online Coping and Emotion Regulation under Distress",
    "v23_B": "Authentic and Less-Ideal Self-Presentation",
    "v23_C": "Covert Social Media Monitoring and Passive Participation",
    "v27_B": "Distress from Missing Online Events",
    "v34": "Offline Bullying Victimization",
    "v36": "Offline Bullying Perpetration",
    "v42": "Delinquent and Risk Behaviors",
    "v52": "Self-Worth and Positive Self-Concept",
}


CATEGORY_RULES = [
    ("Interpersonal Network", lambda code, name: str(code).startswith("ip_")),
    ("SEL / Resilience", lambda code, name: str(code).startswith("v54") or str(code) == "v52"),
    ("Family / Parenting", lambda code, name: str(code) in {"v5", "v6", "v19"}),
    (
        "Online / Digital Life",
        lambda code, name: str(code).startswith(("v22", "v23", "v25", "v26", "v27"))
        or str(code) in {"v28", "v49"},
    ),
    ("Bullying / Victimization", lambda code, name: str(code) in {"v34", "v36", "v38", "v40"}),
    ("Delinquency / Risk Behavior", lambda code, name: str(code) == "v42"),
    ("Demographic / Social Status", lambda code, name: str(code) in {"v1", "1", "v1_male", "1_male", "v3"}),
    ("School Context / Belonging", lambda code, name: str(code) in {"v9", "v12", "v8_03-v8_06"}),
]

CATEGORY_ORDER = [
    "SEL / Resilience",
    "Family / Parenting",
    "Online / Digital Life",
    "Bullying / Victimization",
    "Interpersonal Network",
    "Demographic / Social Status",
    "School Context / Belonging",
    "Delinquency / Risk Behavior",
    "Other",
]

CATEGORY_COLORS = {
    "SEL / Resilience": "#2F6B4F",
    "Family / Parenting": "#B56B45",
    "Online / Digital Life": "#2F5F98",
    "Bullying / Victimization": "#A33A3A",
    "Interpersonal Network": "#6C5A9E",
    "Demographic / Social Status": "#6B7280",
    "School Context / Belonging": "#3D7C7A",
    "Delinquency / Risk Behavior": "#9A7B22",
    "Other": "#8A8A8A",
}

MODEL_ORDER = {
    "LASSO Logistic": 1,
    "Ridge Logistic": 2,
    "Multivariable Logistic": 3,
}

TASK_ORDER = {"W2 -> W2": 1, "W2 -> W3": 2}


def reset_outputs() -> None:
    out_resolved = OUT_DIR.resolve()
    section_resolved = SECTION_DIR.resolve()
    if out_resolved.parent != section_resolved:
        raise RuntimeError(f"Refusing to remove unexpected output path: {out_resolved}")
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    DIAG_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)


def categorize(feature_code: Any, variable: Any) -> str:
    code = "" if pd.isna(feature_code) else str(feature_code)
    name = "" if pd.isna(variable) else str(variable)
    for category, rule in CATEGORY_RULES:
        if rule(code, name):
            return category
    return "Other"


def load_coefficients() -> pd.DataFrame:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source coefficient workbook: {SOURCE_XLSX}")
    df = pd.read_excel(SOURCE_XLSX, sheet_name="AllCoefficientsLong")
    df = df[df["Feature Set"].eq(MAIN_FEATURE_SET)].copy()
    df["Abs Std. B"] = pd.to_numeric(df["Abs Std. B"], errors="coerce")
    df["Std. B"] = pd.to_numeric(df["Std. B"], errors="coerce")
    df["Relative Importance %"] = pd.to_numeric(df["Relative Importance %"], errors="coerce")
    df["Rank by Abs Std. B"] = pd.to_numeric(df["Rank by Abs Std. B"], errors="coerce")
    df["Direction"] = df["Std. B"].apply(lambda v: "Positive" if pd.notna(v) and v > 0 else ("Negative" if pd.notna(v) and v < 0 else "Zero"))
    df["Category"] = [categorize(code, var) for code, var in zip(df["Feature Code"], df["Variable"])]
    df["Variable"] = df["Variable"].astype(str).str.replace("Parent?hild", "Parent-Child", regex=False)
    df["Variable"] = [
        DISPLAY_NAME_OVERRIDES.get(str(code), variable)
        for code, variable in zip(df["Feature Code"], df["Variable"])
    ]
    df["Task Order"] = df["Task"].map(TASK_ORDER).fillna(99)
    df["Model Order"] = df["Model"].map(MODEL_ORDER).fillna(99)
    df = df.sort_values(["Task Order", "Model Order", "Rank by Abs Std. B", "Feature Code"]).reset_index(drop=True)
    return df


def compute_lasso_lambda_path() -> tuple[pd.DataFrame, pd.DataFrame]:
    merged, w2, w3 = im3.read_inputs()
    datasets = {"W2": w2, "W3": w3}
    feature_set = next(fs for fs in im3.FEATURE_SETS if fs["feature_set"] == MAIN_FEATURE_SET)
    rows: list[dict[str, Any]] = []
    path_rows: list[dict[str, Any]] = []

    for task in im3.TASKS:
        feature_df = datasets[task["feature_wave"]]
        y, _ = im3.make_target_for_task(task, datasets, feature_df)
        X, defs, _ = im3.build_feature_set(task, feature_set, merged, datasets)
        valid = y.notna()
        yv = y[valid].astype(int)
        Xv = X.loc[valid].copy()
        X_train, _, y_train, _ = train_test_split(
            Xv,
            yv,
            test_size=im3.t23.TEST_SIZE,
            random_state=im3.t23.RANDOM_STATE,
            stratify=yv,
        )

        imputer = SimpleImputer(strategy="median")
        scaler = StandardScaler()
        X_train_imp = imputer.fit_transform(X_train)
        X_train_scaled = scaler.fit_transform(X_train_imp)

        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=im3.t23.RANDOM_STATE)
        cv_model = LogisticRegressionCV(
            Cs=im3.t23.LOGIT_CS,
            penalty="l1",
            solver="saga",
            cv=cv,
            scoring="roc_auc",
            max_iter=20000,
            n_jobs=None,
            random_state=im3.t23.RANDOM_STATE,
        )
        cv_model.fit(X_train_scaled, y_train)
        selected_c = float(cv_model.C_[0])
        selected_lambda = 1.0 / selected_c if selected_c else np.nan

        coef_by_c: dict[float, np.ndarray] = {}
        for c_value in sorted(float(c) for c in im3.t23.LOGIT_CS):
            model = LogisticRegression(
                C=c_value,
                penalty="l1",
                solver="saga",
                max_iter=20000,
                random_state=im3.t23.RANDOM_STATE,
            )
            model.fit(X_train_scaled, y_train)
            coef = model.coef_[0]
            coef_by_c[c_value] = coef
            lambda_value = 1.0 / c_value
            nonzero_count = int((np.abs(coef) > LASSO_NONZERO_TOL).sum())
            path_rows.append(
                {
                    "Task": task["task"],
                    "C": c_value,
                    "Lambda Approx. 1/C": lambda_value,
                    "N Nonzero Features": nonzero_count,
                }
            )

        selected_grid_c = min(coef_by_c, key=lambda c: abs(c - selected_c))
        selected_coef = coef_by_c[selected_grid_c]
        cs_ascending = sorted(coef_by_c)
        defs_reset = defs.reset_index(drop=True)
        for idx, def_row in defs_reset.iterrows():
            nonzero_cs = [
                c_value
                for c_value in cs_ascending
                if abs(float(coef_by_c[c_value][idx])) > LASSO_NONZERO_TOL
            ]
            entry_c = float(nonzero_cs[0]) if nonzero_cs else np.nan
            entry_lambda = 1.0 / entry_c if pd.notna(entry_c) and entry_c else np.nan
            selected_b = float(selected_coef[idx])
            rows.append(
                {
                    "Task": task["task"],
                    "Feature Code": str(def_row["feature_code"]),
                    "Variable": DISPLAY_NAME_OVERRIDES.get(str(def_row["feature_code"]), str(def_row["feature_name"])),
                    "Model Column": str(def_row["model_column"]),
                    "Source Type": def_row.get("source_type", ""),
                    "Items": def_row.get("items", ""),
                    "Category": categorize(def_row["feature_code"], def_row["feature_name"]),
                    "Selected C": selected_c,
                    "Selected Lambda Approx. 1/C": selected_lambda,
                    "Entry C": entry_c,
                    "Entry Lambda Approx. 1/C": entry_lambda,
                    "Coefficient at Selected Lambda": selected_b,
                    "Selected at Best Lambda": abs(selected_b) > LASSO_NONZERO_TOL,
                    "Ever Selected on Lambda Path": bool(nonzero_cs),
                }
            )

    path_df = pd.DataFrame(rows)
    if not path_df.empty:
        path_df["Lambda Entry RI %"] = 0.0
        path_df["Rank by Lambda Entry"] = np.nan
        for task, idx in path_df.groupby("Task").groups.items():
            sub_idx = list(idx)
            entry = pd.to_numeric(path_df.loc[sub_idx, "Entry Lambda Approx. 1/C"], errors="coerce")
            denom = entry.fillna(0).sum()
            if denom > 0:
                path_df.loc[sub_idx, "Lambda Entry RI %"] = entry.fillna(0) / denom * 100.0
            path_df.loc[sub_idx, "Rank by Lambda Entry"] = entry.rank(method="min", ascending=False)
        path_df["Top 20 by Lambda Entry"] = path_df["Rank by Lambda Entry"].le(TOP_N)
        path_df["Direction at Selected Lambda"] = path_df["Coefficient at Selected Lambda"].apply(
            lambda v: "Positive" if pd.notna(v) and v > 0 else ("Negative" if pd.notna(v) and v < 0 else "Zero")
        )
        path_df["_task_order"] = path_df["Task"].map(TASK_ORDER).fillna(99)
        path_df = path_df.sort_values(["_task_order", "Rank by Lambda Entry", "Feature Code"]).drop(columns="_task_order").reset_index(drop=True)

    return path_df, pd.DataFrame(path_rows)


def add_lambda_metrics(coef: pd.DataFrame, lambda_path: pd.DataFrame) -> pd.DataFrame:
    if lambda_path.empty:
        out = coef.copy()
        out["Coefficient RI %"] = out["Relative Importance %"]
        return out
    lambda_cols = [
        "Task",
        "Feature Code",
        "Selected C",
        "Selected Lambda Approx. 1/C",
        "Entry C",
        "Entry Lambda Approx. 1/C",
        "Lambda Entry RI %",
        "Rank by Lambda Entry",
        "Top 20 by Lambda Entry",
        "Coefficient at Selected Lambda",
        "Selected at Best Lambda",
        "Ever Selected on Lambda Path",
    ]
    out = coef.merge(lambda_path[lambda_cols], on=["Task", "Feature Code"], how="left")
    out["Coefficient RI %"] = out["Relative Importance %"]
    return out


def top_n(df: pd.DataFrame, model: str, n: int = TOP_N) -> pd.DataFrame:
    sub = df[df["Model"].eq(model)].copy()
    out = sub[sub["Rank by Abs Std. B"].le(n)].copy()
    out = out.sort_values(["Task Order", "Rank by Abs Std. B", "Feature Code"])
    cols = [
        "Task",
        "Model",
        "Rank by Abs Std. B",
        "Variable",
        "Feature Code",
        "Category",
        "Std. B",
        "Abs Std. B",
        "Relative Importance %",
        "Coefficient RI %",
        "Rank by Lambda Entry",
        "Lambda Entry RI %",
        "Entry Lambda Approx. 1/C",
        "Entry C",
        "Selected Lambda Approx. 1/C",
        "Selected C",
        "Selected at Best Lambda",
        "Direction",
        "Selected by LASSO",
        "Is Interpersonal Feature",
        "Items",
    ]
    return out[[c for c in cols if c in out.columns]]


def lambda_top_n(df: pd.DataFrame, n: int = TOP_N) -> pd.DataFrame:
    sub = df[df["Model"].eq("LASSO Logistic")].copy()
    out = sub[sub["Rank by Lambda Entry"].le(n)].copy()
    out = out.sort_values(["Task Order", "Rank by Lambda Entry", "Feature Code"])
    cols = [
        "Task",
        "Model",
        "Rank by Lambda Entry",
        "Variable",
        "Feature Code",
        "Category",
        "Entry Lambda Approx. 1/C",
        "Entry C",
        "Lambda Entry RI %",
        "Rank by Abs Std. B",
        "Std. B",
        "Coefficient RI %",
        "Selected Lambda Approx. 1/C",
        "Selected C",
        "Selected by LASSO",
        "Selected at Best Lambda",
        "Direction",
        "Is Interpersonal Feature",
        "Items",
    ]
    return out[[c for c in cols if c in out.columns]]


def compare_coeff_vs_lambda_top20(coeff_top20: pd.DataFrame, lambda_top20: pd.DataFrame) -> pd.DataFrame:
    coeff = coeff_top20[["Task", "Feature Code", "Variable", "Category", "Rank by Abs Std. B", "Coefficient RI %"]].copy()
    lam = lambda_top20[["Task", "Feature Code", "Variable", "Category", "Rank by Lambda Entry", "Lambda Entry RI %", "Entry Lambda Approx. 1/C"]].copy()
    merged = coeff.merge(lam, on=["Task", "Feature Code"], how="outer", suffixes=(" Coeff", " Lambda"))
    merged["Variable"] = merged["Variable Coeff"].fillna(merged["Variable Lambda"])
    merged["Category"] = merged["Category Coeff"].fillna(merged["Category Lambda"])
    merged["In Coefficient Top20"] = merged["Rank by Abs Std. B"].notna()
    merged["In Lambda Top20"] = merged["Rank by Lambda Entry"].notna()
    merged["Rank Difference Lambda minus Coefficient"] = merged["Rank by Lambda Entry"] - merged["Rank by Abs Std. B"]
    merged["Top20 Pattern"] = merged.apply(
        lambda row: "In both Top20"
        if row["In Coefficient Top20"] and row["In Lambda Top20"]
        else ("Coefficient Top20 only" if row["In Coefficient Top20"] else "Lambda Top20 only"),
        axis=1,
    )
    drop_cols = [c for c in ["Variable Coeff", "Variable Lambda", "Category Coeff", "Category Lambda"] if c in merged.columns]
    merged = merged.drop(columns=drop_cols)
    merged["_task_order"] = merged["Task"].map(TASK_ORDER).fillna(99)
    pattern_order = {"In both Top20": 1, "Coefficient Top20 only": 2, "Lambda Top20 only": 3}
    merged["_pattern_order"] = merged["Top20 Pattern"].map(pattern_order).fillna(99)
    return merged.sort_values(["_task_order", "_pattern_order", "Rank by Abs Std. B", "Rank by Lambda Entry", "Feature Code"]).drop(columns=["_task_order", "_pattern_order"]).reset_index(drop=True)


def lasso_top20_by_task(lasso_top20: pd.DataFrame, task: str) -> pd.DataFrame:
    return lasso_top20[lasso_top20["Task"].eq(task)].copy()


def category_summary(lasso_top20: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for task, sub in lasso_top20.groupby("Task", sort=False):
        for cat in CATEGORY_ORDER:
            g = sub[sub["Category"].eq(cat)]
            if g.empty:
                continue
            rows.append(
                {
                    "Task": task,
                    "Category": cat,
                    "N Top 20 Features": int(len(g)),
                    "Relative Importance Sum %": float(g["Relative Importance %"].sum(skipna=True)),
                    "Mean Relative Importance %": float(g["Relative Importance %"].mean(skipna=True)),
                    "Top Feature": g.sort_values("Rank by Abs Std. B").iloc[0]["Variable"],
                    "Top Feature Rank": int(g["Rank by Abs Std. B"].min()),
                }
            )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["_task_order"] = out["Task"].map(TASK_ORDER).fillna(99)
    out["_category_order"] = out["Category"].map({c: i for i, c in enumerate(CATEGORY_ORDER, 1)}).fillna(99)
    out = out.sort_values(["_task_order", "_category_order"]).drop(columns=["_task_order", "_category_order"])
    return out.reset_index(drop=True)


def category_summary_lambda(lambda_top20: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for task, sub in lambda_top20.groupby("Task", sort=False):
        for cat in CATEGORY_ORDER:
            g = sub[sub["Category"].eq(cat)]
            if g.empty:
                continue
            rows.append(
                {
                    "Task": task,
                    "Category": cat,
                    "N Top 20 Features": int(len(g)),
                    "Lambda Entry RI Sum %": float(g["Lambda Entry RI %"].sum(skipna=True)),
                    "Mean Lambda Entry RI %": float(g["Lambda Entry RI %"].mean(skipna=True)),
                    "Top Feature": g.sort_values("Rank by Lambda Entry").iloc[0]["Variable"],
                    "Top Feature Rank": int(g["Rank by Lambda Entry"].min()),
                }
            )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out["_task_order"] = out["Task"].map(TASK_ORDER).fillna(99)
    out["_category_order"] = out["Category"].map({c: i for i, c in enumerate(CATEGORY_ORDER, 1)}).fillna(99)
    out = out.sort_values(["_task_order", "_category_order"]).drop(columns=["_task_order", "_category_order"])
    return out.reset_index(drop=True)


def category_pivot(summary: pd.DataFrame) -> pd.DataFrame:
    if summary.empty:
        return summary
    count = summary.pivot(index="Category", columns="Task", values="N Top 20 Features")
    imp = summary.pivot(index="Category", columns="Task", values="Relative Importance Sum %")
    out = pd.DataFrame(index=CATEGORY_ORDER)
    for task in TASK_ORDER:
        out[f"{task} N Top 20"] = count.get(task)
        out[f"{task} Relative Importance Sum %"] = imp.get(task)
    out = out.dropna(how="all").reset_index(names="Category")
    imp_cols = [c for c in out.columns if c.endswith("Relative Importance Sum %")]
    out["Combined Relative Importance Sum %"] = out[imp_cols].fillna(0).sum(axis=1)
    out = out.sort_values(["Combined Relative Importance Sum %", "Category"], ascending=[False, True]).reset_index(drop=True)
    return out


def category_pivot_lambda(summary: pd.DataFrame) -> pd.DataFrame:
    if summary.empty:
        return summary
    count = summary.pivot(index="Category", columns="Task", values="N Top 20 Features")
    imp = summary.pivot(index="Category", columns="Task", values="Lambda Entry RI Sum %")
    out = pd.DataFrame(index=CATEGORY_ORDER)
    for task in TASK_ORDER:
        out[f"{task} N Top 20"] = count.get(task)
        out[f"{task} Lambda Entry RI Sum %"] = imp.get(task)
    out = out.dropna(how="all").reset_index(names="Category")
    imp_cols = [c for c in out.columns if c.endswith("Lambda Entry RI Sum %")]
    out["Combined Lambda Entry RI Sum %"] = out[imp_cols].fillna(0).sum(axis=1)
    out = out.sort_values(["Combined Lambda Entry RI Sum %", "Category"], ascending=[False, True]).reset_index(drop=True)
    return out


def shared_top20(lasso_top20: pd.DataFrame) -> pd.DataFrame:
    w2 = lasso_top20[lasso_top20["Task"].eq("W2 -> W2")].copy()
    w3 = lasso_top20[lasso_top20["Task"].eq("W2 -> W3")].copy()
    shared = w2.merge(
        w3,
        on="Feature Code",
        how="inner",
        suffixes=(" W2->W2", " W2->W3"),
    )
    if shared.empty:
        return shared
    rows = []
    for _, row in shared.iterrows():
        rows.append(
            {
                "Feature Code": row["Feature Code"],
                "Variable": row["Variable W2->W2"],
                "Category": row["Category W2->W2"],
                "W2 -> W2 Rank": row["Rank by Abs Std. B W2->W2"],
                "W2 -> W2 Std. B": row["Std. B W2->W2"],
                "W2 -> W2 Relative Importance %": row["Relative Importance % W2->W2"],
                "W2 -> W3 Rank": row["Rank by Abs Std. B W2->W3"],
                "W2 -> W3 Std. B": row["Std. B W2->W3"],
                "W2 -> W3 Relative Importance %": row["Relative Importance % W2->W3"],
                "Same Direction": row["Direction W2->W2"] == row["Direction W2->W3"],
            }
        )
    out = pd.DataFrame(rows)
    return out.sort_values(["W2 -> W2 Rank", "W2 -> W3 Rank"]).reset_index(drop=True)


def interpersonal_summary(df: pd.DataFrame) -> pd.DataFrame:
    lasso = df[(df["Model"].eq("LASSO Logistic")) & (df["Is Interpersonal Feature"].astype(bool))].copy()
    rows = []
    for task, sub in lasso.groupby("Task", sort=False):
        selected = sub[sub["Selected by LASSO"].astype(bool)]
        top20 = sub[sub["Rank by Abs Std. B"].le(TOP_N)]
        removed = sub[~sub["Selected by LASSO"].astype(bool)]
        rows.append(
            {
                "Task": task,
                "N Interpersonal Features": int(len(sub)),
                "N Selected by LASSO": int(len(selected)),
                "N Removed by LASSO": int(len(removed)),
                "N in LASSO Top 20": int(len(top20)),
                "Interpersonal Relative Importance Sum %": float(sub["Relative Importance %"].sum(skipna=True)),
                "Top Interpersonal Feature": sub.sort_values("Rank by Abs Std. B").iloc[0]["Variable"],
                "Top Interpersonal Rank": int(sub["Rank by Abs Std. B"].min()),
                "Removed Features": "; ".join(removed["Variable"].astype(str).tolist()),
            }
        )
    return pd.DataFrame(rows)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 72))
            ws.column_dimensions[letter].width = max(12, max_len + 2)
    wb.save(path)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(MAIN_XLSX, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
    format_workbook(MAIN_XLSX)


def wrap_label(value: Any, width: int = 24) -> str:
    text = "" if pd.isna(value) else str(value)
    return "\n".join(textwrap.wrap(text, width=width, break_long_words=False))


def save_lasso_top20_bar_chart(df: pd.DataFrame, task: str) -> Path:
    sub = lasso_top20_by_task(df, task).sort_values("Rank by Abs Std. B")
    safe_task = task.lower().replace(" ", "").replace("->", "_to_")
    path = FIG_DIR / f"lasso_top20_relative_importance_{safe_task}.png"
    if sub.empty:
        return path

    fig, ax = plt.subplots(figsize=(18, 8))
    colors = [CATEGORY_COLORS.get(cat, CATEGORY_COLORS["Other"]) for cat in sub["Category"]]
    ax.bar(range(len(sub)), sub["Relative Importance %"], color=colors)
    ax.set_title(f"LASSO Top 20 Relative Importance: {task}", fontsize=16, weight="bold")
    ax.set_ylabel("Relative Importance (%)")
    ax.set_xlabel("Variable")
    ax.set_xticks(range(len(sub)))
    ax.set_xticklabels([wrap_label(v, 18) for v in sub["Variable"]], rotation=65, ha="right", fontsize=8)
    ax.grid(axis="y", alpha=0.25)
    handles = [
        plt.Rectangle((0, 0), 1, 1, color=CATEGORY_COLORS[cat])
        for cat in CATEGORY_ORDER
        if cat in set(sub["Category"])
    ]
    labels = [cat for cat in CATEGORY_ORDER if cat in set(sub["Category"])]
    ax.legend(handles, labels, loc="upper right", fontsize=9)
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def save_category_summary_chart(category_wide: pd.DataFrame) -> Path:
    path = FIG_DIR / "lasso_top20_category_relative_importance_summary.png"
    if category_wide.empty:
        return path

    plot_df = category_wide.copy()
    if "Combined Relative Importance Sum %" in plot_df.columns:
        plot_df = plot_df.sort_values(
            ["Combined Relative Importance Sum %", "Category"],
            ascending=[False, True],
        )
    plot_df = plot_df.set_index("Category")
    cols = [c for c in plot_df.columns if c.endswith("Relative Importance Sum %")]
    cols = [c for c in cols if not c.startswith("Combined ")]
    fig, ax = plt.subplots(figsize=(13, 7))
    x = range(len(plot_df.index))
    width = 0.36
    for i, col in enumerate(cols):
        offset = (i - (len(cols) - 1) / 2) * width
        label = col.replace(" Relative Importance Sum %", "")
        ax.bar([v + offset for v in x], plot_df[col], width=width, label=label)
    ax.set_title("LASSO Top 20 Relative Importance by Conceptual Category", fontsize=15, weight="bold")
    ax.set_ylabel("Relative Importance Sum (%)")
    ax.set_xlabel("Conceptual Category")
    ax.set_xticks(list(x))
    ax.set_xticklabels([wrap_label(v, 18) for v in plot_df.index], rotation=35, ha="right")
    ax.grid(axis="y", alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def save_shared_top20_chart(shared: pd.DataFrame) -> Path:
    path = FIG_DIR / "shared_lasso_top20_relative_importance.png"
    if shared.empty:
        return path

    plot_df = shared.sort_values("W2 -> W2 Rank").copy()
    fig, ax = plt.subplots(figsize=(15, 8))
    y = range(len(plot_df))
    height = 0.36
    ax.barh([v + height / 2 for v in y], plot_df["W2 -> W2 Relative Importance %"], height=height, label="W2 -> W2")
    ax.barh([v - height / 2 for v in y], plot_df["W2 -> W3 Relative Importance %"], height=height, label="W2 -> W3")
    ax.set_title("Shared LASSO Top 20 Features Across Prediction Tasks", fontsize=15, weight="bold")
    ax.set_xlabel("Relative Importance (%)")
    ax.set_yticks(list(y))
    ax.set_yticklabels([wrap_label(v, 34) for v in plot_df["Variable"]], fontsize=9)
    ax.invert_yaxis()
    ax.grid(axis="x", alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def write_chart_index(paths: list[Path]) -> pd.DataFrame:
    rows = []
    for path in paths:
        rows.append(
            {
                "Figure": path.name,
                "Path": str(path),
                "Purpose": {
                    "lasso_top20_relative_importance_w2_to_w2.png": "Top 20 LASSO predictors for W2 -> W2.",
                    "lasso_top20_relative_importance_w2_to_w3.png": "Top 20 LASSO predictors for W2 -> W3.",
                    "lasso_top20_category_relative_importance_summary.png": "Category-level relative-importance comparison across tasks.",
                    "shared_lasso_top20_relative_importance.png": "Variables appearing in both W2 -> W2 and W2 -> W3 LASSO Top 20 lists.",
                }.get(path.name, ""),
            }
        )
    return pd.DataFrame(rows)


def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df.empty:
        return "_No rows available._"
    show = df if max_rows is None else df.head(max_rows)
    return show.to_markdown(index=False)


def write_summary(
    lasso_top20: pd.DataFrame,
    lambda_top20_df: pd.DataFrame,
    top20_comparison: pd.DataFrame,
    category_sum: pd.DataFrame,
    category_wide: pd.DataFrame,
    category_sum_lambda: pd.DataFrame,
    category_wide_lambda: pd.DataFrame,
    shared: pd.DataFrame,
    ip_summary: pd.DataFrame,
    figure_paths: list[Path],
) -> None:
    lines = [
        "# LASSO Top 20 Feature Importance Summary",
        "",
        "## Purpose",
        "",
        "This section identifies the most important predictors after the interpersonal incremental modeling step. The main table uses the LASSO model with the drop + decomposition + 12 interpersonal feature set.",
        "",
        "## Main Interpretation",
        "",
        "- `03_interpersonal_incremental_modeling` asks whether the 12 interpersonal indicators improve model performance.",
        "- `04_feature_importance_top20` now reports two LASSO ranking approaches: coefficient-based RI and lambda-entry-based RI.",
        "- LASSO is the primary model here because it can shrink weak features to zero and therefore supports feature-selection interpretation.",
        "- Coefficient-based RI is the original ranking: `abs(standardized coefficient) / sum(abs(standardized coefficients)) * 100` at the selected lambda.",
        "- Lambda-entry RI uses the LASSO regularization path: features that become non-zero at stronger penalty levels receive higher lambda-entry scores.",
        "",
        "## Main Outputs",
        "",
        f"- Workbook: `{MAIN_XLSX}`",
        f"- Source coefficients: `{SOURCE_XLSX}`",
        f"- Figures folder: `{FIG_DIR}`",
        "",
        "## Figures",
        "",
        md_table(write_chart_index(figure_paths)),
        "",
        "## Category-Level Summary: Coefficient-Based RI",
        "",
        md_table(category_wide),
        "",
        "## Category-Level Summary: Lambda-Entry RI",
        "",
        md_table(category_wide_lambda),
        "",
        "## Coefficient-Based Top 20 vs Lambda-Entry Top 20",
        "",
        md_table(top20_comparison),
        "",
        "## Shared LASSO Top 20 Features Across W2 -> W2 and W2 -> W3",
        "",
        md_table(shared),
        "",
        "## Interpersonal Features in LASSO",
        "",
        md_table(ip_summary),
        "",
        "## W2 -> W2 LASSO Top 20: Coefficient-Based RI",
        "",
        md_table(lasso_top20_by_task(lasso_top20, "W2 -> W2")),
        "",
        "## W2 -> W3 LASSO Top 20: Coefficient-Based RI",
        "",
        md_table(lasso_top20_by_task(lasso_top20, "W2 -> W3")),
        "",
        "## W2 -> W2 LASSO Top 20: Lambda-Entry RI",
        "",
        md_table(lasso_top20_by_task(lambda_top20_df, "W2 -> W2")),
        "",
        "## W2 -> W3 LASSO Top 20: Lambda-Entry RI",
        "",
        md_table(lasso_top20_by_task(lambda_top20_df, "W2 -> W3")),
        "",
        "## Suggested Paper Logic",
        "",
        "The coefficient-based Top 20 is easier to interpret as the final selected-model effect magnitude. The lambda-entry Top 20 is closer to the LASSO path logic requested by the advisor: variables entering the model earlier under stronger regularization are treated as stronger feature-selection signals.",
        "",
        "For the manuscript, the most defensible strategy is to report both: coefficient-based RI for effect-size interpretation at the selected lambda, and lambda-entry RI as a feature-selection robustness/path-based ranking. Features that appear in both Top 20 lists are the most stable candidates for interpretation and interaction analysis.",
        "",
        "## 為什麼 Coefficient-Based RI 和 Lambda-Entry RI 會差很多？",
        "",
        "這兩種 RI 不是同一種重要性指標，它們回答的是不同問題。",
        "",
        "### 1. Coefficient-Based RI 問的是：在最後選定的 lambda 下，誰的係數比較大？",
        "",
        "Coefficient-Based RI 使用 selected lambda 下的 standardized coefficient：",
        "",
        "```text",
        "Coefficient-Based RI = abs(standardized coefficient) / sum(abs(standardized coefficients)) * 100",
        "```",
        "",
        "所以它比較像是在問：",
        "",
        "> 在最後這個 LASSO 模型裡，哪個變數的最終效果量比較大？",
        "",
        "這個方法容易解釋成 effect-size ranking。若某個變數在最後模型中的係數較大，它的 coefficient-based RI 就會比較高。",
        "",
        "### 2. Lambda-Entry RI 問的是：誰在更強的 LASSO 懲罰下就先被選進模型？",
        "",
        "Lambda-Entry RI 使用 LASSO regularization path：",
        "",
        "```text",
        "Lambda-Entry RI = entry lambda / sum(entry lambda across features in the task) * 100",
        "lambda = 1 / C",
        "```",
        "",
        "這個方法比較像是在問：",
        "",
        "> 當 LASSO 懲罰很強、模型只能留下少數最穩定特徵時，哪些變數會最早變成 non-zero？",
        "",
        "因此 Lambda-Entry RI 不是看最後係數有多大，而是看 feature 在 path 上「多早進場」。越早在強懲罰下被選進來，entry lambda 越大，排名越前面。",
        "",
        "### 3. 為什麼排序會不同？",
        "",
        "排序不同是正常的，因為兩者衡量的不是同一件事：",
        "",
        "| 指標 | 主要看什麼 | 比較適合解釋成什麼 |",
        "|---|---|---|",
        "| Coefficient-Based RI | selected lambda 下的最終係數大小 | 最終模型中的 effect magnitude |",
        "| Lambda-Entry RI | 變數在 LASSO path 中第一次 non-zero 的 lambda | feature-selection strength / path robustness |",
        "",
        "舉例來說，有些變數可能很早就進入模型，但最後係數不大；這種變數在 Lambda-Entry RI 會較前面，但在 Coefficient-Based RI 不一定前面。相反地，有些變數可能比較晚才進入模型，但在 selected lambda 下係數變大；這種變數在 Coefficient-Based RI 會較前面，但在 Lambda-Entry RI 不一定前面。",
        "",
        "### 4. 目前結果的具體差異",
        "",
        "在 `W2 -> W2` 中，兩種 Top20 有 17 個重疊，但也有差異：",
        "",
        "- Coefficient Top20 only：`Social Awareness & Relationship Skills`, `Online Total Nominations`, `Outgoing Friendship Nominations`。",
        "- Lambda Top20 only：`Cyberbullying Perpetration`, `Sent Positive Tie Ratio`, `Covert Social Media Monitoring and Passive Participation`。",
        "",
        "在 `W2 -> W3` 中，兩種 Top20 有 16 個重疊，差異更明顯：",
        "",
        "- Coefficient Top20 only：`Self-Management`, `Online Total Nominations`, `Reciprocal Friendship Ties`, `Motivation & Goal Setting`。",
        "- Lambda Top20 only：`Offline Bullying Victimization`, `Offline Bullying Perpetration`, `Cyberbullying Victimization`, `Perceived Effectiveness of School-based Digital/Technology Learning`。",
        "",
        "這代表有些人際網絡特徵在 selected lambda 下仍有一定係數，因此進入 coefficient-based Top20；但它們不是在最強懲罰階段最早被選進來的特徵，所以在 lambda-entry Top20 中較弱或沒有出現。",
        "",
        "### 5. 建議論文怎麼使用",
        "",
        "建議不要把兩種 RI 混成同一個概念。可以這樣分工：",
        "",
        "- 主要解釋穩定特徵：優先看兩種 Top20 都出現的 variables。",
        "- 解釋最終模型效果大小：看 Coefficient-Based RI。",
        "- 回應老師對 LASSO path / lambda 的要求：看 Lambda-Entry RI。",
        "- 若某個特徵只在 coefficient Top20 出現，要保守說它是在 selected lambda 下有較大係數，但不是 path 上最早進入的穩定特徵。",
        "- 若某個特徵只在 lambda Top20 出現，要保守說它很早被 LASSO path 選進來，但在最後 selected lambda 下的係數大小不一定進入前 20。",
        "",
        "因此目前最適合寫成：",
        "",
        "> Coefficient-based RI was used to describe effect magnitude in the selected LASSO model, whereas lambda-entry RI was used as a path-based feature-selection signal. Predictors appearing in both rankings were treated as the most stable and interpretable features.",
        "",
    ]
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def write_diagnostics(df: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> None:
    payload = {
        "source_xlsx": str(SOURCE_XLSX),
        "main_feature_set": MAIN_FEATURE_SET,
        "top_n": TOP_N,
        "lambda_nonzero_tolerance": LASSO_NONZERO_TOL,
        "lambda_path_note": "Lambda is approximated as 1/C because scikit-learn uses C as inverse regularization strength.",
        "n_source_rows": int(len(df)),
        "tasks": sorted(df["Task"].dropna().unique().tolist()),
        "models": sorted(df["Model"].dropna().unique().tolist()),
        "categories": CATEGORY_ORDER,
        "figures": [str(p) for p in sorted(FIG_DIR.glob("*.png"))],
        "sheets": {name: {"rows": int(len(sheet_df)), "columns": list(sheet_df.columns)} for name, sheet_df in sheets.items()},
    }
    DIAGNOSTICS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    reset_outputs()
    coef_raw = load_coefficients()
    lambda_path, lambda_path_summary = compute_lasso_lambda_path()
    coef = add_lambda_metrics(coef_raw, lambda_path)
    lasso_top20 = top_n(coef, "LASSO Logistic")
    lasso_lambda_top20 = lambda_top_n(coef)
    top20_comparison = compare_coeff_vs_lambda_top20(lasso_top20, lasso_lambda_top20)
    ridge_top20 = top_n(coef, "Ridge Logistic")
    logistic_top20 = top_n(coef, "Multivariable Logistic")
    cat_sum = category_summary(lasso_top20)
    cat_wide = category_pivot(cat_sum)
    cat_sum_lambda = category_summary_lambda(lasso_lambda_top20)
    cat_wide_lambda = category_pivot_lambda(cat_sum_lambda)
    shared = shared_top20(lasso_top20)
    ip_sum = interpersonal_summary(coef)
    figure_paths = [
        save_lasso_top20_bar_chart(lasso_top20, "W2 -> W2"),
        save_lasso_top20_bar_chart(lasso_top20, "W2 -> W3"),
        save_category_summary_chart(cat_wide),
        save_shared_top20_chart(shared),
    ]

    readme = pd.DataFrame(
        [
            {
                "Item": "Primary model",
                "Description": "LASSO Logistic using decomposed + 12 interpersonal features.",
            },
            {
                "Item": "Primary tasks",
                "Description": "W2 -> W2 and W2 -> W3.",
            },
            {
                "Item": "Relative Importance %",
                "Description": "abs(standardized coefficient) divided by the sum of abs(standardized coefficients) within each task/model, multiplied by 100.",
            },
            {
                "Item": "Coefficient RI %",
                "Description": "Same as the original Relative Importance %, retained explicitly as the coefficient-based ranking.",
            },
            {
                "Item": "Lambda Entry RI %",
                "Description": "Entry lambda divided by the sum of entry lambdas within each task, multiplied by 100. Lambda is approximated as 1/C.",
            },
            {
                "Item": "Rank by Lambda Entry",
                "Description": "Ranks features by the strongest regularization level at which they first become non-zero on the LASSO path.",
            },
            {
                "Item": "Category Summary",
                "Description": "Counts and RI sums for LASSO Top 20 features by conceptual domain. Coefficient-based and lambda-entry versions are both provided.",
            },
        ]
    )

    sheets = {
        "ReadMe": readme,
        "LASSO_Top20_Combined": lasso_top20,
        "LASSO_Top20_W2toW2": lasso_top20_by_task(lasso_top20, "W2 -> W2"),
        "LASSO_Top20_W2toW3": lasso_top20_by_task(lasso_top20, "W2 -> W3"),
        "Lambda_Top20_Combined": lasso_lambda_top20,
        "Lambda_Top20_W2toW2": lasso_top20_by_task(lasso_lambda_top20, "W2 -> W2"),
        "Lambda_Top20_W2toW3": lasso_top20_by_task(lasso_lambda_top20, "W2 -> W3"),
        "Coeff_vs_Lambda_Top20": top20_comparison,
        "LambdaPath_AllFeatures": lambda_path,
        "LambdaPath_C_Summary": lambda_path_summary,
        "CategorySummary": cat_sum,
        "CategorySummaryWide": cat_wide,
        "LambdaCategorySummary": cat_sum_lambda,
        "LambdaCategorySummaryWide": cat_wide_lambda,
        "SharedTop20": shared,
        "InterpersonalSummary": ip_sum,
        "Ridge_Top20_Reference": ridge_top20,
        "Logistic_Top20_Reference": logistic_top20,
        "FigureIndex": write_chart_index(figure_paths),
    }
    write_workbook(sheets)
    write_summary(
        lasso_top20,
        lasso_lambda_top20,
        top20_comparison,
        cat_sum,
        cat_wide,
        cat_sum_lambda,
        cat_wide_lambda,
        shared,
        ip_sum,
        figure_paths,
    )
    write_diagnostics(coef, sheets)

    print(f"Wrote {MAIN_XLSX}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Wrote {DIAGNOSTICS_JSON}")
    print("\nCategory summary:")
    print(cat_wide.to_string(index=False))
    print("\nInterpersonal summary:")
    print(ip_sum.to_string(index=False))


if __name__ == "__main__":
    main()
