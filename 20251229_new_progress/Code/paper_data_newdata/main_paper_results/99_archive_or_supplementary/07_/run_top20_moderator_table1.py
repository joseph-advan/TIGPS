from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats


SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
INTERACTION_DIR = PAPER_RESULTS_DIR / "06_interaction_analysis"
if str(INTERACTION_DIR) not in sys.path:
    sys.path.insert(0, str(INTERACTION_DIR))

import run_interaction_analysis as interaction  # noqa: E402


OUT_DIR = SECTION_DIR / "outputs"
DIAG_DIR = OUT_DIR / "diagnostics"
COMBINED_XLSX = OUT_DIR / "top20_moderator_table1_combined.xlsx"
SUMMARY_MD = OUT_DIR / "TOP20_MODERATOR_TABLE1_SUMMARY_ZH.md"
DIAGNOSTICS_JSON = DIAG_DIR / "top20_moderator_table1_diagnostics.json"


TABLE_SPECS = [
    {
        "id": "w2_to_w2_piu",
        "task": "W2 -> W2",
        "sheet": "W2toW2_PIU",
        "file": "table1_w2_to_w2_by_problematic_internet_use_top20.xlsx",
        "moderator_id": "problematic_internet_use",
        "high_label": "High Problematic Internet Use",
        "low_label": "Low Problematic Internet Use",
    },
    {
        "id": "w2_to_w3_piu",
        "task": "W2 -> W3",
        "sheet": "W2toW3_PIU",
        "file": "table1_w2_to_w3_by_problematic_internet_use_top20.xlsx",
        "moderator_id": "problematic_internet_use",
        "high_label": "High Problematic Internet Use",
        "low_label": "Low Problematic Internet Use",
    },
    {
        "id": "w2_to_w2_online_activity",
        "task": "W2 -> W2",
        "sheet": "W2toW2_OnlineActivity",
        "file": "table1_w2_to_w2_by_online_activity_top20.xlsx",
        "moderator_id": "online_activity",
        "high_label": "High Online Activity",
        "low_label": "Low Online Activity",
    },
    {
        "id": "w2_to_w3_online_activity",
        "task": "W2 -> W3",
        "sheet": "W2toW3_OnlineActivity",
        "file": "table1_w2_to_w3_by_online_activity_top20.xlsx",
        "moderator_id": "online_activity",
        "high_label": "High Online Activity",
        "low_label": "Low Online Activity",
    },
]


def format_p(value: float) -> str:
    if pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def format_mean_sd(values: pd.Series) -> str:
    clean = pd.to_numeric(values, errors="coerce").dropna()
    if clean.empty:
        return ""
    return f"{clean.mean():.2f} ({clean.std(ddof=1):.2f})"


def cohen_d(high: pd.Series, low: pd.Series) -> float:
    high = pd.to_numeric(high, errors="coerce").dropna()
    low = pd.to_numeric(low, errors="coerce").dropna()
    if len(high) < 2 or len(low) < 2:
        return np.nan
    var_high = high.var(ddof=1)
    var_low = low.var(ddof=1)
    denom_n = len(high) + len(low) - 2
    if denom_n <= 0:
        return np.nan
    pooled = math.sqrt(((len(high) - 1) * var_high + (len(low) - 1) * var_low) / denom_n)
    if pooled == 0 or pd.isna(pooled):
        return np.nan
    return float((high.mean() - low.mean()) / pooled)


def welch_p(high: pd.Series, low: pd.Series) -> float:
    high = pd.to_numeric(high, errors="coerce").dropna()
    low = pd.to_numeric(low, errors="coerce").dropna()
    if len(high) < 2 or len(low) < 2:
        return np.nan
    result = stats.ttest_ind(high, low, equal_var=False, nan_policy="omit")
    return float(result.pvalue) if not pd.isna(result.pvalue) else np.nan


def is_binary(values: pd.Series) -> bool:
    unique = set(pd.to_numeric(values, errors="coerce").dropna().astype(float).unique())
    return bool(unique) and unique.issubset({0.0, 1.0})


def readme_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Item": "Purpose",
                "Description": "Table 1 style comparison of section 04 LASSO Top20 feature scores by high/low moderator group.",
            },
            {
                "Item": "Moderators",
                "Description": "W2 Online Activity median split; W2 Problematic Internet Use v28 median split.",
            },
            {
                "Item": "Rows",
                "Description": "Only section 04 LASSO Top20 features are included. For PIU grouping, v28 itself is skipped to avoid comparing the grouping variable against itself.",
            },
            {
                "Item": "Statistics",
                "Description": "Continuous and binary model features are summarized as mean (SD). p-values use Welch t-test. Between-group difference uses Cohen's d.",
            },
            {
                "Item": "Direction",
                "Description": "Cohen's d is High group mean minus Low group mean, divided by pooled SD.",
            },
        ]
    )


def build_table(
    *,
    spec: dict[str, str],
    X: pd.DataFrame,
    candidates: pd.DataFrame,
    moderator_binary: pd.Series,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    high_label = spec["high_label"]
    low_label = spec["low_label"]
    task_candidates = candidates[candidates["Task"].eq(spec["task"])].copy()
    rows: list[dict[str, Any]] = [
        {
            "Task": spec["task"],
            "Question ID": np.nan,
            "LASSO Rank": np.nan,
            "Feature Code": np.nan,
            "Category": np.nan,
            "Variable": "N",
            high_label: int(moderator_binary.eq(1).sum()),
            low_label: int(moderator_binary.eq(0).sum()),
            "Total": int(moderator_binary.notna().sum()),
            "High n": int(moderator_binary.eq(1).sum()),
            "Low n": int(moderator_binary.eq(0).sum()),
            "Total valid n": int(moderator_binary.notna().sum()),
            "p-value": np.nan,
            "Between-group difference": np.nan,
            "Between-group difference type": np.nan,
            "Feature value type": np.nan,
            "Items": np.nan,
            "Note": "Moderator complete rows only.",
        }
    ]
    skipped: list[dict[str, Any]] = []

    for _, cand in task_candidates.iterrows():
        code = str(cand["Feature Code"])
        variable = str(cand["Variable"])
        if spec["moderator_id"] == "problematic_internet_use" and code == interaction.PROBLEMATIC_INTERNET_USE_FEATURE_CODE:
            skipped.append(
                {
                    "Task": spec["task"],
                    "Moderator": "Problematic Internet Use",
                    "Feature": variable,
                    "Feature Code": code,
                    "Reason": "v28 defines the high/low PIU group, so it is skipped as a focal Table 1 feature.",
                }
            )
            continue

        model_col = str(cand.get("Model Column", ""))
        if not model_col or model_col not in X.columns:
            skipped.append(
                {
                    "Task": spec["task"],
                    "Moderator": spec["moderator_id"],
                    "Feature": variable,
                    "Feature Code": code,
                    "Reason": f"Missing model column: {model_col}",
                }
            )
            continue

        values = pd.to_numeric(X[model_col], errors="coerce")
        frame = pd.DataFrame({"value": values, "group": moderator_binary}).dropna(subset=["group", "value"])
        high = frame.loc[frame["group"].eq(1), "value"]
        low = frame.loc[frame["group"].eq(0), "value"]
        total = frame["value"]
        p_value = welch_p(high, low)
        d_value = cohen_d(high, low)
        binary = is_binary(total)
        rows.append(
            {
                "Task": spec["task"],
                "Question ID": cand.get("Items", ""),
                "LASSO Rank": cand.get("LASSO Top20 Rank", cand.get("Rank by Abs Std. B", np.nan)),
                "Feature Code": code,
                "Category": cand.get("Category", ""),
                "Variable": f"{variable}, mean (SD)",
                high_label: format_mean_sd(high),
                low_label: format_mean_sd(low),
                "Total": format_mean_sd(total),
                "High n": int(high.notna().sum()),
                "Low n": int(low.notna().sum()),
                "Total valid n": int(total.notna().sum()),
                "p-value": format_p(p_value),
                "Between-group difference": "" if pd.isna(d_value) else round(d_value, 3),
                "Between-group difference type": "Cohen's d",
                "Feature value type": "binary 0/1" if binary else "continuous score",
                "Items": cand.get("Items", ""),
                "Note": "Binary feature is shown as mean proportion (SD)." if binary else "",
            }
        )

    table = pd.DataFrame(rows)
    if len(table) > 1:
        n_row = table.iloc[[0]]
        feature_rows = table.iloc[1:].copy()
        feature_rows["LASSO Rank Sort"] = pd.to_numeric(feature_rows["LASSO Rank"], errors="coerce")
        feature_rows = feature_rows.sort_values(["LASSO Rank Sort", "Feature Code"]).drop(columns=["LASSO Rank Sort"])
        table = pd.concat([n_row, feature_rows], ignore_index=True)
    return table, pd.DataFrame(skipped)


def build_all_tables() -> tuple[dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame]:
    feature_df, _, X, _, _, candidates = interaction.build_base_inputs()
    _, high_online, online_diag = interaction.make_high_online_activity_w2(feature_df)
    _, high_piu, piu_diag = interaction.make_high_problematic_internet_use_w2(X)
    moderators = {
        "online_activity": high_online,
        "problematic_internet_use": high_piu,
    }
    moderator_defs = pd.DataFrame(
        [
            {
                "Moderator": "Online Activity",
                **online_diag,
            },
            {
                "Moderator": "Problematic Internet Use",
                **piu_diag,
            },
        ]
    )

    sheets: dict[str, pd.DataFrame] = {"ReadMe": readme_sheet(), "ModeratorDefinitions": moderator_defs}
    skipped_all: list[pd.DataFrame] = []
    for spec in TABLE_SPECS:
        table, skipped = build_table(
            spec=spec,
            X=X,
            candidates=candidates,
            moderator_binary=moderators[spec["moderator_id"]],
        )
        sheets[spec["sheet"]] = table
        if not skipped.empty:
            skipped_all.append(skipped)

    skipped_df = pd.concat(skipped_all, ignore_index=True) if skipped_all else pd.DataFrame()
    sheets["SkippedFeatures"] = skipped_df
    return sheets, moderator_defs, skipped_df


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 80))
            ws.column_dimensions[letter].width = max(12, max_len + 2)
    wb.save(path)


def write_single_workbooks(sheets: dict[str, pd.DataFrame]) -> list[dict[str, str]]:
    outputs: list[dict[str, str]] = []
    for spec in TABLE_SPECS:
        out_path = OUT_DIR / spec["file"]
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            readme_sheet().to_excel(writer, sheet_name="ReadMe", index=False)
            sheets["ModeratorDefinitions"].to_excel(writer, sheet_name="ModeratorDefinitions", index=False)
            sheets[spec["sheet"]].to_excel(writer, sheet_name="Table1", index=False)
            skipped = sheets["SkippedFeatures"]
            if not skipped.empty:
                skipped[skipped["Task"].eq(spec["task"])].to_excel(writer, sheet_name="SkippedFeatures", index=False)
        format_workbook(out_path)
        outputs.append({"Table": spec["sheet"], "Path": str(out_path)})
    return outputs


def write_summary(sheets: dict[str, pd.DataFrame], single_outputs: list[dict[str, str]]) -> None:
    lines = [
        "# Top20 Moderator Table 1 Summary",
        "",
        "## 目的",
        "",
        "這個 07 分析針對 04 的 LASSO Top20 特徵，建立以 moderator 高低組為分組依據的 Table 1。",
        "",
        "## 四張 Table 1",
        "",
        "- `W2 -> W2`：依 W2 Problematic Internet Use 高低組比較 Top20 特徵。",
        "- `W2 -> W3`：依 W2 Problematic Internet Use 高低組比較 Top20 特徵。",
        "- `W2 -> W2`：依 W2 Online Activity 高低組比較 Top20 特徵。",
        "- `W2 -> W3`：依 W2 Online Activity 高低組比較 Top20 特徵。",
        "",
        "## 統計呈現",
        "",
        "- 每個特徵以 `mean (SD)` 呈現。",
        "- p-value 使用 Welch t-test。",
        "- Between-group difference 使用 Cohen's d，方向是 High group mean - Low group mean。",
        "- Binary 0/1 特徵也以 mean(SD) 呈現，mean 可視為比例。",
        "",
        "## 注意事項",
        "",
        "`v28` 是 Problematic Internet Use 高低組的分組依據，因此在 PIU Table 1 中不作為 focal feature 比較，已記錄在 `SkippedFeatures`。",
        "",
        "## Outputs",
        "",
        f"- Combined workbook: `{COMBINED_XLSX}`",
    ]
    for item in single_outputs:
        lines.append(f"- {item['Table']}: `{item['Path']}`")
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    DIAG_DIR.mkdir(parents=True, exist_ok=True)
    sheets, moderator_defs, skipped_df = build_all_tables()
    with pd.ExcelWriter(COMBINED_XLSX, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(COMBINED_XLSX)
    single_outputs = write_single_workbooks(sheets)
    write_summary(sheets, single_outputs)
    diagnostics = {
        "combined_xlsx": str(COMBINED_XLSX),
        "summary_md": str(SUMMARY_MD),
        "single_outputs": single_outputs,
        "sheets": {name: {"rows": int(len(df)), "columns": list(df.columns)} for name, df in sheets.items()},
        "moderator_definitions": moderator_defs.to_dict(orient="records"),
        "skipped_features": skipped_df.to_dict(orient="records"),
    }
    DIAGNOSTICS_JSON.write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {COMBINED_XLSX}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Wrote {DIAGNOSTICS_JSON}")
    for item in single_outputs:
        print(f"Wrote {item['Path']}")


if __name__ == "__main__":
    main()
