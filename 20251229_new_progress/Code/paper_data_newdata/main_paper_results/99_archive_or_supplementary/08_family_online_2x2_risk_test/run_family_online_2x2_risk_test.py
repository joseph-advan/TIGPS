from __future__ import annotations

import math
import shutil
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
ROOT = SCRIPT_PATH.parents[4]
IX_DIR = PAPER_RESULTS_DIR / "06_interaction_analysis"
if str(IX_DIR) not in sys.path:
    sys.path.insert(0, str(IX_DIR))

import run_interaction_analysis as ix  # noqa: E402

OUT_DIR = SECTION_DIR / "outputs"
OUT_XLSX = OUT_DIR / "family_cohesion_online_activity_2x2_risk_test.xlsx"
SUMMARY_MD = OUT_DIR / "FAMILY_COHESION_ONLINE_ACTIVITY_2X2_RISK_TEST_ZH.md"

FAMILY_CODE = "v5"
FAMILY_COL = "feature_v5"
FAMILY_NAME = "Family Cohesion and Support (Family Functioning)"
REFERENCE_GROUP = "High Family Cohesion + Low Online Activity"


def fmt_p(p: float | None) -> str:
    if p is None or pd.isna(p):
        return ""
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def safe_float(x: Any) -> float:
    if x is None or pd.isna(x):
        return np.nan
    return float(x)


def inverse_logit(x: float) -> float:
    return float(1 / (1 + math.exp(-x)))


def reset_outputs() -> None:
    if OUT_DIR.exists():
        resolved = OUT_DIR.resolve()
        if resolved.parent != SECTION_DIR.resolve():
            raise RuntimeError(f"Refusing to remove unexpected output path: {resolved}")
        shutil.rmtree(OUT_DIR)
    OUT_DIR.mkdir(parents=True, exist_ok=True)


def build_base() -> tuple[pd.DataFrame, dict[str, pd.DataFrame], pd.DataFrame, pd.DataFrame, pd.Series, pd.Series, dict[str, Any]]:
    feature_df, datasets, X, feature_defs, feature_diag, _ = ix.build_base_inputs()
    if FAMILY_COL not in X.columns:
        raise KeyError(f"Missing {FAMILY_COL}; cannot run 2x2 family-online test.")
    online_score, high_online, online_diag = ix.make_high_online_activity_w2(feature_df)
    return feature_df, datasets, X, feature_defs, online_score, high_online, {**feature_diag, **online_diag}


def make_family_group(family_score: pd.Series) -> tuple[pd.Series, dict[str, Any]]:
    score = pd.to_numeric(family_score, errors="coerce")
    median = float(score.median(skipna=True))
    group = pd.Series(np.nan, index=score.index, dtype=object)
    group.loc[score.notna() & score.gt(median)] = "High Family Cohesion"
    group.loc[score.notna() & score.le(median)] = "Low Family Cohesion"
    diag = {
        "family_feature_code": FAMILY_CODE,
        "family_feature_column": FAMILY_COL,
        "family_feature_name": FAMILY_NAME,
        "family_split_definition": "High Family Cohesion = feature_v5 > W2 median; Low Family Cohesion = feature_v5 <= W2 median.",
        "family_median": median,
        "high_family_n": int(group.eq("High Family Cohesion").sum()),
        "low_family_n": int(group.eq("Low Family Cohesion").sum()),
        "missing_family_n": int(group.isna().sum()),
    }
    return group, diag


def make_frame(task: dict[str, Any], feature_df: pd.DataFrame, datasets: dict[str, pd.DataFrame], X: pd.DataFrame, online_score: pd.Series, high_online: pd.Series) -> tuple[pd.DataFrame, dict[str, Any]]:
    y, target_diag = ix.make_target_for_task(task, datasets, feature_df)
    family_score = pd.to_numeric(X[FAMILY_COL], errors="coerce")
    family_group, family_diag = make_family_group(family_score)
    online_group = pd.Series(np.nan, index=feature_df.index, dtype=object)
    online_group.loc[pd.to_numeric(high_online, errors="coerce").eq(1)] = "High Online Activity"
    online_group.loc[pd.to_numeric(high_online, errors="coerce").eq(0)] = "Low Online Activity"
    frame = pd.DataFrame(
        {
            "student_id": feature_df["student_id"],
            "target": pd.to_numeric(y, errors="coerce"),
            "family_score": family_score,
            "family_group": family_group,
            "online_score": pd.to_numeric(online_score, errors="coerce"),
            "online_group": online_group,
            "high_online": pd.to_numeric(high_online, errors="coerce"),
        },
        index=feature_df.index,
    )
    if "feature_v1_male" in X.columns:
        frame["gender_male"] = pd.to_numeric(X["feature_v1_male"], errors="coerce")
    frame = frame.dropna(subset=["target", "family_score", "family_group", "online_group", "high_online"]).copy()
    frame["risk_group"] = frame["family_group"].astype(str) + " + " + frame["online_group"].astype(str)
    frame["low_family"] = frame["family_group"].eq("Low Family Cohesion").astype(int)
    frame["low_family_x_high_online"] = frame["low_family"] * frame["high_online"]
    diag = {"Task": task["Task"], **target_diag, **family_diag, "analysis_n": int(len(frame))}
    return frame, diag


def group_summary(frame: pd.DataFrame, task_name: str) -> pd.DataFrame:
    order = [
        "High Family Cohesion + Low Online Activity",
        "High Family Cohesion + High Online Activity",
        "Low Family Cohesion + Low Online Activity",
        "Low Family Cohesion + High Online Activity",
    ]
    rows = []
    for group in order:
        sub = frame[frame["risk_group"].eq(group)]
        n = len(sub)
        high_n = int(sub["target"].eq(1).sum()) if n else 0
        pct = high_n / n * 100 if n else np.nan
        rows.append(
            {
                "Task": task_name,
                "Risk Group": group,
                "Family Cohesion Group": group.split(" + ")[0],
                "Online Activity Group": group.split(" + ")[1],
                "N": n,
                "High Psychological Distress N": high_n,
                "High Psychological Distress %": pct,
                "Family Cohesion Mean": safe_float(sub["family_score"].mean()),
                "Family Cohesion SD": safe_float(sub["family_score"].std(ddof=1)),
                "Online Activity Mean": safe_float(sub["online_score"].mean()),
                "Online Activity SD": safe_float(sub["online_score"].std(ddof=1)),
            }
        )
    out = pd.DataFrame(rows)
    out["Risk Rank Within Task"] = out["High Psychological Distress %"].rank(method="first", ascending=False).astype("Int64")
    return out.sort_values(["Task", "Risk Rank Within Task", "Risk Group"])


def fit_group_logit(frame: pd.DataFrame, task_name: str, adjusted: bool) -> tuple[pd.DataFrame, pd.DataFrame]:
    model_df = frame.copy()
    terms = []
    for group in sorted(model_df["risk_group"].dropna().unique()):
        if group == REFERENCE_GROUP:
            continue
        col = "group__" + group.replace(" ", "_").replace("+", "plus")
        model_df[col] = model_df["risk_group"].eq(group).astype(int)
        terms.append(col)
    if adjusted and "gender_male" in model_df.columns and model_df["gender_male"].nunique(dropna=True) > 1:
        terms.append("gender_male")
    model_df = model_df.dropna(subset=["target"] + terms).copy()
    fit = sm.Logit(model_df["target"], sm.add_constant(model_df[terms], has_constant="add")).fit(disp=False, maxiter=300)
    rows = []
    for term in ["const"] + terms:
        if term == "const":
            label = REFERENCE_GROUP
            comparison = "Reference intercept"
        elif term == "gender_male":
            label = "Gender: Male (vs Female)"
            comparison = "Gender adjustment covariate"
        else:
            raw = term.removeprefix("group__").replace("plus", "+").replace("_", " ")
            label = raw
            comparison = f"{raw} vs {REFERENCE_GROUP}"
        b = safe_float(fit.params.get(term, np.nan))
        se = safe_float(fit.bse.get(term, np.nan))
        p = safe_float(fit.pvalues.get(term, np.nan))
        rows.append(
            {
                "Task": task_name,
                "Model": "Gender-adjusted group logistic" if adjusted else "Unadjusted group logistic",
                "Term": label,
                "Comparison": comparison,
                "B": b,
                "SE": se,
                "p-value": p,
                "p-value formatted": fmt_p(p),
                "OR": math.exp(b) if not pd.isna(b) else np.nan,
                "Reference Group": REFERENCE_GROUP,
                "N": int(len(model_df)),
            }
        )
    pred_rows = []
    for group in [REFERENCE_GROUP] + [r["Term"] for r in rows if r["Term"] not in {REFERENCE_GROUP, "Gender: Male (vs Female)"}]:
        linear = safe_float(fit.params.get("const", 0.0))
        if group != REFERENCE_GROUP:
            term = "group__" + group.replace(" ", "_").replace("+", "plus")
            linear += safe_float(fit.params.get(term, 0.0))
        # Set gender_male to 0 for adjusted predictions to keep group comparisons on the same baseline.
        pred_rows.append(
            {
                "Task": task_name,
                "Model": "Gender-adjusted group logistic" if adjusted else "Unadjusted group logistic",
                "Risk Group": group,
                "Predicted Logit": linear,
                "Predicted Probability": inverse_logit(linear),
                "Prediction Note": "gender_male fixed at 0 when adjustment is included" if adjusted else "unadjusted group prediction",
            }
        )
    return pd.DataFrame(rows), pd.DataFrame(pred_rows)


def fit_teacher_equivalent(frame: pd.DataFrame, task_name: str) -> pd.DataFrame:
    terms = ["low_family", "high_online", "low_family_x_high_online"]
    if "gender_male" in frame.columns and frame["gender_male"].nunique(dropna=True) > 1:
        terms.append("gender_male")
    model_df = frame.dropna(subset=["target"] + terms).copy()
    fit = sm.Logit(model_df["target"], sm.add_constant(model_df[terms], has_constant="add")).fit(disp=False, maxiter=300)
    mapping = {
        "const": "b0: High Family Cohesion + Low Online Activity intercept",
        "low_family": "b1: Low Family Cohesion difference when Online Activity is low",
        "high_online": "b2: High Online Activity difference when Family Cohesion is high",
        "low_family_x_high_online": "b3: extra difference for Low Family Cohesion + High Online Activity",
        "gender_male": "Gender adjustment covariate",
    }
    rows = []
    for term in ["const"] + terms:
        b = safe_float(fit.params.get(term, np.nan))
        rows.append(
            {
                "Task": task_name,
                "Term": term,
                "Meaning": mapping.get(term, term),
                "B": b,
                "SE": safe_float(fit.bse.get(term, np.nan)),
                "p-value": safe_float(fit.pvalues.get(term, np.nan)),
                "p-value formatted": fmt_p(safe_float(fit.pvalues.get(term, np.nan))),
                "OR": math.exp(b) if not pd.isna(b) else np.nan,
                "N": int(len(model_df)),
            }
        )
    return pd.DataFrame(rows)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    wb = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 80) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)
    wb.save(OUT_XLSX)


def write_summary(group_df: pd.DataFrame, teacher_df: pd.DataFrame, diag_df: pd.DataFrame) -> None:
    primary = group_df[group_df["Task"].eq("W2 -> W3")].sort_values("Risk Rank Within Task")
    top = primary.iloc[0]
    low_high = primary[primary["Risk Group"].eq("Low Family Cohesion + High Online Activity")].iloc[0]
    b3 = teacher_df[(teacher_df["Task"].eq("W2 -> W3")) & (teacher_df["Term"].eq("low_family_x_high_online"))].iloc[0]
    lines = [
        "# Family Cohesion x Online Activity 2x2 Risk Test",
        "",
        "## 這個測試在回答什麼",
        "",
        "這個 08 測試不是取代 06 的交互作用模型，而是把 06 中最重要的 Family Cohesion x Online Activity 結果改成更直覺的 2x2 組別比較。",
        "",
        "分組方式：",
        "",
        "- Family Cohesion：使用 W2 `v5` 題組分數，以 W2 中位數切成 High / Low。",
        "- Online Activity：使用 W2 `v21_3` 到 `v21_6` 加總分數，以 W2 中位數切成 High / Low。",
        "- Outcome：High Psychological Distress，使用各任務的心理困擾題組加總後以中位數切分。",
        "",
        "四組為：",
        "",
        "1. High Family Cohesion + Low Online Activity",
        "2. High Family Cohesion + High Online Activity",
        "3. Low Family Cohesion + Low Online Activity",
        "4. Low Family Cohesion + High Online Activity",
        "",
        "## 主要結果：W2 -> W3",
        "",
        f"W2 -> W3 中，高心理困擾比例最高的組別是 `{top['Risk Group']}`，比例為 {top['High Psychological Distress %']:.1f}%（n={int(top['N'])}）。",
        f"你關心的 `Low Family Cohesion + High Online Activity` 組別，高心理困擾比例為 {low_high['High Psychological Distress %']:.1f}%（n={int(low_high['N'])}），風險排名第 {int(low_high['Risk Rank Within Task'])}。",
        "",
        "## 2x2 interaction 係數怎麼看",
        "",
        "這裡也用老師給的公式概念重新估計一次，但把 Family Cohesion 也切成 High / Low：",
        "",
        "```text",
        "logit(P(High Psychological Distress = 1))",
        "= b0 + b1 * LowFamily + b2 * HighOnline + b3 * LowFamily * HighOnline",
        "```",
        "",
        "在這個設定中：",
        "",
        "- `b0`：High Family Cohesion + Low Online Activity 的 baseline logit。",
        "- `b1`：在 Low Online Activity 裡，Low Family 比 High Family 多出的差異。",
        "- `b2`：在 High Family Cohesion 裡，High Online 比 Low Online 多出的差異。",
        "- `b3`：Low Family + High Online 這個組合是否有額外加乘風險。",
        "",
        f"W2 -> W3 的 `b3` = {b3['B']:.4f}，p = {b3['p-value formatted']}。",
        "",
    ]
    if b3["p-value"] < 0.05:
        lines.append("這代表 `Low Family Cohesion + High Online Activity` 有統計上顯著的額外組合效果。")
    else:
        lines.append("這代表目前沒有足夠證據說 `Low Family Cohesion + High Online Activity` 有超過兩個主效果相加之外的額外加乘風險。")
    lines += [
        "",
        "## 可以怎麼寫",
        "",
        "較保守、符合目前統計結果的寫法：",
        "",
        "> We further examined whether students with lower family cohesion and high online activity constituted a higher-risk subgroup. A 2x2 group comparison was conducted using median splits of W2 family cohesion and W2 online activity. This analysis provides an intuitive subgroup-level description of future psychological distress risk, complementing the continuous interaction model in Section 06.",
        "",
        "中文解釋：這個測試可以用來說明不同家庭支持與網路活躍組合下，未來高心理困擾比例是否不同。但如果 `b3` 不顯著，就不要說明確存在加乘反效果；可以說是描述性風險分層。",
        "",
        "## 輸出檔案",
        "",
        f"- `{OUT_XLSX.name}`：包含 group summary、group logistic、2x2 teacher-equivalent interaction、diagnostics。",
        "",
        "## Diagnostics",
        "",
    ]
    for _, row in diag_df.iterrows():
        lines.append(f"- {row['Metric']}: {row['Value']}")
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    reset_outputs()
    feature_df, datasets, X, feature_defs, online_score, high_online, base_diag = build_base()
    group_rows = []
    group_model_rows = []
    pred_rows = []
    teacher_rows = []
    diag_rows = []
    for task in ix.TASKS:
        frame, diag = make_frame(task, feature_df, datasets, X, online_score, high_online)
        group_rows.append(group_summary(frame, task["Task"]))
        for adjusted in [False, True]:
            m, p = fit_group_logit(frame, task["Task"], adjusted=adjusted)
            group_model_rows.append(m)
            pred_rows.append(p)
        teacher_rows.append(fit_teacher_equivalent(frame, task["Task"]))
        for k, v in {**base_diag, **diag}.items():
            if isinstance(v, (list, dict)):
                v = str(v)
            diag_rows.append({"Task": task["Task"], "Metric": k, "Value": v})
    group_df = pd.concat(group_rows, ignore_index=True)
    model_df = pd.concat(group_model_rows, ignore_index=True)
    pred_df = pd.concat(pred_rows, ignore_index=True)
    teacher_df = pd.concat(teacher_rows, ignore_index=True)
    diag_df = pd.DataFrame(diag_rows).drop_duplicates()
    readme = pd.DataFrame(
        [
            {"Item": "Purpose", "Description": "2x2 subgroup test of Family Cohesion high/low by Online Activity high/low."},
            {"Item": "Family split", "Description": "W2 feature_v5 > median = High Family Cohesion; <= median = Low Family Cohesion."},
            {"Item": "Online split", "Description": "sum W2 v21_3-v21_6 > median = High Online Activity; <= median = Low Online Activity."},
            {"Item": "Primary interpretation", "Description": "Use W2 -> W3 to discuss future psychological distress risk subgroup patterns."},
            {"Item": "Caution", "Description": "Low Family Cohesion means lower-than-or-equal-to-median family cohesion, not absence of family support."},
        ]
    )
    write_workbook(
        {
            "ReadMe": readme,
            "GroupSummary": group_df,
            "GroupLogistic": model_df,
            "PredictedProbabilities": pred_df,
            "TeacherEquivalent2x2": teacher_df,
            "Diagnostics": diag_df,
        }
    )
    write_summary(group_df, teacher_df, diag_df)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {SUMMARY_MD}")


if __name__ == "__main__":
    main()
