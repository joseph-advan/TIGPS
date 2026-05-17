from __future__ import annotations

import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
ROOT = SCRIPT_PATH.parents[5]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"
SOURCE_SCRIPT = CODE_DIR / "Ridge_lasso" / "run_ridge_lasso_shap_three_scenarios.py"
SOURCE_OUT = CODE_DIR / "Ridge_lasso" / "outputs" / "model_results"
OUT_DIR = SCRIPT_PATH.parent / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def main() -> None:
    subprocess.run([sys.executable, str(SOURCE_SCRIPT)], check=True)
    for src in SOURCE_OUT.glob("ridge_lasso_three_scenarios_*"):
        if src.is_file():
            shutil.copy2(src, OUT_DIR / src.name)

    summary = pd.read_csv(SOURCE_OUT / "ridge_lasso_three_scenarios_summary.csv", encoding="utf-8-sig")
    rel = pd.read_csv(SOURCE_OUT / "ridge_lasso_three_scenarios_relative_importance.csv", encoding="utf-8-sig")
    shap = pd.read_csv(SOURCE_OUT / "ridge_lasso_three_scenarios_shap_importance.csv", encoding="utf-8-sig")
    keep = ["w2_predict_w2", "w2_predict_w3"]
    summary = summary[summary["scenario"].isin(keep)].copy()
    rel = rel[rel["scenario"].isin(keep)].copy()
    shap = shap[shap["scenario"].isin(keep)].copy()
    summary["Task"] = summary["scenario"].map({"w2_predict_w2": "W2 -> W2", "w2_predict_w3": "W2 -> W3"})
    xlsx = OUT_DIR / "ridge_lasso_regularized_performance.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Performance")
        rel.to_excel(writer, index=False, sheet_name="RelativeImportance")
        shap.to_excel(writer, index=False, sheet_name="SHAPImportance")
    style_xlsx(xlsx)
    (OUT_DIR / "RIDGE_LASSO_REGULARIZED_SUMMARY.md").write_text(
        "# Ridge / LASSO Regularized Models\n\n"
        "This wrapper reruns the existing Ridge/LASSO drop + decomposition script and copies the W2 -> W2 and W2 -> W3 outputs into this model-performance folder.\n\n"
        f"Source script: `{SOURCE_SCRIPT}`\n\n"
        f"Output workbook: `{xlsx}`\n",
        encoding="utf-8",
    )
    print(f"Wrote {xlsx}")


if __name__ == "__main__":
    main()
