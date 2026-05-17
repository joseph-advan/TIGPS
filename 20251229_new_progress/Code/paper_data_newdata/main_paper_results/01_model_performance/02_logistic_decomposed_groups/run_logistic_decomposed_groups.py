from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
ROOT = SCRIPT_PATH.parents[5]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"
TABLES_SCRIPT_DIR = CODE_DIR / "tables" / "scripts"
if str(TABLES_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(TABLES_SCRIPT_DIR))

import build_table2_table3_drop_decomposition as t23  # noqa: E402

OUT_DIR = SCRIPT_PATH.parent / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TASKS = [
    {"Task": "W2 -> W2", "Scenario": "w2_predict_w2", "Feature Wave": "W2", "Target Wave": "W2", "Feature Target Group": "v55", "Target Items": t23.TARGET_W2_ITEMS},
    {"Task": "W2 -> W3", "Scenario": "w2_predict_w3", "Feature Wave": "W2", "Target Wave": "W3", "Feature Target Group": "54", "Target Items": t23.TARGET_W3_ITEMS},
]


def write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def make_target(task: dict[str, Any], datasets: dict[str, pd.DataFrame], feature_df: pd.DataFrame) -> pd.Series:
    target_df = datasets[task["Target Wave"]]
    _, y_raw, _ = t23.make_target(target_df, task["Target Items"])
    target_map = pd.DataFrame({"student_id": target_df["student_id"], "target": y_raw}).drop_duplicates("student_id", keep="first")
    y = feature_df[["student_id"]].merge(target_map, on="student_id", how="left")["target"]
    y.index = feature_df.index
    return y


def main() -> None:
    merged_path = t23.core.pick_first_existing_path(t23.core.MERGED_PATH_CANDIDATES)
    merged = pd.read_csv(merged_path, dtype=str, encoding="utf-8-sig")
    for c in ["Year", "Group_ID", "Question_ID"]:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).str.strip()
    datasets = {
        "W2": t23.core.normalize_student_id(pd.read_csv(t23.W2_DATA, encoding="utf-8-sig", low_memory=False)),
        "W3": t23.core.normalize_student_id(pd.read_csv(t23.W3_DATA, encoding="utf-8-sig", low_memory=False)),
    }

    perf_rows = []
    coef_rows = []
    diag_rows = []
    for task in TASKS:
        feature_df = datasets[task["Feature Wave"]]
        y = make_target(task, datasets, feature_df)
        X, defs, diag = t23.build_drop_decomposition_features(feature_df, merged, task["Feature Wave"], task["Feature Target Group"])
        coef, perf = t23.fit_model_comparison(y, X, defs)
        row = perf[perf["Model"].eq("Multivariable Logistic")].copy()
        row.insert(0, "Task", task["Task"])
        row.insert(1, "Scenario", task["Scenario"])
        row.insert(2, "Feature Set", "drop_plus_decomposition")
        row.insert(3, "Feature Wave", task["Feature Wave"])
        row.insert(4, "Target Wave", task["Target Wave"])
        perf_rows.append(row)
        csub = coef[["Model Column", "Variable", "Feature Code", "Source Type", "Items", "Multivariable Logistic Std. B"]].copy()
        csub.insert(0, "Task", task["Task"])
        csub.insert(1, "Scenario", task["Scenario"])
        coef_rows.append(csub)
        diag_rows.append({"Task": task["Task"], "Scenario": task["Scenario"], **diag, "n_final_features": int(X.shape[1]), "target_non_missing": int(y.notna().sum())})

    perf_all = pd.concat(perf_rows, ignore_index=True)
    coef_all = pd.concat(coef_rows, ignore_index=True)
    diag_df = pd.DataFrame(diag_rows)
    xlsx = OUT_DIR / "logistic_decomposed_groups_performance.xlsx"
    write_xlsx(xlsx, {"Performance": perf_all, "Coefficients": coef_all, "Diagnostics": diag_df})
    perf_all.to_csv(OUT_DIR / "logistic_decomposed_groups_performance.csv", index=False, encoding="utf-8-sig")
    md = OUT_DIR / "LOGISTIC_DECOMPOSED_GROUPS_SUMMARY.md"
    md.write_text(
        "# Decomposed Logistic Baseline\n\n"
        "Plain multivariable Logistic regression using the current drop + decomposition feature set.\n\n"
        "Tasks: W2 -> W2 and W2 -> W3.\n\n"
        f"Output workbook: `{xlsx}`\n",
        encoding="utf-8",
    )
    print(f"Wrote {xlsx}")


if __name__ == "__main__":
    main()
