from __future__ import annotations

import shutil
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
ROOT = SCRIPT_PATH.parents[5]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"
SOURCE_SCRIPT = CODE_DIR / "GNN_baseline" / "run_graphsage_three_tasks.py"
SOURCE_MODEL_OUT = CODE_DIR / "GNN_baseline" / "outputs" / "model_results"
SOURCE_DIAG_OUT = CODE_DIR / "GNN_baseline" / "outputs" / "diagnostics"
OUT_DIR = SCRIPT_PATH.parent / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def main() -> None:
    subprocess.run([sys.executable, str(SOURCE_SCRIPT)], check=True)
    for src in SOURCE_MODEL_OUT.glob("graphsage_three_tasks_*"):
        if src.is_file():
            shutil.copy2(src, OUT_DIR / src.name)
    diag = SOURCE_DIAG_OUT / "graphsage_three_tasks_diagnostics.json"
    if diag.exists():
        shutil.copy2(diag, OUT_DIR / diag.name)

    summary = pd.read_csv(SOURCE_MODEL_OUT / "graphsage_three_tasks_summary.csv", encoding="utf-8-sig")
    seeds = pd.read_csv(SOURCE_MODEL_OUT / "graphsage_three_tasks_seed_metrics.csv", encoding="utf-8-sig")
    keep = ["w2_self", "w2_predict_w3"]
    summary = summary[summary["scenario"].isin(keep)].copy()
    seeds = seeds[seeds["scenario"].isin(keep)].copy()
    summary["Task"] = summary["scenario"].map({"w2_self": "W2 -> W2", "w2_predict_w3": "W2 -> W3"})
    xlsx = OUT_DIR / "graphsage_gnn_performance.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        seeds.to_excel(writer, index=False, sheet_name="SeedMetrics")
    style_xlsx(xlsx)
    (OUT_DIR / "GRAPHSAGE_GNN_SUMMARY.md").write_text(
        "# GraphSAGE GNN Performance\n\n"
        "This wrapper reruns the existing GraphSAGE three-task script and copies the W2 -> W2 and W2 -> W3 outputs into this model-performance folder.\n\n"
        "Version rule: GraphSAGE is run only with the current drop + decomposition node features. No no-drop GNN version is run or reported.\n\n"
        "GraphSAGE metrics are means over 5 random seeds, not CV5 folds.\n\n"
        f"Source script: `{SOURCE_SCRIPT}`\n\n"
        f"Output workbook: `{xlsx}`\n",
        encoding="utf-8",
    )
    print(f"Wrote {xlsx}")


if __name__ == "__main__":
    main()
