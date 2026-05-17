from __future__ import annotations

import importlib.util
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
ROOT = SCRIPT_PATH.parents[5]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"
SOURCE_SCRIPT = CODE_DIR / "logistic_baseline" / "build_logistic_median_split_combined_with_precision_recall.py"
OUT_DIR = SCRIPT_PATH.parent / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def load_source_module():
    spec = importlib.util.spec_from_file_location("logistic_original_source", SOURCE_SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {SOURCE_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def main() -> None:
    src = load_source_module()
    merged_path = src.pick_first_existing_path(src.MERGED_PATH_CANDIDATES)
    w2_raw = src.normalize_student_id(pd.read_csv(src.W2_DATA_PATH, low_memory=False))
    w3_raw = src.normalize_student_id(pd.read_csv(src.W3_DATA_PATH, low_memory=False))
    merged = pd.read_csv(merged_path, dtype=str)
    for c in ["Year", "Group_ID", "Question_ID"]:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).str.strip()

    no_drop_df, no_drop_details = src.evaluate_version(
        version_name="original_groups_no_drop",
        merged=merged,
        w2_raw=w2_raw,
        w3_raw=w3_raw,
        w2_feature_groups=src.W2_FEATURE_GROUP_IDS,
        w3_feature_groups=src.W3_FEATURE_GROUP_IDS,
    )
    drop_df, drop_details = src.evaluate_version(
        version_name="original_groups_drop_groups",
        merged=merged,
        w2_raw=w2_raw,
        w3_raw=w3_raw,
        w2_feature_groups=src.W2_FEATURE_GROUP_IDS_DROP,
        w3_feature_groups=src.W3_FEATURE_GROUP_IDS_DROP,
    )
    delta_df = src.build_delta_table(no_drop_df, drop_df)
    drop_item_df = src.build_drop_item_table(
        merged,
        df_by_year={"W2": w2_raw, "W3": w3_raw},
        drop_by_year={"W2": src.W2_DROP_GROUP_IDS, "W3": src.W3_DROP_GROUP_IDS},
    )

    keep = ["w2_self", "w2_predict_w3"]
    no_drop_df = no_drop_df[no_drop_df["scenario"].isin(keep)].copy()
    drop_df = drop_df[drop_df["scenario"].isin(keep)].copy()
    delta_df = delta_df[delta_df["scenario"].isin(keep)].copy()

    for df, version in [(no_drop_df, "original_groups_no_drop"), (drop_df, "original_groups_drop_groups")]:
        df.insert(0, "Feature Set", version)
        df.insert(0, "Model", "Original-group Logistic")
        df["Task"] = df["scenario"].map({"w2_self": "W2 -> W2", "w2_predict_w3": "W2 -> W3"})

    combined = pd.concat([no_drop_df, drop_df], ignore_index=True)
    xlsx = OUT_DIR / "logistic_original_groups_performance.xlsx"
    write_xlsx(
        xlsx,
        {
            "Combined": combined,
            "NoDrop": no_drop_df,
            "DropGroups": drop_df,
            "DropMinusNoDrop": delta_df,
            "DroppedItems": drop_item_df,
        },
    )
    combined.to_csv(OUT_DIR / "logistic_original_groups_performance.csv", index=False, encoding="utf-8-sig")
    md = OUT_DIR / "LOGISTIC_ORIGINAL_GROUPS_SUMMARY.md"
    md.write_text(
        "# Original-group Logistic Baseline\n\n"
        "This output reruns the existing non-decomposed/original-group Logistic baseline for W2 -> W2 and W2 -> W3.\n\n"
        "Two versions are included: original groups without dropping configured groups, and original groups with the current drop-groups rule.\n\n"
        f"Source script: `{SOURCE_SCRIPT}`\n\n"
        f"Output workbook: `{xlsx}`\n",
        encoding="utf-8",
    )
    print(f"Wrote {xlsx}")


if __name__ == "__main__":
    main()
