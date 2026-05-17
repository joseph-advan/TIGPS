from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
MODEL_DIR = SCRIPT_PATH.parents[1]
OUT_DIR = SCRIPT_PATH.parent / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PATHS = {
    "original": MODEL_DIR / "01_logistic_original_groups" / "outputs" / "logistic_original_groups_performance.xlsx",
    "decomposed": MODEL_DIR / "02_logistic_decomposed_groups" / "outputs" / "logistic_decomposed_groups_performance.xlsx",
    "ridge_lasso": MODEL_DIR / "03_ridge_lasso_regularized" / "outputs" / "ridge_lasso_regularized_performance.xlsx",
    "graphsage": MODEL_DIR / "04_graphsage_gnn" / "outputs" / "graphsage_gnn_performance.xlsx",
}


def num(value: Any) -> float:
    return float(pd.to_numeric(value, errors="coerce")) if pd.notna(pd.to_numeric(value, errors="coerce")) else np.nan


def rows_original() -> list[dict[str, Any]]:
    df = pd.read_excel(PATHS["original"], sheet_name="Combined")
    rows = []
    for _, r in df.iterrows():
        rows.append(
            {
                "Task": r["Task"],
                "Model Family": "Logistic baseline",
                "Model": "Original-group Logistic",
                "Feature Set": r["Feature Set"],
                "Feature Set Detail": "Original/non-decomposed questionnaire groups",
                "Metric Basis": "CV5 mean",
                "N": r["n_rows_modeling"],
                "N Features": r["n_features_used"],
                "AUC": r["cv5_auc_mean"],
                "Accuracy": r["cv5_accuracy_mean"],
                "F1": r["cv5_f1_mean"],
                "Precision": r["cv5_precision_mean"],
                "Recall": r["cv5_recall_mean"],
                "Test AUC": r["test_auc"],
                "Test Accuracy": r["test_accuracy"],
                "Test F1": r["test_f1"],
                "Scenario": r["scenario"],
                "Source File": str(PATHS["original"]),
            }
        )
    return rows


def rows_decomposed() -> list[dict[str, Any]]:
    df = pd.read_excel(PATHS["decomposed"], sheet_name="Performance")
    rows = []
    for _, r in df.iterrows():
        rows.append(
            {
                "Task": r["Task"],
                "Model Family": "Logistic baseline",
                "Model": "Decomposed Logistic",
                "Feature Set": r["Feature Set"],
                "Feature Set Detail": "Drop + decomposition features",
                "Metric Basis": "CV5 mean",
                "N": r["N"],
                "N Features": r["N features"],
                "AUC": r["CV auc mean"],
                "Accuracy": r["CV accuracy mean"],
                "F1": r["CV f1 mean"],
                "Precision": r["CV precision mean"],
                "Recall": r["CV recall mean"],
                "Test AUC": r["Test AUC"],
                "Test Accuracy": r["Test Accuracy"],
                "Test F1": r["Test F1"],
                "Scenario": r["Scenario"],
                "Source File": str(PATHS["decomposed"]),
            }
        )
    return rows


def rows_ridge_lasso() -> list[dict[str, Any]]:
    df = pd.read_excel(PATHS["ridge_lasso"], sheet_name="Performance")
    rows = []
    for _, r in df.iterrows():
        model = "LASSO Logistic" if str(r["model_type"]).lower() == "lasso" else "Ridge Logistic"
        rows.append(
            {
                "Task": r["Task"],
                "Model Family": "Regularized logistic",
                "Model": model,
                "Feature Set": "drop_plus_decomposition",
                "Feature Set Detail": "Drop + decomposition features",
                "Metric Basis": "CV5 mean",
                "N": r["n_rows_modeling"],
                "N Features": r["n_features_used"],
                "AUC": r["cv5_auc_mean"],
                "Accuracy": r["cv5_accuracy_mean"],
                "F1": r["cv5_f1_mean"],
                "Precision": r["cv5_precision_mean"],
                "Recall": r["cv5_recall_mean"],
                "Test AUC": r["test_auc"],
                "Test Accuracy": r["test_accuracy"],
                "Test F1": r["test_f1"],
                "Scenario": r["scenario"],
                "Source File": str(PATHS["ridge_lasso"]),
            }
        )
    return rows


def rows_graphsage() -> list[dict[str, Any]]:
    df = pd.read_excel(PATHS["graphsage"], sheet_name="Summary")
    rows = []
    for _, r in df.iterrows():
        rows.append(
            {
                "Task": r["Task"],
                "Model Family": "GNN",
                "Model": "GraphSAGE",
                "Feature Set": "drop_plus_decomposition_plus_graph_edges",
                "Feature Set Detail": "Drop + decomposition node features with W2 peer nomination graph edges",
                "Metric Basis": "5-seed heldout test mean",
                "N": r["n_nodes_modeling"],
                "N Features": r["n_features_used"],
                "AUC": r["test_auc_mean"],
                "Accuracy": r["test_accuracy_mean"],
                "F1": r["test_f1_mean"],
                "Precision": r["test_precision_mean"],
                "Recall": r["test_recall_mean"],
                "Test AUC": r["test_auc_mean"],
                "Test Accuracy": r["test_accuracy_mean"],
                "Test F1": r["test_f1_mean"],
                "Scenario": r["scenario"],
                "Source File": str(PATHS["graphsage"]),
            }
        )
    return rows


def build_wide(main: pd.DataFrame) -> pd.DataFrame:
    cols = ["AUC", "Accuracy", "F1", "Precision", "Recall"]
    chunks = []
    for task, sub in main.groupby("Task", sort=False):
        wide = sub.pivot_table(index=["Task"], columns="Model", values=cols, aggfunc="first")
        wide.columns = [f"{metric} - {model}" for metric, model in wide.columns]
        chunks.append(wide.reset_index())
    return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()


def build_deltas(main: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for task, sub in main.groupby("Task", sort=False):
        gnn = sub[sub["Model"].eq("GraphSAGE")]
        if gnn.empty:
            continue
        gnn_auc = num(gnn.iloc[0]["AUC"])
        gnn_f1 = num(gnn.iloc[0]["F1"])
        non_gnn = sub[~sub["Model"].eq("GraphSAGE")].copy()
        best_auc_row = non_gnn.loc[pd.to_numeric(non_gnn["AUC"], errors="coerce").idxmax()]
        best_f1_row = non_gnn.loc[pd.to_numeric(non_gnn["F1"], errors="coerce").idxmax()]
        decomp = sub[sub["Model"].eq("Decomposed Logistic")]
        rows.append(
            {
                "Task": task,
                "GraphSAGE AUC": gnn_auc,
                "Best non-GNN AUC": num(best_auc_row["AUC"]),
                "Best non-GNN AUC model": best_auc_row["Model"],
                "GraphSAGE minus best non-GNN AUC": gnn_auc - num(best_auc_row["AUC"]),
                "GraphSAGE F1": gnn_f1,
                "Best non-GNN F1": num(best_f1_row["F1"]),
                "Best non-GNN F1 model": best_f1_row["Model"],
                "GraphSAGE minus best non-GNN F1": gnn_f1 - num(best_f1_row["F1"]),
                "Decomposed Logistic AUC": num(decomp.iloc[0]["AUC"]) if not decomp.empty else np.nan,
                "GraphSAGE minus Decomposed Logistic AUC": gnn_auc - num(decomp.iloc[0]["AUC"]) if not decomp.empty else np.nan,
            }
        )
    return pd.DataFrame(rows)


def add_accuracy_rank(main: pd.DataFrame) -> pd.DataFrame:
    """Sort models within each prediction task by descending Accuracy."""
    ranked = main.copy()
    ranked["_task_order"] = ranked["Task"].map({"W2 -> W2": 0, "W2 -> W3": 1}).fillna(99)
    ranked["_accuracy_numeric"] = pd.to_numeric(ranked["Accuracy"], errors="coerce")
    ranked = ranked.sort_values(
        ["_task_order", "_accuracy_numeric", "Model", "Feature Set"],
        ascending=[True, False, True, True],
        kind="stable",
    )
    ranked["Accuracy Rank Within Task"] = (
        ranked.groupby("Task")["_accuracy_numeric"].rank(method="first", ascending=False).astype(int)
    )
    ranked = ranked.drop(columns=["_task_order", "_accuracy_numeric"])
    cols = list(ranked.columns)
    cols.remove("Accuracy Rank Within Task")
    insert_at = cols.index("Model") + 1
    cols = cols[:insert_at] + ["Accuracy Rank Within Task"] + cols[insert_at:]
    return ranked[cols]


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)
    wb.save(path)


def main() -> None:
    missing = [str(p) for p in PATHS.values() if not p.exists()]
    if missing:
        raise FileNotFoundError("Missing model outputs:\n" + "\n".join(missing))
    main_df = pd.DataFrame(rows_original() + rows_decomposed() + rows_ridge_lasso() + rows_graphsage())
    order = {"W2 -> W2": 0, "W2 -> W3": 1}
    model_order = {
        "Original-group Logistic": 0,
        "Decomposed Logistic": 1,
        "LASSO Logistic": 2,
        "Ridge Logistic": 3,
        "GraphSAGE": 4,
    }
    main_df["_task_order"] = main_df["Task"].map(order)
    main_df["_model_order"] = main_df["Model"].map(model_order)
    main_df = main_df.sort_values(["_task_order", "_model_order", "Feature Set"]).drop(columns=["_task_order", "_model_order"])
    ranked_df = add_accuracy_rank(main_df)
    w2_ranked = ranked_df[ranked_df["Task"].eq("W2 -> W2")].copy()
    w3_ranked = ranked_df[ranked_df["Task"].eq("W2 -> W3")].copy()
    wide = build_wide(main_df)
    deltas = build_deltas(main_df)
    source_index = pd.DataFrame([
        {"Model group": "Original-group Logistic", "Folder": str(MODEL_DIR / "01_logistic_original_groups"), "Output": str(PATHS["original"])},
        {"Model group": "Decomposed Logistic", "Folder": str(MODEL_DIR / "02_logistic_decomposed_groups"), "Output": str(PATHS["decomposed"])},
        {"Model group": "Ridge/LASSO", "Folder": str(MODEL_DIR / "03_ridge_lasso_regularized"), "Output": str(PATHS["ridge_lasso"])},
        {"Model group": "GraphSAGE", "Folder": str(MODEL_DIR / "04_graphsage_gnn"), "Output": str(PATHS["graphsage"])},
    ])

    xlsx = OUT_DIR / "model_comparison_all_w2w2_w2w3.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        main_df.to_excel(writer, index=False, sheet_name="MainComparison")
        ranked_df.to_excel(writer, index=False, sheet_name="RankedByAccuracy")
        w2_ranked.to_excel(writer, index=False, sheet_name="W2_to_W2_Ranked")
        w3_ranked.to_excel(writer, index=False, sheet_name="W2_to_W3_Ranked")
        wide.to_excel(writer, index=False, sheet_name="WideComparison")
        deltas.to_excel(writer, index=False, sheet_name="KeyDeltas")
        source_index.to_excel(writer, index=False, sheet_name="SourceIndex")
    style_xlsx(xlsx)
    ranked_df.to_csv(OUT_DIR / "model_comparison_all_w2w2_w2w3.csv", index=False, encoding="utf-8-sig")

    md = OUT_DIR / "MODEL_COMPARISON_ALL_SUMMARY.md"
    show_delta = deltas.copy()
    for c in show_delta.columns:
        if pd.api.types.is_numeric_dtype(show_delta[c]):
            show_delta[c] = show_delta[c].map(lambda v: "" if pd.isna(v) else f"{float(v):.4f}")
    md.write_text(
        "# All Model Performance Comparison\n\n"
        "This folder compares all model-performance outputs generated under `01_model_performance`.\n\n"
        "## Included Models\n\n"
        "- Original-group Logistic baseline: non-decomposed/original questionnaire groups. This is the only place where no-drop is intentionally reported.\n"
        "- Decomposed Logistic baseline: drop + decomposition features.\n"
        "- LASSO Logistic and Ridge Logistic: regularized models with drop + decomposition features.\n"
        "- GraphSAGE: GNN using drop + decomposition node features and peer nomination graph edges. No no-drop GNN version is run or reported.\n\n"
        "## Important Metric Note\n\n"
        "Logistic, LASSO, and Ridge rows use CV5 mean metrics in the main `AUC`, `Accuracy`, `F1`, `Precision`, and `Recall` columns. GraphSAGE uses 5-seed held-out test mean metrics, not CV5 folds.\n\n"
        "## Key Deltas\n\n"
        + show_delta.to_markdown(index=False)
        + "\n\n"
        f"Output workbook: `{xlsx}`\n",
        encoding="utf-8",
    )
    print(f"Wrote {xlsx}")
    print(deltas.to_string(index=False))


if __name__ == "__main__":
    main()
