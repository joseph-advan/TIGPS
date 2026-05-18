from __future__ import annotations

import json
import shutil
import textwrap
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
SECTION_DIR = SCRIPT_PATH.parent
PAPER_RESULTS_DIR = SECTION_DIR.parent
SOURCE_XLSX = (
    PAPER_RESULTS_DIR
    / "04_feature_importance_top20"
    / "outputs"
    / "lasso_top20_feature_importance_with_categories.xlsx"
)

OUT_DIR = SECTION_DIR / "outputs"
FIG_DIR = OUT_DIR / "figures"
DIAG_DIR = OUT_DIR / "diagnostics"
MAIN_XLSX = OUT_DIR / "category_level_interpretation.xlsx"
SUMMARY_MD = OUT_DIR / "CATEGORY_LEVEL_INTERPRETATION_SUMMARY_ZH.md"
DIAGNOSTICS_JSON = DIAG_DIR / "category_level_interpretation_diagnostics.json"

CATEGORY_ORDER = [
    "SEL / Resilience",
    "Online / Digital Life",
    "Family / Parenting",
    "Bullying / Victimization",
    "Interpersonal Network",
    "Demographic / Social Status",
    "Delinquency / Risk Behavior",
    "Other",
]

CATEGORY_COLORS = {
    "SEL / Resilience": "#2F6B4F",
    "Online / Digital Life": "#2F5F98",
    "Family / Parenting": "#B56B45",
    "Bullying / Victimization": "#A33A3A",
    "Interpersonal Network": "#6C5A9E",
    "Demographic / Social Status": "#6B7280",
    "Delinquency / Risk Behavior": "#9A7B22",
    "Other": "#8A8A8A",
}

INTERACTION_DOMAIN_PRIORITY = {
    "SEL / Resilience": "Primary protective moderator",
    "Family / Parenting": "Secondary contextual moderator",
    "Bullying / Victimization": "Risk-context moderator",
    "Online / Digital Life": "Digital-life covariate or mechanism",
}


def reset_outputs() -> None:
    out_resolved = OUT_DIR.resolve()
    section_resolved = SECTION_DIR.resolve()
    if out_resolved.parent != section_resolved:
        raise RuntimeError(f"Refusing to remove unexpected output path: {out_resolved}")
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    DIAG_DIR.mkdir(parents=True, exist_ok=True)


def read_source() -> dict[str, pd.DataFrame]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE_XLSX}")
    sheets = {
        "category_wide": pd.read_excel(SOURCE_XLSX, sheet_name="CategorySummaryWide"),
        "lasso_combined": pd.read_excel(SOURCE_XLSX, sheet_name="LASSO_Top20_Combined"),
        "shared": pd.read_excel(SOURCE_XLSX, sheet_name="SharedTop20"),
        "interpersonal": pd.read_excel(SOURCE_XLSX, sheet_name="InterpersonalSummary"),
    }
    return {name: sanitize_frame(df) for name, df in sheets.items()}


def sanitize_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return (
        value.replace("Parent?hild", "Parent-Child")
        .replace("Parent?Child", "Parent-Child")
        .replace("Parent–Child", "Parent-Child")
    )


def sanitize_frame(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].map(sanitize_text)
    return out


def category_summary_long(category_wide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in category_wide.iterrows():
        category = row["Category"]
        for task in ["W2 -> W2", "W2 -> W3"]:
            n_col = f"{task} N Top 20"
            imp_col = f"{task} Relative Importance Sum %"
            rows.append(
                {
                    "Task": task,
                    "Category": category,
                    "N Top 20 Features": row.get(n_col),
                    "Relative Importance Sum %": row.get(imp_col),
                }
            )
    out = pd.DataFrame(rows)
    out["_category_order"] = out["Category"].map({c: i for i, c in enumerate(CATEGORY_ORDER)}).fillna(99)
    out["_task_order"] = out["Task"].map({"W2 -> W2": 1, "W2 -> W3": 2}).fillna(99)
    return out.sort_values(["_task_order", "_category_order"]).drop(columns=["_category_order", "_task_order"])


def safe_float(value: object) -> float:
    if value is None or pd.isna(value):
        return 0.0
    return float(value)


def safe_int(value: object) -> int:
    if value is None or pd.isna(value):
        return 0
    return int(value)


def build_category_narrative(category_wide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in category_wide.iterrows():
        category = str(row["Category"])
        w2_imp = safe_float(row.get("W2 -> W2 Relative Importance Sum %", 0))
        w3_imp = safe_float(row.get("W2 -> W3 Relative Importance Sum %", 0))
        w2_n = safe_int(row.get("W2 -> W2 N Top 20", 0))
        w3_n = safe_int(row.get("W2 -> W3 N Top 20", 0))
        mean_imp = (w2_imp + w3_imp) / 2
        diff = w3_imp - w2_imp

        if category == "SEL / Resilience":
            interpretation = "Dominant domain across both prediction tasks; supports a strong individual-capacity interpretation."
            paper_use = "Use as the primary explanatory domain and a main candidate for protective interaction analysis."
        elif category == "Online / Digital Life":
            interpretation = "Second-largest domain, especially stronger in W2 -> W3; supports the digital-life pathway."
            paper_use = "Use to connect the study to online activity and to define digital exposure/mechanism variables."
        elif category == "Interpersonal Network":
            interpretation = "Present but not dominant; weaker in longitudinal prediction."
            paper_use = "Use as evidence that peer-network structure adds limited explanatory weight compared with individual-level domains."
        elif category == "Bullying / Victimization":
            interpretation = "More prominent longitudinally than cross-sectionally; supports risk-context interpretation."
            paper_use = "Use as a risk domain and candidate interaction/context variable."
        elif category == "Family / Parenting":
            interpretation = "Moderate and stable contextual domain."
            paper_use = "Use as family context or adjustment domain."
        elif category == "Demographic / Social Status":
            interpretation = "Gender remains consistently selected, but this domain should be treated as background adjustment, not mechanism."
            paper_use = "Use as covariate/background interpretation."
        elif category == "Delinquency / Risk Behavior":
            interpretation = "Smaller domain but appears in both tasks."
            paper_use = "Use as secondary behavioral-risk context."
        else:
            interpretation = "Low or unclassified contribution."
            paper_use = "Review only if needed."

        rows.append(
            {
                "Category": category,
                "W2 -> W2 N Top 20": w2_n,
                "W2 -> W2 Relative Importance Sum %": w2_imp,
                "W2 -> W3 N Top 20": w3_n,
                "W2 -> W3 Relative Importance Sum %": w3_imp,
                "Mean Relative Importance %": mean_imp,
                "Longitudinal Minus Cross-sectional %": diff,
                "Interpretation": interpretation,
                "Paper Use": paper_use,
            }
        )
    out = pd.DataFrame(rows)
    out["_order"] = out["Category"].map({c: i for i, c in enumerate(CATEGORY_ORDER)}).fillna(99)
    return out.sort_values(["_order"]).drop(columns="_order").reset_index(drop=True)


def build_stable_categories(category_wide: pd.DataFrame) -> pd.DataFrame:
    narr = build_category_narrative(category_wide)
    out = narr[
        (pd.to_numeric(narr["W2 -> W2 N Top 20"], errors="coerce").fillna(0) > 0)
        & (pd.to_numeric(narr["W2 -> W3 N Top 20"], errors="coerce").fillna(0) > 0)
    ].copy()
    out["Stability Flag"] = out["Mean Relative Importance %"].apply(
        lambda v: "High stable domain" if v >= 15 else ("Moderate stable domain" if v >= 5 else "Low stable domain")
    )
    return out.sort_values("Mean Relative Importance %", ascending=False).reset_index(drop=True)


def build_candidate_interactions(lasso: pd.DataFrame, shared: pd.DataFrame) -> pd.DataFrame:
    shared_codes = set(shared["Feature Code"].astype(str))
    sub = lasso[lasso["Category"].isin(INTERACTION_DOMAIN_PRIORITY)].copy()
    sub["Feature Code"] = sub["Feature Code"].astype(str)
    sub["Appears in Both Tasks"] = sub["Feature Code"].isin(shared_codes)
    sub["Interaction Role"] = sub["Category"].map(INTERACTION_DOMAIN_PRIORITY)
    sub["Priority Score"] = (
        pd.to_numeric(sub["Relative Importance %"], errors="coerce").fillna(0)
        + sub["Appears in Both Tasks"].astype(int) * 5
        + sub["Category"].eq("SEL / Resilience").astype(int) * 3
    )
    sub["Recommended Use"] = sub.apply(recommended_interaction_use, axis=1)
    keep_cols = [
        "Task",
        "Variable",
        "Feature Code",
        "Category",
        "Rank by Abs Std. B",
        "Std. B",
        "Relative Importance %",
        "Direction",
        "Appears in Both Tasks",
        "Interaction Role",
        "Recommended Use",
        "Priority Score",
        "Items",
    ]
    out = sub[[c for c in keep_cols if c in sub.columns]].copy()
    return out.sort_values(["Priority Score", "Task"], ascending=[False, True]).reset_index(drop=True)


def recommended_interaction_use(row: pd.Series) -> str:
    category = str(row.get("Category", ""))
    direction = str(row.get("Direction", ""))
    variable = str(row.get("Variable", ""))
    if category == "SEL / Resilience" and direction == "Negative":
        return "Strong candidate protective moderator for High Online Activity interaction."
    if category == "SEL / Resilience":
        return "Candidate SEL moderator; inspect item direction before final interpretation."
    if category == "Bullying / Victimization":
        return "Candidate risk-context interaction with High Online Activity."
    if category == "Family / Parenting" and direction == "Negative":
        return "Candidate contextual protective moderator."
    if category == "Online / Digital Life":
        if "Problematic Internet" in variable:
            return "Digital risk mechanism; usually model as exposure/covariate rather than protective moderator."
        return "Digital-life mechanism/covariate; use carefully with High Online Activity to avoid redundancy."
    return "Secondary candidate."


def wrap_label(value: Any, width: int = 24) -> str:
    text = "" if pd.isna(value) else str(value)
    return "\n".join(textwrap.wrap(text, width=width, break_long_words=False))


def save_category_bar(category_wide: pd.DataFrame) -> Path:
    path = FIG_DIR / "category_level_relative_importance_bar.png"
    plot_df = category_wide.set_index("Category")
    cols = [c for c in plot_df.columns if c.endswith("Relative Importance Sum %")]
    fig, ax = plt.subplots(figsize=(13, 7))
    x = range(len(plot_df.index))
    width = 0.36
    for i, col in enumerate(cols):
        offset = (i - (len(cols) - 1) / 2) * width
        label = col.replace(" Relative Importance Sum %", "")
        ax.bar([v + offset for v in x], plot_df[col], width=width, label=label)
    ax.set_title("Category-Level Relative Importance from LASSO Top 20", fontsize=15, weight="bold")
    ax.set_ylabel("Relative Importance Sum (%)")
    ax.set_xticks(list(x))
    ax.set_xticklabels([wrap_label(v, 18) for v in plot_df.index], rotation=35, ha="right")
    ax.grid(axis="y", alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def save_domain_story_chart(narrative: pd.DataFrame) -> Path:
    path = FIG_DIR / "domain_story_mean_importance.png"
    plot_df = narrative.sort_values("Mean Relative Importance %", ascending=True).copy()
    fig, ax = plt.subplots(figsize=(11, 6.5))
    colors = [CATEGORY_COLORS.get(c, CATEGORY_COLORS["Other"]) for c in plot_df["Category"]]
    ax.barh(plot_df["Category"], plot_df["Mean Relative Importance %"], color=colors)
    ax.set_title("Average Domain Importance Across W2->W2 and W2->W3", fontsize=15, weight="bold")
    ax.set_xlabel("Mean Relative Importance Sum (%)")
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def save_interaction_candidate_chart(candidates: pd.DataFrame) -> Path:
    path = FIG_DIR / "top_interaction_candidate_variables.png"
    if candidates.empty:
        return path
    plot_df = candidates.drop_duplicates("Feature Code").sort_values("Priority Score", ascending=False).head(12)
    fig, ax = plt.subplots(figsize=(12, 7))
    colors = [CATEGORY_COLORS.get(c, CATEGORY_COLORS["Other"]) for c in plot_df["Category"]]
    ax.barh([wrap_label(v, 32) for v in plot_df["Variable"]], plot_df["Priority Score"], color=colors)
    ax.invert_yaxis()
    ax.set_title("Candidate Variables for Interaction Analysis", fontsize=15, weight="bold")
    ax.set_xlabel("Priority Score")
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=220)
    plt.close(fig)
    return path


def write_xlsx(sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(MAIN_XLSX, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    format_workbook(MAIN_XLSX)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 75))
            ws.column_dimensions[letter].width = max(12, max_len + 2)
    wb.save(path)


def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df.empty:
        return "_No rows available._"
    show = df if max_rows is None else df.head(max_rows)
    return show.to_markdown(index=False)


def write_summary(
    narrative: pd.DataFrame,
    stable: pd.DataFrame,
    candidates: pd.DataFrame,
    interpersonal: pd.DataFrame,
    figures: pd.DataFrame,
) -> None:
    lines = [
        "# Category-Level Interpretation Summary",
        "",
        "## 這個資料夾在做什麼",
        "",
        "05 不是重新訓練模型，而是把 04 的 LASSO Top 20 結果整理成可以放進論文的 conceptual-domain interpretation。也就是從單一變項移到較大的概念類別，例如 SEL / Resilience、Online / Digital Life、Family / Parenting、Bullying / Victimization、Interpersonal Network。",
        "",
        "## 主要結論",
        "",
        "1. SEL / Resilience 是兩個任務中最主要的類別，代表個人心理能力、自我價值與社會情緒能力是最強的解釋方向。",
        "2. Online / Digital Life 是第二個重要類別，尤其在 W2 -> W3 中更明顯，支持後續討論 digital life 與心理困擾的關係。",
        "3. Interpersonal Network 有訊號，但不是主導類別；W2 -> W3 中只剩 1 個 interpersonal feature 進入 LASSO Top 20。",
        "4. 這個結果支持目前的論文敘事：GNN 沒有明顯優於線性模型，且網絡結構本身不是最主要的預測來源；更需要往 SEL、resilience、online/digital life、family、bullying 等個人與情境特徵解釋。",
        "",
        "## Category Narrative",
        "",
        md_table(narrative),
        "",
        "## Stable Domains Across W2 -> W2 and W2 -> W3",
        "",
        md_table(stable),
        "",
        "## Interpersonal Network Interpretation",
        "",
        md_table(interpersonal),
        "",
        "## Candidate Variables for 06 Interaction Analysis",
        "",
        "這些候選變項不是最終模型，而是根據 04 的 Top20 與 shared Top20 整理出的下一步交互作用分析候選清單。",
        "",
        md_table(candidates.head(20)),
        "",
        "## Figures",
        "",
        md_table(figures),
        "",
        "## 建議接到 06 的分析邏輯",
        "",
        "下一步可以把 High Online Activity 當作 digital exposure，並測試它與 SEL / Resilience、Self-Worth、Bullying / Victimization、Family Context 的交互作用。如果交互作用項為負，代表該特徵可能削弱 high online activity 與 high psychological distress 的關聯，也就是 potential protective factor。",
        "",
    ]
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def build_figure_index(paths: list[Path]) -> pd.DataFrame:
    purpose = {
        "category_level_relative_importance_bar.png": "Compare category-level summed relative importance across W2 -> W2 and W2 -> W3.",
        "domain_story_mean_importance.png": "Show average domain importance across the two main prediction tasks.",
        "top_interaction_candidate_variables.png": "List high-priority variables to consider in 06 interaction analysis.",
    }
    return pd.DataFrame(
        [{"Figure": p.name, "Path": str(p), "Purpose": purpose.get(p.name, "")} for p in paths]
    )


def write_diagnostics(sheets: dict[str, pd.DataFrame], figures: pd.DataFrame) -> None:
    payload = {
        "source_xlsx": str(SOURCE_XLSX),
        "output_xlsx": str(MAIN_XLSX),
        "summary_md": str(SUMMARY_MD),
        "figures": figures.to_dict(orient="records"),
        "sheets": {name: {"rows": int(len(df)), "columns": list(df.columns)} for name, df in sheets.items()},
    }
    DIAGNOSTICS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    reset_outputs()
    source = read_source()
    category_wide = source["category_wide"]
    lasso = source["lasso_combined"]
    shared = source["shared"]
    interpersonal = source["interpersonal"]

    category_long = category_summary_long(category_wide)
    narrative = build_category_narrative(category_wide)
    stable = build_stable_categories(category_wide)
    candidates = build_candidate_interactions(lasso, shared)

    figure_paths = [
        save_category_bar(category_wide),
        save_domain_story_chart(narrative),
        save_interaction_candidate_chart(candidates),
    ]
    figures = build_figure_index(figure_paths)

    readme = pd.DataFrame(
        [
            {
                "Item": "Purpose",
                "Description": "Translate 04 LASSO Top20 feature importance into conceptual domain-level paper interpretation.",
            },
            {"Item": "Source", "Description": str(SOURCE_XLSX)},
            {"Item": "Primary Model", "Description": "LASSO Logistic, decomposed + 12 interpersonal features."},
            {"Item": "Main Tasks", "Description": "W2 -> W2 and W2 -> W3."},
        ]
    )

    sheets = {
        "ReadMe": readme,
        "CategorySummaryLong": category_long,
        "CategoryNarrative": narrative,
        "StableDomains": stable,
        "CandidateInteractions": candidates,
        "InterpersonalInterpretation": interpersonal,
        "SharedTop20Source": shared,
        "FigureIndex": figures,
    }
    write_xlsx(sheets)
    write_summary(narrative, stable, candidates, interpersonal, figures)
    write_diagnostics(sheets, figures)

    print(f"Wrote {MAIN_XLSX}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"Wrote {DIAGNOSTICS_JSON}")
    print("\nCategory narrative:")
    print(narrative.to_string(index=False))
    print("\nTop candidate interactions:")
    print(candidates.head(12).to_string(index=False))


if __name__ == "__main__":
    main()
