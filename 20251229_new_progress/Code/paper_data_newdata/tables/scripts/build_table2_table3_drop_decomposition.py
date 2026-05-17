from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression, LogisticRegressionCV
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    confusion_matrix,
    f1_score,
    make_scorer,
    precision_score,
    recall_score,
    roc_auc_score,
)
from sklearn.model_selection import StratifiedKFold, cross_validate, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

SCRIPT_PATH = Path(__file__).resolve()
TABLES_DIR = SCRIPT_PATH.parents[1]
ROOT = SCRIPT_PATH.parents[4]
CODE_DIR = ROOT / "Code" / "paper_data_newdata"

CORE_DIR = CODE_DIR / "Interpersonal_features"
FD_DIR = CODE_DIR / "Feature_Decomposition"
for p in [CORE_DIR, FD_DIR]:
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

import run_interpersonal_feature_logistic_comparison as core  # noqa: E402
import build_binary_drop_then_split_baseline as fd  # noqa: E402

W2_DATA = ROOT / "Data" / "testing_clean" / "W2" / "TIGPS_W2_studentdata_ver6.csv"
W3_DATA = ROOT / "Data" / "testing_clean" / "W3" / "TIGPS_W3_student_studentdata_ver5.csv"
FORMAL_NAMES_PATH = ROOT / "Data" / "otherData" / "論文使用之題組正式名稱.csv"
TABLE2_OUT = TABLES_DIR / "table2" / "outputs"
TABLE2_DOCS = TABLES_DIR / "table2" / "docs"
TABLE3_OUT = TABLES_DIR / "table3" / "outputs"
TABLE3_DOCS = TABLES_DIR / "table3" / "docs"
DIAG_OUT = TABLES_DIR / "diagnostics"

RANDOM_STATE = 42
TEST_SIZE = 0.2
LOGIT_CS = np.logspace(-4, 4, 41)
TARGET_W2_ITEMS = [f"v55_{i}" for i in range(1, 15)]
TARGET_W3_ITEMS = [f"54-{i}" for i in range(1, 15)]

SCENARIOS = {
    "W2": {
        "slug": "w2_2024",
        "title": "W2 2024",
        "target_group_id": "v55",
        "target_items": TARGET_W2_ITEMS,
    },
    "W3": {
        "slug": "w3_2025",
        "title": "W3 2025",
        "target_group_id": "54",
        "target_items": TARGET_W3_ITEMS,
    },
}


def ensure_dirs() -> None:
    for p in [TABLE2_OUT, TABLE2_DOCS, TABLE3_OUT, TABLE3_DOCS, DIAG_OUT]:
        p.mkdir(parents=True, exist_ok=True)


def min_valid_count(n_items: int) -> int:
    return max(1, int(math.ceil(n_items * 0.5)))


def scale_score(df: pd.DataFrame, columns: list[str], agg: str = "mean") -> pd.Series:
    existing = [c for c in columns if c in df.columns]
    if not existing:
        return pd.Series(np.nan, index=df.index, dtype="float")
    values = df[existing].apply(pd.to_numeric, errors="coerce")
    valid_count = values.notna().sum(axis=1)
    min_valid = min_valid_count(len(existing))
    if agg == "sum":
        score = values.sum(axis=1, min_count=min_valid)
    elif agg == "mean":
        score = values.mean(axis=1, skipna=True)
        score[valid_count < min_valid] = np.nan
    else:
        raise ValueError(f"Unsupported agg: {agg}")
    return pd.to_numeric(score, errors="coerce")


def make_target(df: pd.DataFrame, items: list[str]) -> tuple[pd.Series, pd.Series, dict[str, Any]]:
    score = scale_score(df, items, agg="sum")
    median = float(score.median(skipna=True))
    binary = pd.Series(np.nan, index=df.index, dtype="float")
    binary.loc[score.notna()] = score.loc[score.notna()].ge(median).astype(int)
    diag = {
        "target_items": items,
        "target_score_aggregation": "sum",
        "target_min_valid_items": min_valid_count(len(items)),
        "target_median_cutoff": median,
        "target_non_missing": int(score.notna().sum()),
        "target_positive": int(binary.eq(1).sum()),
        "target_negative": int(binary.eq(0).sum()),
    }
    return score, binary, diag


def format_float(value: float | None, digits: int = 3) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def format_p(value: float | None) -> str:
    if value is None or pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def safe_exp(value: float | None) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        return float(np.exp(value))
    except Exception:
        return None


def markdown_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    show = df if max_rows is None else df.head(max_rows)
    cols = list(show.columns)
    lines = ["| " + " | ".join(cols) + " |", "| " + " | ".join(["---"] * len(cols)) + " |"]
    for _, row in show.iterrows():
        vals = [str(row[col]).replace("|", "/") for col in cols]
        lines.append("| " + " | ".join(vals) + " |")
    if max_rows is not None and len(df) > max_rows:
        lines.append(f"\nShowing first {max_rows} of {len(df)} rows. See XLSX for full table.")
    return "\n".join(lines)


def write_xlsx(df: pd.DataFrame, path: Path, sheet_name: str) -> None:
    sheet_name = sheet_name[:31] or "Sheet1"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 60)


def write_grouped_table2_xlsx(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Table2"
    groups = [
        ("Variable Information", [("Variable", "Variable"), ("Feature Code", "Feature Code"), ("Items", "Items")]),
        ("Univariate Logistic", [("B", "Univariate B"), ("SE", "Univariate SE"), ("p-value", "Univariate p-value")]),
        ("LASSO Logistic", [("Std. B", "LASSO Std. B"), ("Relative Importance %", "LASSO Relative Importance %")]),
        ("Ridge Logistic", [("Std. B", "Ridge Std. B"), ("Relative Importance %", "Ridge Relative Importance %")]),
    ]
    col_idx = 1
    for group_name, columns in groups:
        start = col_idx
        end = col_idx + len(columns) - 1
        ws.cell(row=1, column=start, value=group_name)
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for offset, (display_col, _) in enumerate(columns):
            ws.cell(row=2, column=start + offset, value=display_col)
        col_idx = end + 1
    ordered_columns = [data_col for _, columns in groups for _, data_col in columns]
    for row_idx, (_, row) in enumerate(df[ordered_columns].iterrows(), start=3):
        for col_idx, col in enumerate(ordered_columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col])
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value))
            if ws.cell(row=row_idx, column=col_idx).value is not None
            else 0
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)
    wb.save(path)


def build_table3_performance_table(perf_df: pd.DataFrame, wave_title: str) -> pd.DataFrame:
    """Keep Table 3 focused on model-level predictive performance."""
    rename = {
        "CV auc mean": "CV AUC Mean",
        "CV auc SD": "CV AUC SD",
        "CV accuracy mean": "CV Accuracy Mean",
        "CV accuracy SD": "CV Accuracy SD",
        "CV balanced_accuracy mean": "CV Balanced Accuracy Mean",
        "CV balanced_accuracy SD": "CV Balanced Accuracy SD",
        "CV precision mean": "CV Precision Mean",
        "CV precision SD": "CV Precision SD",
        "CV recall mean": "CV Recall Mean",
        "CV recall SD": "CV Recall SD",
        "CV f1 mean": "CV F1 Mean",
        "CV f1 SD": "CV F1 SD",
        "CV sensitivity mean": "CV Sensitivity Mean",
        "CV sensitivity SD": "CV Sensitivity SD",
        "CV specificity mean": "CV Specificity Mean",
        "CV specificity SD": "CV Specificity SD",
    }
    out = perf_df.rename(columns=rename).copy()
    out.insert(0, "Wave", wave_title)
    ordered = [
        "Wave",
        "Model",
        "N",
        "N features",
        "Selected C",
        "CV AUC Mean",
        "CV AUC SD",
        "CV Accuracy Mean",
        "CV Accuracy SD",
        "CV Balanced Accuracy Mean",
        "CV Balanced Accuracy SD",
        "CV Precision Mean",
        "CV Precision SD",
        "CV Recall Mean",
        "CV Recall SD",
        "CV F1 Mean",
        "CV F1 SD",
        "CV Sensitivity Mean",
        "CV Sensitivity SD",
        "CV Specificity Mean",
        "CV Specificity SD",
        "Test AUC",
        "Test Accuracy",
        "Test Balanced Accuracy",
        "Test Precision",
        "Test Recall",
        "Test F1",
    ]
    return out[[c for c in ordered if c in out.columns]]


def write_grouped_table3_xlsx(df: pd.DataFrame, path: Path, sheet_name: str = "Table3") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Table3"
    groups = [
        ("Model Information", ["Wave", "Model", "N", "N features", "Selected C"]),
        (
            "5-fold Cross-Validation",
            [
                "CV AUC Mean",
                "CV AUC SD",
                "CV Accuracy Mean",
                "CV Accuracy SD",
                "CV Balanced Accuracy Mean",
                "CV Balanced Accuracy SD",
                "CV Precision Mean",
                "CV Precision SD",
                "CV Recall Mean",
                "CV Recall SD",
                "CV F1 Mean",
                "CV F1 SD",
                "CV Sensitivity Mean",
                "CV Sensitivity SD",
                "CV Specificity Mean",
                "CV Specificity SD",
            ],
        ),
        (
            "Holdout Test Set",
            ["Test AUC", "Test Accuracy", "Test Balanced Accuracy", "Test Precision", "Test Recall", "Test F1"],
        ),
    ]
    ordered_columns = [col for _, cols in groups for col in cols if col in df.columns]
    col_idx = 1
    for group_name, columns in groups:
        existing = [col for col in columns if col in df.columns]
        if not existing:
            continue
        start = col_idx
        end = col_idx + len(existing) - 1
        ws.cell(row=1, column=start, value=group_name)
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for offset, col in enumerate(existing):
            ws.cell(row=2, column=start + offset, value=col)
        col_idx = end + 1
    for row_idx, (_, row) in enumerate(df[ordered_columns].iterrows(), start=3):
        for col_idx, col in enumerate(ordered_columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col])
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            len(str(ws.cell(row=row_idx, column=col_idx).value))
            if ws.cell(row=row_idx, column=col_idx).value is not None
            else 0
            for row_idx in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)
    wb.save(path)


def write_analysis_logic_xlsx(df: pd.DataFrame, path: Path, wave_title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out = df.copy()
    out["Univariate B numeric"] = pd.to_numeric(out["Univariate B"], errors="coerce")
    out["Univariate p numeric"] = pd.to_numeric(out["Univariate p-value"].replace("<0.001", "0.0005"), errors="coerce")
    out["LASSO Std. B numeric"] = pd.to_numeric(out["LASSO Std. B"], errors="coerce")
    out["Ridge Std. B numeric"] = pd.to_numeric(out["Ridge Std. B"], errors="coerce")
    out["LASSO Importance numeric"] = pd.to_numeric(out["LASSO Relative Importance %"], errors="coerce")
    out["Ridge Importance numeric"] = pd.to_numeric(out["Ridge Relative Importance %"], errors="coerce")
    lasso_cutoff = out["LASSO Importance numeric"].quantile(0.75)
    ridge_cutoff = out["Ridge Importance numeric"].quantile(0.75)
    out["Univariate significant"] = out["Univariate p numeric"].lt(0.05)
    out["LASSO high importance"] = out["LASSO Importance numeric"].ge(lasso_cutoff)
    out["Ridge high importance"] = out["Ridge Importance numeric"].ge(ridge_cutoff)
    out["Supported by all three"] = out[
        ["Univariate significant", "LASSO high importance", "Ridge high importance"]
    ].all(axis=1)
    out["Logic note"] = np.select(
        [
            out["Supported by all three"],
            out["Univariate significant"] & (out["LASSO high importance"] | out["Ridge high importance"]),
            out["Univariate significant"],
            out["LASSO high importance"] | out["Ridge high importance"],
        ],
        [
            "Significant univariate association and high regularized-model importance in both LASSO and Ridge.",
            "Significant univariate association and high importance in at least one regularized model.",
            "Significant in univariate model only; lower regularized-model importance.",
            "High regularized-model importance without univariate p < 0.05.",
        ],
        default="Not prominent by the current summary rules.",
    )
    summary_cols = [
        "Variable",
        "Feature Code",
        "Univariate B",
        "Univariate p-value",
        "LASSO Std. B",
        "LASSO Relative Importance %",
        "Ridge Std. B",
        "Ridge Relative Importance %",
        "Univariate significant",
        "LASSO high importance",
        "Ridge high importance",
        "Supported by all three",
        "Logic note",
    ]
    top_univ = out.loc[out["Univariate significant"]].copy()
    top_univ["abs_univ_b"] = top_univ["Univariate B numeric"].abs()
    top_univ = top_univ.sort_values("abs_univ_b", ascending=False).head(10)[summary_cols]
    top_lasso = out.sort_values("LASSO Importance numeric", ascending=False).head(10)[summary_cols]
    top_ridge = out.sort_values("Ridge Importance numeric", ascending=False).head(10)[summary_cols]
    all_three = out.loc[out["Supported by all three"]].sort_values(
        ["LASSO Importance numeric", "Ridge Importance numeric"], ascending=False
    )[summary_cols]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        intro = pd.DataFrame(
            [
                {"Item": "Wave", "Definition": wave_title},
                {"Item": "Univariate significant", "Definition": "Univariate logistic p-value < 0.05."},
                {"Item": "LASSO high importance", "Definition": f"LASSO relative importance >= wave-specific 75th percentile ({lasso_cutoff:.2f}%)."},
                {"Item": "Ridge high importance", "Definition": f"Ridge relative importance >= wave-specific 75th percentile ({ridge_cutoff:.2f}%)."},
                {"Item": "Supported by all three", "Definition": "Univariate significant + LASSO high importance + Ridge high importance."},
            ]
        )
        intro.to_excel(writer, index=False, sheet_name="Logic")
        top_univ.to_excel(writer, index=False, sheet_name="Top_Univariate")
        top_lasso.to_excel(writer, index=False, sheet_name="Top_LASSO")
        top_ridge.to_excel(writer, index=False, sheet_name="Top_Ridge")
        all_three.to_excel(writer, index=False, sheet_name="All_Three")
        out[summary_cols].to_excel(writer, index=False, sheet_name="All_Variables")
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            for col_idx in range(1, worksheet.max_column + 1):
                max_len = max(
                    len(str(worksheet.cell(row=row_idx, column=col_idx).value))
                    if worksheet.cell(row=row_idx, column=col_idx).value is not None
                    else 0
                    for row_idx in range(1, worksheet.max_row + 1)
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)


def normalize_group_id(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text


def load_formal_name_map() -> dict[str, dict[str, str]]:
    mapping = {"W2": {}, "W3": {}}
    if not FORMAL_NAMES_PATH.exists():
        return mapping
    names = pd.read_csv(FORMAL_NAMES_PATH, encoding="utf-8-sig")
    for _, row in names.iterrows():
        english = str(row.get("Suggested formal English name", "")).strip()
        if not english:
            continue
        w2_id = normalize_group_id(row.get("W2 Group ID", ""))
        w3_id = normalize_group_id(row.get("W3 Group ID", ""))
        if w2_id:
            mapping["W2"][w2_id] = english
        if w3_id:
            mapping["W3"][w3_id] = english
    return mapping


def display_feature_name(item: dict[str, Any], wave: str, formal_name_map: dict[str, dict[str, str]]) -> str:
    if item.get("is_gender_dummy"):
        return "Gender: Male (vs Female)"
    if item.get("subscale_name_en"):
        return str(item["subscale_name_en"])
    source_group_id = normalize_group_id(item.get("source_group_id", ""))
    used_items = [str(value) for value in item.get("used_items", [])]
    if wave == "W2" and source_group_id == "v52" and used_items and all(value.startswith("v52_") for value in used_items):
        # W2 uses v52 for self-rated health, but v52_1-v52_3 are the Self-Worth scale.
        if "v52_" in formal_name_map.get(wave, {}):
            return formal_name_map[wave]["v52_"]
    if source_group_id in formal_name_map.get(wave, {}):
        return formal_name_map[wave][source_group_id]
    if item.get("formal_group_name_en"):
        return str(item["formal_group_name_en"])
    return str(item.get("feature_name", ""))


def build_drop_decomposition_features(df: pd.DataFrame, merged: pd.DataFrame, wave: str, target_group_id: str) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    config = fd.load_subscale_config(fd.SUBSCALE_CONFIG_PATH)
    formal_name_map = load_formal_name_map()
    split_specs = fd.split_specs_from_config(config, wave)
    direct_specs = fd.direct_feature_specs_from_config(config, wave)
    metadata = fd.feature_metadata_from_config(config, wave)
    feature_group_ids = fd.W2_FEATURE_GROUP_IDS if wave == "W2" else fd.W3_FEATURE_GROUP_IDS
    drop_group_ids = fd.W2_DROP_GROUP_IDS if wave == "W2" else fd.W3_DROP_GROUP_IDS
    feature_group_ids = [gid for gid in feature_group_ids if gid != target_group_id]
    feature_table, feature_meta = fd.build_feature_table(
        df,
        merged,
        year=wave,
        feature_group_ids=feature_group_ids,
        drop_group_ids=drop_group_ids,
        split_specs=split_specs,
        direct_feature_specs=direct_specs,
        feature_metadata=metadata,
    )
    rows = []
    for item in feature_meta["feature_defs"]:
        feature_name = item["feature_name"]
        col = f"feature_{feature_name}"
        if col not in feature_table.columns:
            continue
        english_name = display_feature_name(item, wave, formal_name_map)
        source_type = (
            "categorical_dummy"
            if item.get("is_gender_dummy")
            else (
                "decomposed_subscale"
                if item.get("is_split_feature")
                else ("direct_feature" if item.get("is_direct_feature") else "scale_or_single_item")
            )
        )
        rows.append(
            {
                "wave": wave,
                "feature_code": str(feature_name),
                "feature_name": english_name,
                "model_column": col,
                "source_type": source_type,
                "items": ";".join(item.get("used_items", [])),
                "source_group_id": item.get("source_group_id", ""),
                "score_aggregation": item.get("score_aggregation", "mean"),
                "n_non_missing": int(feature_table[col].notna().sum()),
                "n_unique_non_missing": int(feature_table[col].dropna().nunique()),
            }
        )
    feature_defs = pd.DataFrame(rows)
    keep_cols = [col for col in feature_defs["model_column"] if feature_table[col].notna().sum() > 0 and feature_table[col].dropna().nunique() > 1]
    X = feature_table[keep_cols].copy()
    feature_defs = feature_defs.loc[feature_defs["model_column"].isin(keep_cols)].reset_index(drop=True)
    diag = {
        "wave": wave,
        "feature_set": "drop_plus_decomposition",
        "n_rows": int(len(df)),
        "n_predictor_columns": int(X.shape[1]),
        "drop_group_ids": ";".join(sorted(drop_group_ids)),
        "split_group_ids": ";".join(sorted(split_specs.keys())),
        "direct_feature_ids": ";".join(sorted(direct_specs.keys())),
        "skipped_no_mapping": ";".join(feature_meta.get("skipped_feature_groups_no_mapping_items", [])),
        "skipped_no_columns": ";".join(feature_meta.get("skipped_feature_groups_no_columns", [])),
    }
    return X, feature_defs, diag


def fit_univariate(y: pd.Series, X: pd.DataFrame, feature_defs: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, meta in feature_defs.iterrows():
        col = meta["model_column"]
        model_df = pd.DataFrame({"y": y, "x": X[col]}).dropna()
        row = meta.to_dict()
        row.update({"univariate_n": int(len(model_df)), "univariate_status": ""})
        if len(model_df) < 50 or model_df["y"].nunique() < 2 or model_df["x"].nunique() < 2:
            row.update({"univariate_status": "insufficient_variation"})
            rows.append(row)
            continue
        try:
            fit = sm.Logit(model_df["y"], sm.add_constant(model_df[["x"]], has_constant="add")).fit(disp=False, maxiter=200)
            b = float(fit.params["x"])
            se = float(fit.bse["x"])
            p = float(fit.pvalues["x"])
            ci_low, ci_high = fit.conf_int().loc["x"].tolist()
            row.update(
                {
                    "univariate_b": b,
                    "univariate_se": se,
                    "univariate_p": p,
                    "univariate_or": safe_exp(b),
                    "univariate_or_ci_low": safe_exp(ci_low),
                    "univariate_or_ci_high": safe_exp(ci_high),
                    "univariate_status": "ok",
                }
            )
        except Exception as exc:
            row.update({"univariate_status": f"failed: {type(exc).__name__}"})
        rows.append(row)
    return pd.DataFrame(rows)


def fit_multivariable(y: pd.Series, X: pd.DataFrame, feature_defs: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    model_df = pd.concat([y.rename("y"), X], axis=1).dropna()
    diag = {
        "multivariable_n": int(len(model_df)),
        "multivariable_n_positive": int(model_df["y"].eq(1).sum()) if len(model_df) else 0,
        "multivariable_n_negative": int(model_df["y"].eq(0).sum()) if len(model_df) else 0,
        "multivariable_status": "not_run",
    }
    rows = feature_defs.copy()
    if len(model_df) < 50 or model_df["y"].nunique() < 2:
        diag["multivariable_status"] = "insufficient_rows_or_target_variation"
        return rows, diag
    try:
        fit = sm.Logit(model_df["y"], sm.add_constant(model_df[X.columns], has_constant="add")).fit(disp=False, maxiter=300)
        diag["multivariable_status"] = "ok"
        diag["multivariable_pseudo_r2"] = float(fit.prsquared)
        diag["multivariable_aic"] = float(fit.aic)
        for idx, meta in rows.iterrows():
            col = meta["model_column"]
            if col not in fit.params.index:
                continue
            b = float(fit.params[col])
            se = float(fit.bse[col])
            p = float(fit.pvalues[col])
            ci_low, ci_high = fit.conf_int().loc[col].tolist()
            rows.loc[idx, "multivariable_b"] = b
            rows.loc[idx, "multivariable_se"] = se
            rows.loc[idx, "multivariable_p"] = p
            rows.loc[idx, "multivariable_or"] = safe_exp(b)
            rows.loc[idx, "multivariable_or_ci_low"] = safe_exp(ci_low)
            rows.loc[idx, "multivariable_or_ci_high"] = safe_exp(ci_high)
            rows.loc[idx, "multivariable_n"] = int(len(model_df))
            rows.loc[idx, "multivariable_status"] = "ok"
    except Exception as exc:
        diag["multivariable_status"] = f"failed: {type(exc).__name__}: {exc}"
    return rows, diag


def clean_table2(univ: pd.DataFrame, multi: pd.DataFrame, multi_diag: dict[str, Any]) -> pd.DataFrame:
    merged = univ.copy()
    for col in [
        "multivariable_b", "multivariable_se", "multivariable_p", "multivariable_or",
        "multivariable_or_ci_low", "multivariable_or_ci_high", "multivariable_n", "multivariable_status",
    ]:
        if col in multi.columns:
            merged[col] = multi[col].values
    out = pd.DataFrame({
        "Variable": merged["feature_name"],
        "Feature Code": merged["feature_code"],
        "Model Column": merged["model_column"],
        "Source Type": merged["source_type"],
        "Items": merged["items"],
        "Univariate N": merged.get("univariate_n", ""),
        "Univariate B": merged.get("univariate_b", np.nan).map(format_float),
        "Univariate SE": merged.get("univariate_se", np.nan).map(format_float),
        "Univariate p-value": merged.get("univariate_p", np.nan).map(format_p),
        "Multivariable N": merged.get("multivariable_n", multi_diag.get("multivariable_n", "")),
        "Multivariable B": merged.get("multivariable_b", np.nan).map(format_float),
        "Multivariable SE": merged.get("multivariable_se", np.nan).map(format_float),
        "Multivariable p-value": merged.get("multivariable_p", np.nan).map(format_p),
        "Univariate Status": merged.get("univariate_status", ""),
        "Multivariable Status": merged.get("multivariable_status", multi_diag.get("multivariable_status", "")),
    })
    return out


def build_combined_model_table2(univ: pd.DataFrame, coef_table: pd.DataFrame) -> pd.DataFrame:
    univ_sub = univ[["model_column", "feature_name", "feature_code", "items", "univariate_b", "univariate_se", "univariate_p"]].copy()
    univ_sub = univ_sub.rename(
        columns={
            "model_column": "Model Column",
            "feature_name": "Variable",
            "feature_code": "Feature Code",
            "items": "Items",
        }
    )
    coef_sub = coef_table[
        ["Model Column", "LASSO Logistic Std. B", "Ridge Logistic Std. B"]
    ].copy()
    merged = univ_sub.merge(coef_sub, on="Model Column", how="left")
    for model_prefix in ["LASSO Logistic", "Ridge Logistic"]:
        coef_col = f"{model_prefix} Std. B"
        importance_col = f"{model_prefix} Relative Importance %"
        coef_abs = pd.to_numeric(merged[coef_col], errors="coerce").abs()
        denom = coef_abs.sum()
        merged[importance_col] = np.where(denom > 0, coef_abs / denom * 100, np.nan)
    return pd.DataFrame(
        {
            "Variable": merged["Variable"],
            "Feature Code": merged["Feature Code"],
            "Items": merged["Items"],
            "Univariate B": merged["univariate_b"].map(format_float),
            "Univariate SE": merged["univariate_se"].map(format_float),
            "Univariate p-value": merged["univariate_p"].map(format_p),
            "LASSO Std. B": merged["LASSO Logistic Std. B"],
            "LASSO Relative Importance %": pd.Series(merged["LASSO Logistic Relative Importance %"]).map(lambda x: format_float(x, 2)),
            "Ridge Std. B": merged["Ridge Logistic Std. B"],
            "Ridge Relative Importance %": pd.Series(merged["Ridge Logistic Relative Importance %"]).map(lambda x: format_float(x, 2)),
        }
    )


def specificity_score(y_true, y_pred) -> float:
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return float(tn / (tn + fp)) if (tn + fp) else 0.0


def sensitivity_score(y_true, y_pred) -> float:
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return float(tp / (tp + fn)) if (tp + fn) else 0.0


def model_metrics(y_true: pd.Series, prob: np.ndarray) -> dict[str, float]:
    pred = (prob >= 0.5).astype(int)
    return {
        "accuracy": float(accuracy_score(y_true, pred)),
        "balanced_accuracy": float(balanced_accuracy_score(y_true, pred)),
        "sensitivity": sensitivity_score(y_true, pred),
        "specificity": specificity_score(y_true, pred),
        "precision": float(precision_score(y_true, pred, zero_division=0)),
        "recall": float(recall_score(y_true, pred, zero_division=0)),
        "f1": float(f1_score(y_true, pred, zero_division=0)),
        "auc": float(roc_auc_score(y_true, prob)),
    }


def make_pipeline(model) -> Pipeline:
    return Pipeline(
        steps=[
            ("imputer", SimpleImputer(strategy="median")),
            ("scaler", StandardScaler()),
            ("model", model),
        ]
    )


def fit_model_comparison(y: pd.Series, X: pd.DataFrame, feature_defs: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    valid = y.notna()
    yv = y[valid].astype(int)
    Xv = X.loc[valid].copy()
    X_train, X_test, y_train, y_test = train_test_split(Xv, yv, test_size=TEST_SIZE, random_state=RANDOM_STATE, stratify=yv)
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=RANDOM_STATE)
    models = {
        "Multivariable Logistic": LogisticRegression(penalty=None, solver="lbfgs", max_iter=5000),
        "LASSO Logistic": LogisticRegressionCV(Cs=LOGIT_CS, penalty="l1", solver="saga", cv=cv, scoring="roc_auc", max_iter=20000, n_jobs=None, random_state=RANDOM_STATE),
        "Ridge Logistic": LogisticRegressionCV(Cs=LOGIT_CS, penalty="l2", solver="lbfgs", cv=cv, scoring="roc_auc", max_iter=10000, n_jobs=None),
    }
    perf_rows = []
    coef_frames = []
    scoring = {
        "accuracy": "accuracy",
        "balanced_accuracy": "balanced_accuracy",
        "f1": "f1",
        "auc": "roc_auc",
        "precision": make_scorer(precision_score, zero_division=0),
        "recall": make_scorer(recall_score, zero_division=0),
        "sensitivity": make_scorer(sensitivity_score),
        "specificity": make_scorer(specificity_score),
    }
    for model_name, model in models.items():
        pipe = make_pipeline(model)
        pipe.fit(X_train, y_train)
        prob = pipe.predict_proba(X_test)[:, 1]
        test = model_metrics(y_test, prob)
        fitted = pipe.named_steps["model"]
        selected_c = getattr(fitted, "C_", [np.nan])[0] if hasattr(fitted, "C_") else np.nan
        if model_name == "Multivariable Logistic":
            cv_model = LogisticRegression(penalty=None, solver="lbfgs", max_iter=5000)
        elif model_name == "LASSO Logistic":
            cv_model = LogisticRegression(C=float(selected_c), penalty="l1", solver="saga", max_iter=20000, random_state=RANDOM_STATE)
        else:
            cv_model = LogisticRegression(C=float(selected_c), penalty="l2", solver="lbfgs", max_iter=10000)
        cv_scores = cross_validate(make_pipeline(cv_model), Xv, yv, cv=cv, scoring=scoring, n_jobs=None)
        perf = {
            "Model": model_name,
            "N": int(len(Xv)),
            "N features": int(Xv.shape[1]),
            "Selected C": format_float(selected_c, 6),
            "Test AUC": format_float(test["auc"]),
            "Test Accuracy": format_float(test["accuracy"]),
            "Test Precision": format_float(test["precision"]),
            "Test Recall": format_float(test["recall"]),
            "Test F1": format_float(test["f1"]),
            "Test Balanced Accuracy": format_float(test["balanced_accuracy"]),
        }
        for metric in scoring:
            perf[f"CV {metric} mean"] = format_float(float(np.mean(cv_scores[f"test_{metric}"])))
            perf[f"CV {metric} SD"] = format_float(float(np.std(cv_scores[f"test_{metric}"], ddof=1)))
        perf_rows.append(perf)
        coefs = fitted.coef_[0]
        coef_frame = feature_defs[["model_column", "feature_code", "feature_name", "source_type", "items"]].copy()
        coef_frame["Model"] = model_name
        coef_frame["Standardized Coefficient"] = coefs
        if model_name == "LASSO Logistic":
            coef_frame["Selected by LASSO"] = np.abs(coefs) > 1e-8
        coef_frames.append(coef_frame)
    perf_df = pd.DataFrame(perf_rows)
    coef_long = pd.concat(coef_frames, ignore_index=True)
    coef_wide = feature_defs[["model_column", "feature_code", "feature_name", "source_type", "items"]].copy()
    coef_wide = coef_wide.rename(columns={"feature_name": "Variable", "feature_code": "Feature Code", "model_column": "Model Column", "source_type": "Source Type", "items": "Items"})
    for model_name in models:
        sub = coef_long[coef_long["Model"].eq(model_name)].set_index("model_column")
        coef_wide[f"{model_name} Std. B"] = coef_wide["Model Column"].map(sub["Standardized Coefficient"]).map(lambda x: format_float(x, 4))
    lasso_sub = coef_long[coef_long["Model"].eq("LASSO Logistic")].set_index("model_column")
    coef_wide["Selected by LASSO"] = coef_wide["Model Column"].map(lasso_sub["Selected by LASSO"]).map(lambda value: bool(value) if pd.notna(value) else False)
    return coef_wide, perf_df


def run_wave(wave: str, df: pd.DataFrame, merged: pd.DataFrame) -> dict[str, Any]:
    cfg = SCENARIOS[wave]
    slug = cfg["slug"]
    target_score, y, target_diag = make_target(df, cfg["target_items"])
    X, feature_defs, feature_diag = build_drop_decomposition_features(df, merged, wave, cfg["target_group_id"])
    univ = fit_univariate(y, X, feature_defs)
    multi, multi_diag = fit_multivariable(y, X, feature_defs)
    coef_table, perf_table = fit_model_comparison(y, X, feature_defs)
    table2 = build_combined_model_table2(univ, coef_table)
    write_grouped_table2_xlsx(table2, TABLE2_OUT / f"table2_{slug}_logistic.xlsx")
    write_analysis_logic_xlsx(table2, TABLE2_OUT / f"table2_{slug}_analysis_logic.xlsx", cfg["title"])
    write_xlsx(feature_defs, TABLE2_DOCS / f"feature_dictionary_{slug}.xlsx", "FeatureDictionary")
    table3 = build_table3_performance_table(perf_table, cfg["title"])
    write_grouped_table3_xlsx(table3, TABLE3_OUT / f"table3_{slug}_model_performance.xlsx", "Table3")
    return {"wave": wave, **target_diag, **feature_diag, **multi_diag}


def write_notes(diagnostics: pd.DataFrame) -> None:
    # Notes are kept in XLSX/JSON diagnostics only. The tables folder is kept
    # free of Markdown outputs for the current workflow.
    return


def main() -> None:
    ensure_dirs()
    merged_path = core.pick_first_existing_path(core.MERGED_PATH_CANDIDATES)
    merged = pd.read_csv(merged_path, dtype=str, encoding="utf-8-sig")
    for c in ["Year", "Group_ID", "Question_ID"]:
        if c in merged.columns:
            merged[c] = merged[c].astype(str).str.strip()
    w2 = core.normalize_student_id(pd.read_csv(W2_DATA, encoding="utf-8-sig", low_memory=False))
    w3 = core.normalize_student_id(pd.read_csv(W3_DATA, encoding="utf-8-sig", low_memory=False))
    diagnostics = pd.DataFrame([run_wave("W2", w2, merged), run_wave("W3", w3, merged)])
    write_xlsx(diagnostics, DIAG_OUT / "table2_table3_drop_decomposition_diagnostics.xlsx", "Diagnostics")
    (DIAG_OUT / "table2_table3_drop_decomposition_diagnostics.json").write_text(json.dumps(diagnostics.to_dict(orient="records"), ensure_ascii=False, indent=2), encoding="utf-8")
    write_notes(diagnostics)
    print("Wrote drop + decomposition Table 2/3 outputs.")
    print(diagnostics.to_string(index=False))


if __name__ == "__main__":
    main()
