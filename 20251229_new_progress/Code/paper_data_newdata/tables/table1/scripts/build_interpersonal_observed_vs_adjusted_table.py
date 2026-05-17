from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats


SCRIPT_PATH = Path(__file__).resolve()
TABLE1_DIR = SCRIPT_PATH.parents[1]
TABLES_DIR = TABLE1_DIR.parent
ROOT = TABLE1_DIR.parents[3]
TABLE_SCRIPTS_DIR = TABLES_DIR / "scripts"
if str(TABLE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(TABLE_SCRIPTS_DIR))

import build_table2_table3_drop_decomposition as t23  # noqa: E402


W2_DATA = ROOT / "Data" / "testing_clean" / "W2" / "TIGPS_W2_studentdata_ver6.csv"
W3_DATA = ROOT / "Data" / "testing_clean" / "W3" / "TIGPS_W3_student_studentdata_ver5.csv"
INTERPERSONAL_DIR = ROOT / "Code" / "paper_data_newdata" / "Interpersonal_features" / "outputs" / "features"
W2_INTERPERSONAL = INTERPERSONAL_DIR / "interpersonal_features_w2.csv"
W3_INTERPERSONAL = INTERPERSONAL_DIR / "interpersonal_features_w3.csv"
OUT_DIR = TABLE1_DIR / "outputs" / "05_interpersonal_feature_comparison"
OUT_PATH = OUT_DIR / "table1_interpersonal_observed_vs_class_adjusted_psychological_distress.xlsx"


FEATURES = [
    {
        "feature": "Online total nominations",
        "observed_col": "ip_online_total",
        "adjusted_col": "ip_online_total_rate_class",
    },
    {
        "feature": "Offline total nominations",
        "observed_col": "ip_offline_total",
        "adjusted_col": "ip_offline_total_rate_class",
    },
    {
        "feature": "Outgoing friendship nominations",
        "observed_col": "ip_out_friend_total",
        "adjusted_col": "ip_out_friend_total_rate_class",
    },
    {
        "feature": "Incoming friendship nominations",
        "observed_col": "ip_in_friend_total",
        "adjusted_col": "ip_in_friend_total_rate_class",
    },
    {
        "feature": "Outgoing negative nominations",
        "observed_col": "ip_out_enemy_total",
        "adjusted_col": "ip_out_enemy_total_rate_class",
    },
    {
        "feature": "Incoming negative nominations",
        "observed_col": "ip_in_enemy_total",
        "adjusted_col": "ip_in_enemy_total_rate_class",
    },
    {
        "feature": "Reciprocal friendship ties",
        "observed_col": "ip_reciprocal_friend_count",
        "adjusted_col": "ip_reciprocal_friend_count_rate_class",
    },
    {
        "feature": "Reciprocal negative ties",
        "observed_col": "ip_reciprocal_enemy_count",
        "adjusted_col": "ip_reciprocal_enemy_count_rate_class",
    },
    {
        "feature": "Sent positive tie ratio",
        "observed_col": "ip_sent_like_ratio",
        "adjusted_col": "ip_sent_like_ratio",
    },
    {
        "feature": "Received positive tie ratio",
        "observed_col": "ip_received_like_ratio",
        "adjusted_col": "ip_received_like_ratio",
    },
    {
        "feature": "Sent network valence",
        "observed_col": "ip_sent_net",
        "adjusted_col": "ip_sent_net_rate_class",
    },
    {
        "feature": "Received network valence",
        "observed_col": "ip_received_net",
        "adjusted_col": "ip_received_net_rate_class",
    },
]


def format_mean_sd(values: pd.Series) -> str:
    values = pd.to_numeric(values, errors="coerce").dropna()
    if values.empty:
        return ""
    return f"{values.mean():.2f} ({values.std(ddof=1):.2f})"


def format_p(value: float | None) -> str:
    if value is None or pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def p_value(high: pd.Series, low: pd.Series) -> float | None:
    high = pd.to_numeric(high, errors="coerce").dropna()
    low = pd.to_numeric(low, errors="coerce").dropna()
    if len(high) < 2 or len(low) < 2:
        return None
    _, p = stats.ttest_ind(high, low, equal_var=False, nan_policy="omit")
    return float(p)


def make_distress_group(df: pd.DataFrame, wave: str) -> pd.Series:
    items = t23.TARGET_W2_ITEMS if wave == "W2" else t23.TARGET_W3_ITEMS
    _, binary, _ = t23.make_target(df, items)
    group = pd.Series(pd.NA, index=df.index, dtype="object")
    group.loc[binary.eq(1)] = "High Psychological Distress"
    group.loc[binary.eq(0)] = "Low Psychological Distress"
    return group


def load_wave_table(wave: str, survey_path: Path, interpersonal_path: Path) -> pd.DataFrame:
    survey = t23.core.normalize_student_id(pd.read_csv(survey_path, encoding="utf-8-sig", low_memory=False))
    group = make_distress_group(survey, wave)
    ip = t23.core.normalize_student_id(pd.read_csv(interpersonal_path, encoding="utf-8-sig", low_memory=False))
    ip["ip_online_total"] = (
        pd.to_numeric(ip.get("ip_out_online_friend"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_in_online_friend"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_out_online_enemy"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_in_online_enemy"), errors="coerce").fillna(0.0)
    )
    ip["ip_offline_total"] = (
        pd.to_numeric(ip.get("ip_out_offline_friend"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_in_offline_friend"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_out_offline_enemy"), errors="coerce").fillna(0.0)
        + pd.to_numeric(ip.get("ip_in_offline_enemy"), errors="coerce").fillna(0.0)
    )
    ip["ip_online_total_rate_class"] = ip["ip_online_total"] / pd.to_numeric(ip["ip_class_size_minus1"], errors="coerce")
    ip["ip_offline_total_rate_class"] = ip["ip_offline_total"] / pd.to_numeric(ip["ip_class_size_minus1"], errors="coerce")
    if "ip_sent_net_rate_class" not in ip.columns:
        ip["ip_sent_net_rate_class"] = pd.to_numeric(ip["ip_sent_net"], errors="coerce") / pd.to_numeric(
            ip["ip_class_size_minus1"], errors="coerce"
        )
    if "ip_received_net_rate_class" not in ip.columns:
        ip["ip_received_net_rate_class"] = pd.to_numeric(ip["ip_received_net"], errors="coerce") / pd.to_numeric(
            ip["ip_class_size_minus1"], errors="coerce"
        )
    needed = ["student_id"] + sorted({spec["observed_col"] for spec in FEATURES} | {spec["adjusted_col"] for spec in FEATURES})
    ip = ip[[c for c in needed if c in ip.columns]].drop_duplicates(subset=["student_id"], keep="first")
    data = survey[["student_id"]].merge(ip, on="student_id", how="left")
    data["distress_group"] = group.values

    rows = []
    for spec in FEATURES:
        observed = pd.to_numeric(data[spec["observed_col"]], errors="coerce")
        adjusted = pd.to_numeric(data[spec["adjusted_col"]], errors="coerce")
        high = data["distress_group"].eq("High Psychological Distress")
        low = data["distress_group"].eq("Low Psychological Distress")
        rows.append(
            {
                "Feature": spec["feature"],
                "Observed p-value": format_p(p_value(observed[high], observed[low])),
                "Class-adjusted p-value": format_p(p_value(adjusted[high], adjusted[low])),
                "Observed High Distress, mean (SD)": format_mean_sd(observed[high]),
                "Class-adjusted High Distress, mean (SD)": format_mean_sd(adjusted[high]),
                "Observed Low Distress, mean (SD)": format_mean_sd(observed[low]),
                "Class-adjusted Low Distress, mean (SD)": format_mean_sd(adjusted[low]),
                "Observed source column": spec["observed_col"],
                "Class-adjusted source column": spec["adjusted_col"],
            }
        )
    return pd.DataFrame(rows)


def write_output(w2: pd.DataFrame, w3: pd.DataFrame) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    readme = pd.DataFrame(
        [
            {
                "Item": "Purpose",
                "Description": "Compare interpersonal features before and after respondent-class adjustment for psychological distress groups.",
            },
            {
                "Item": "Observed features",
                "Description": "Raw observed interpersonal counts, ratios, or valence scores from the peer nomination network.",
            },
            {
                "Item": "Class-adjusted features",
                "Description": "Nomination and reciprocal count features divided by same-class respondents minus one. Ratio and valence features are not class-size denominators, so their observed and adjusted columns use the same source.",
            },
            {
                "Item": "Source",
                "Description": "This workbook is computed directly from interpersonal_features_w2.csv/interpersonal_features_w3.csv and the W2/W3 cleaned student datasets, not from the combined Table 1 workbook.",
            },
        ]
    )
    all_df = pd.concat([w2.assign(Wave="W2 2024"), w3.assign(Wave="W3 2025")], ignore_index=True)
    all_df = all_df[["Wave"] + list(w2.columns)]

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="ReadMe", index=False)
        w2.to_excel(writer, sheet_name="W2_2024", index=False)
        w3.to_excel(writer, sheet_name="W3_2025", index=False)
        all_df.to_excel(writer, sheet_name="All", index=False)

    wb = load_workbook(OUT_PATH)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 58))
            ws.column_dimensions[get_column_letter(idx)].width = max(14, max_len + 2)
    wb.save(OUT_PATH)


def main() -> None:
    w2 = load_wave_table("W2", W2_DATA, W2_INTERPERSONAL)
    w3 = load_wave_table("W3", W3_DATA, W3_INTERPERSONAL)
    write_output(w2, w3)
    print(f"Wrote {OUT_PATH}")
    print(w2.to_string(index=False))
    print(w3.to_string(index=False))


if __name__ == "__main__":
    main()
